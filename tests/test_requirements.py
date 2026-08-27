from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

import backend_core
from backend_core import BackendStore
from maochao_rpa import Account, MaochaoRPA, RunResult, SupplierRef


class RequirementsTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        backend_core.DB_PATH = self.root / "backend" / "rpa.db"
        backend_core.WORK_DIR = self.root / "backend"
        backend_core.WORKER_HEARTBEAT_PATH = self.root / "backend" / "worker_heartbeat.json"
        config = {
            "login_url": "https://example.invalid/",
            "chrome_executable_path": "chrome",
            "data_root": str(self.root / "data"),
            "download_root": str(self.root / "downloads"),
            "log_dir": str(self.root / "logs"),
            "screenshot_dir": str(self.root / "screenshots"),
            "accounts_source": "json",
            "accounts_db_path": str(self.root / "accounts" / "accounts.db"),
            "accounts_db_key_path": str(self.root / "accounts" / ".key"),
            "accounts": [
                {
                    "key": "tmall_test_01", "name": "测试账号", "username": "tester", "password": "secret",
                    "port": 19231, "profile_dir": str(self.root / "profile"),
                    "download_dir": str(self.root / "downloads" / "tmall_test_01"),
                    "tasks": list(backend_core.TASKS), "enabled": True,
                }
            ],
            "direct_urls": {}, "selectors": {}, "cleanup": {},
        }
        self.config_path = self.root / "config.json"
        self.config_path.write_text(json.dumps(config, ensure_ascii=False), encoding="utf-8")
        self.store = BackendStore(self.config_path)
        self.store.account_store.upsert_account(config["accounts"][0], base_dir=self.root)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def test_supply_chain_operator_and_account_ownership(self) -> None:
        supply_a = self.store.create_supply_chain_user("supply-a", "供应链A", "pass-a")
        supply_b = self.store.create_supply_chain_user("supply-b", "供应链B", "pass-b")
        self.assertEqual(
            self.store.authenticate_supply_chain_user("supply-a", "pass-a")["user_id"],
            supply_a["user_id"],
        )
        operator = self.store.create_operator("运营甲", supply_a["user_id"])
        self.assertEqual(operator["supply_chain_user_id"], supply_a["user_id"])
        moved = self.store.update_operator(operator["operator_id"], supply_chain_user_id=supply_b["user_id"])
        self.assertEqual(moved["supply_chain_user_id"], supply_b["user_id"])
        self.store.set_account_owner("tmall_test_01", "supply_chain", supply_a["user_id"])
        self.assertEqual(self.store.get_account_owner("tmall_test_01")["creator_user_id"], supply_a["user_id"])

    def test_all_operator_schedule_is_dynamic(self) -> None:
        operator = self.store.create_operator("运营甲")
        schedule = self.store.create_schedule(
            task_keys=["channel-goods"], operator_ids=[], all_operators=True,
            enabled=True, time_of_day="09:00", weekdays=list(range(7)), headed=True,
        )
        self.assertTrue(schedule["all_operators"])
        self.assertEqual(schedule["operator_ids"], [])
        self.store.create_operator("运营乙")
        self.assertEqual(len([item for item in self.store.list_operators() if item["active"]]), 2)
        self.assertIsNotNone(operator)

    def test_claim_removes_lock_left_by_cancelled_run(self) -> None:
        operator = self.store.create_operator("运营甲")
        self.store.upsert_account_suppliers("tmall_test_01", [
            {"supplier_id": "supplier-1", "supplier_name": "供应商1"},
        ])
        self.store.set_operator_suppliers(operator["operator_id"], "tmall_test_01", ["supplier-1"])
        stale = self.store.create_run(
            ["channel-goods"], ["tmall_test_01"], True, True,
            operator_id=operator["operator_id"],
            suppliers=[{"account_key": "tmall_test_01", "supplier_id": "supplier-1", "supplier_name": "供应商1"}],
        )
        account = self.store.list_accounts()[0]
        self.store.cancel_pending_run(stale["run_id"])
        self.store.lock_accounts(stale["run_id"], [account])
        pending = self.store.create_run(
            ["channel-goods"], ["tmall_test_01"], True, False,
            operator_id=operator["operator_id"],
            suppliers=[{"account_key": "tmall_test_01", "supplier_id": "supplier-1", "supplier_name": "供应商1"}],
        )
        claimed = self.store.claim_next_pending_run()
        self.assertEqual(claimed["run_id"], pending["run_id"])

    def test_item_id_config_rejects_unknown_and_replaces_affected_supplier(self) -> None:
        self.store.upsert_account_suppliers("tmall_test_01", [
            {"supplier_id": "supplier-1", "supplier_name": "供应商1"},
        ])
        bad = self.store.replace_item_id_config(
            [{"account_key": "missing", "supplier_id": "supplier-1", "item_id": "100"}],
            original_name="bad.xlsx", uploaded_by_role="admin",
        )
        self.assertEqual(bad["status"], "rejected")
        first = self.store.replace_item_id_config(
            [
                {"account_key": "tmall_test_01", "supplier_id": "supplier-1", "item_id": "100"},
                {"account_key": "tmall_test_01", "supplier_id": "supplier-1", "item_id": "101"},
            ],
            original_name="first.xlsx", uploaded_by_role="admin",
        )
        self.assertEqual(first["row_count"], 2)
        second = self.store.replace_item_id_config(
            [{"account_key": "tmall_test_01", "supplier_id": "supplier-1", "item_id": "200"}],
            original_name="second.xlsx", uploaded_by_role="supply_chain", uploaded_by_user_id="supply-a",
        )
        self.assertEqual(second["row_count"], 1)
        self.assertEqual(self.store.list_item_ids("tmall_test_01", "supplier-1"), ["200"])
        rolled_back = self.store.rollback_item_id_config(first["upload_id"])
        self.assertEqual(rolled_back["row_count"], 2)
        self.assertEqual(self.store.list_item_ids("tmall_test_01", "supplier-1"), ["100", "101"])
        uploads = {item["upload_id"]: item for item in self.store.list_item_id_uploads()}
        self.assertEqual(uploads[first["upload_id"]]["status"], "active")
        self.assertEqual(uploads[second["upload_id"]]["status"], "superseded")
        cleared = self.store.replace_item_id_config(
            [{"account_key": "tmall_test_01", "supplier_id": "supplier-1", "item_id": ""}],
            original_name="clear.xlsx", uploaded_by_role="supply_chain",
        )
        self.assertEqual(cleared["row_count"], 0)
        self.assertEqual(self.store.list_item_ids("tmall_test_01", "supplier-1"), [])
        self.store.rollback_item_id_config(first["upload_id"])
        self.assertEqual(self.store.list_item_ids("tmall_test_01", "supplier-1"), ["100", "101"])

    def test_channel_goods_batch_merge_keeps_one_header_and_deduplicates(self) -> None:
        from openpyxl import Workbook, load_workbook

        cleaned = self.root / "data" / "operator" / "20260827" / "supplier-1" / "cleaned"
        cleaned.mkdir(parents=True)
        paths = []
        for index, rows in enumerate(([("100", "A"), ("101", "B")], [("101", "B"), ("102", "C")]), start=1):
            path = cleaned / f"batch{index}.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "库位明细"
            sheet.append(["货品ID", "名称"])
            for row in rows:
                sheet.append(row)
            workbook.save(path)
            paths.append(path)
        rpa = object.__new__(MaochaoRPA)
        rpa.settings = SimpleNamespace(data_root=self.root / "data")
        rpa._active_operator_name = "operator"
        rpa._current_supplier = SupplierRef("supplier-1", "供应商1", "tmall_test_01")
        account = Account(
            key="tmall_test_01", name="测试账号", username="", password="", port=19231,
            profile_dir=self.root / "profile", download_dir=self.root / "downloads",
            supplier_names=[], tasks=[], note="", xpath_vars={}, selector_overrides={}, enabled=True,
        )
        results = [
            RunResult(task="channel-goods", title="库位明细", account=account.key, status="ok", cleaned_file=str(path))
            for path in paths
        ]
        merged = rpa._merge_channel_goods_results(results, account)
        self.assertEqual(Path(merged.cleaned_file).parent.name, "库位明细")
        self.assertEqual(Path(merged.cleaned_file).name, "供应商1_库位明细.xlsx")
        workbook = load_workbook(merged.cleaned_file, read_only=True, data_only=True)
        values = list(workbook["库位明细"].values)
        workbook.close()
        self.assertEqual(values[0], ("货品ID", "名称"))
        self.assertEqual(values[1:], [("100", "A"), ("101", "B"), ("102", "C")])
        self.assertTrue(all(not path.exists() for path in paths))

    def test_shared_folder_and_file_prefix_use_supplier_name(self) -> None:
        rpa = object.__new__(MaochaoRPA)
        rpa.settings = SimpleNamespace(data_root=self.root / "data")
        rpa._active_operator_name = "静吟"
        rpa._current_supplier = SupplierRef("115468372", "广州七邦科技集团有限公司-口腔-寄售", "tmall_test_01")
        account = Account(
            key="tmall_test_01", name="测试账号", username="", password="", port=19231,
            profile_dir=self.root / "profile", download_dir=self.root / "downloads",
            supplier_names=[], tasks=[], note="", xpath_vars={}, selector_overrides={}, enabled=True,
        )
        raw_dir, cleaned_dir = rpa._account_data_dirs(account, "po-list")
        self.assertIn("广州七邦科技集团有限公司-口腔-寄售", raw_dir.parts)
        self.assertNotIn("115468372", raw_dir.parts + cleaned_dir.parts)
        self.assertTrue(rpa._supplier_prefix().startswith("广州七邦科技"))
        self.assertIn("_raw_archive", raw_dir.parts)
        self.assertEqual(cleaned_dir.name, "补货单列表")
        self.assertNotIn("cleaned", cleaned_dir.parts)


if __name__ == "__main__":
    unittest.main()
