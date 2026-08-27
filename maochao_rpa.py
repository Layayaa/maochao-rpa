# -*- coding: utf-8 -*-
"""
猫超补货数据源下载 RPA
====================
覆盖需求文档中本次纳入的 6 个数据源：
  1、实时库存
  2、库存分析 - 品仓明细表
  3、系统单
  4、补货单列表
  10、库位明细
  11、调拨单

设计约束：
  - 一个账号对应一个独立 Chrome remote debugging port 和 user-data-dir；
  - 初期顺序执行，不并发；
  - 账号放在加密 SQLite 安全库中，XPath 留在 JSON 配置并支持账号级覆盖；
  - 下载文件归档到共享盘/<人员>/YYYYMMDD/<供应商>/raw，并清洗到 cleaned。
"""

from __future__ import annotations

import argparse
import base64
import csv
from difflib import SequenceMatcher
import json
import os
import re
import shutil
import socket
import subprocess
import sys
import threading
import time
import traceback
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Iterable

from account_store import AccountStore, account_context, deep_merge, render_tree, unresolved_templates


BASE = Path(__file__).resolve().parent
DEFAULT_CONFIG = BASE / "config.example.json"
DEFAULT_SHARED_DATA_ROOT = r"\\172.17.17.3\公司共享文件夹\第一事业部\阿滨组\供应链"


@dataclass
class Account:
    key: str
    name: str
    username: str
    password: str
    port: int
    profile_dir: Path
    download_dir: Path
    supplier_names: list[str]
    tasks: list[str]
    note: str = ""
    xpath_vars: dict[str, Any] = field(default_factory=dict)
    selector_overrides: dict[str, Any] = field(default_factory=dict)
    enabled: bool = True


@dataclass
class Settings:
    config_path: Path
    accounts_source: str
    accounts_db_path: Path
    accounts_db_key_path: Path
    login_url: str
    chrome_executable_path: str
    data_root: Path
    log_dir: Path
    screenshot_dir: Path
    headless: bool
    download_timeout_sec: int
    task_timeout_sec: int
    poll_interval_sec: float
    accounts: list[Account]
    direct_urls: dict[str, str]
    selectors: dict[str, Any]
    cleanup: dict[str, Any]


@dataclass
class SupplierRef:
    supplier_id: str
    supplier_name: str
    account_key: str = ""


@dataclass
class RunResult:
    task: str
    title: str
    account: str
    status: str
    raw_file: str = ""
    cleaned_file: str = ""
    screenshot: str = ""
    error: str = ""
    started_at: str = ""
    finished_at: str = ""
    note: str = ""
    supplier_id: str = ""
    supplier_name: str = ""


TASKS: dict[str, dict[str, str]] = {
    "realtime-inventory": {
        "title": "1、实时库存",
        "file_task_text": "导出 实时库存",
        "prefix": "01_realtime_inventory",
    },
    "pincang-detail": {
        "title": "2、库存分析-品仓明细表",
        "file_task_text": "导出 品仓明细表",
        "prefix": "02_pincang_detail",
    },
    "system-order": {
        "title": "3、系统单",
        "file_task_text": "导出 PO明细确认分页导出",
        "prefix": "03_system_order",
    },
    "po-list": {
        "title": "4、补货单列表",
        "file_task_text": "PO明细分页导出",
        "prefix": "04_po_list",
    },
    "channel-goods": {
        "title": "10、库位明细",
        "file_task_text": "货品生命周期导出结果",
        "prefix": "10_channel_goods",
    },
    "transfer-order": {
        "title": "11、调拨单",
        "file_task_text": "导出 调拨单货品明细",
        "prefix": "11_transfer_order",
    },
}

TASK_ALIASES = {
    "1": "realtime-inventory",
    "实时库存": "realtime-inventory",
    "2": "pincang-detail",
    "品仓明细": "pincang-detail",
    "3": "system-order",
    "系统单": "system-order",
    "4": "po-list",
    "补货单": "po-list",
    "10": "channel-goods",
    "库位明细": "channel-goods",
    "11": "transfer-order",
    "调拨单": "transfer-order",
}

TASK_FRAME_HINTS = {
    "realtime-inventory": "inventory_realtime_search",
    "pincang-detail": "ai_tj_inventory_3",
    "system-order": "purchase_order_list_v4",
    "po-list": "purchase_order_list_v4",
    "channel-goods": "merchandise_channel_store",
    "transfer-order": "purchase_transfer_order_list_v4",
}

REQUIRED_SELECTORS: dict[str, tuple[str, ...]] = {
    "realtime-inventory": (
        "realtime.menu_inventory",
        "realtime.menu_inventory_query",
        "realtime.supplier_field",
        "realtime.query_button",
        "realtime.export_button",
        "realtime.export_all_option",
    ),
    "pincang-detail": (
        "pincang.menu_tianji",
        "pincang.menu_inventory_analysis",
        "pincang.tab_pincang_detail",
        "pincang.export_button",
    ),
    "system-order": (
        "purchase.menu_purchase",
        "purchase.menu_replenishment_order",
        "purchase.po_status_field",
        "purchase.query_button",
        "system_order.import_button",
        "system_order.import_confirm_option",
        "system_order.dialog_export_data",
    ),
    "po-list": (
        "purchase.menu_purchase",
        "purchase.menu_replenishment_order",
        "po_list.more_button",
        "po_list.start_date_input",
        "po_list.end_date_input",
        "po_list.date_confirm_button",
        "purchase.po_status_field",
        "purchase.query_button",
        "po_list.export_button",
    ),
    "channel-goods": (
        "channel_goods.menu_goods",
        "channel_goods.menu_channel_goods",
        "channel_goods.filter_button",
        "channel_goods.export_button",
    ),
    "transfer-order": (
        "purchase.menu_purchase",
        "transfer_order.menu_transfer_order",
        "transfer_order.export_button",
        "transfer_order.export_goods_detail_option",
    ),
}


def _require_playwright():
    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError(
            "缺少 Playwright 依赖。请先运行：\n"
            "  python3 -m pip install -r requirements.txt\n"
            "或使用 WorkBuddy Python 执行同样命令。"
        ) from exc
    return sync_playwright, PlaywrightTimeoutError


def _resolve_path(raw: str | None, base: Path, default: str) -> Path:
    value = (raw or default).strip()
    if value.startswith(("\\\\", "//")):
        return Path(value)
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = base / path
    return path


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        text = text[1:-1].strip()
    return text


PLACEHOLDER_SUPPLIER_NAMES = frozenset(
    {"全部", "请选择", "无数据", "暂无数据", "所有", "全部供应商", "所有供应商"}
)


def is_placeholder_supplier(supplier_id: str = "", supplier_name: str = "") -> bool:
    """右上角「全部」等占位项不是真实供应商，不能同步、勾选或执行。"""
    name = _clean_text(supplier_name)
    sid = _clean_text(supplier_id)
    if name in PLACEHOLDER_SUPPLIER_NAMES or sid in PLACEHOLDER_SUPPLIER_NAMES:
        return True
    if sid.startswith("name:") and sid[5:] in PLACEHOLDER_SUPPLIER_NAMES:
        return True
    return False


def _as_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [_clean_text(item) for item in value if _clean_text(item)]
    if isinstance(value, str):
        return [_clean_text(item) for item in re.split(r"[,，\s]+", value) if _clean_text(item)]
    return [_clean_text(value)] if _clean_text(value) else []


def _as_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    return {}


def _to_bool(value: Any, default: bool = True) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    text = _clean_text(value).lower()
    if text in {"1", "true", "yes", "y", "on", "是"}:
        return True
    if text in {"0", "false", "no", "n", "off", "否"}:
        return False
    return default


def _slug(value: str) -> str:
    value = re.sub(r"\s+", "_", value.strip())
    value = re.sub(r'[\\/:*?"<>|]+', "_", value)
    return value or "item"


def _months_ago(base: date, months: int) -> date:
    year = base.year
    month = base.month - months
    while month <= 0:
        month += 12
        year -= 1
    month_lengths = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28,
                     31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day = min(base.day, month_lengths[month - 1])
    return date(year, month, day)


def _xpath_literal(value: str) -> str:
    if "'" not in value:
        return f"'{value}'"
    if '"' not in value:
        return f'"{value}"'
    parts = value.split("'")
    return "concat(" + ', "\'", '.join(f"'{part}'" for part in parts) + ")"


def _pw_selector(selector: str) -> str:
    selector = selector.strip()
    if selector.startswith("xpath="):
        return selector
    if selector.startswith("/") or selector.startswith("("):
        return f"xpath={selector}"
    return selector


def load_settings(config_path: Path) -> Settings:
    config_path = config_path.expanduser().resolve()
    with config_path.open("r", encoding="utf-8") as f:
        raw = json.load(f)
    base = config_path.parent

    default_download_root = _resolve_path(raw.get("download_root"), base, "./downloads")
    configured_data_root = _clean_text(raw.get("data_root"))
    if configured_data_root in {"", ".", "./data", "data"}:
        configured_data_root = DEFAULT_SHARED_DATA_ROOT
    accounts_source = (_clean_text(raw.get("accounts_source", "db")) or "db").lower()
    accounts_db_path = _resolve_path(raw.get("accounts_db_path"), base, "./accounts/maochao_accounts.db")
    accounts_db_key_path = _resolve_path(raw.get("accounts_db_key_path"), base, "./accounts/.secret_key")

    if accounts_source == "db":
        store = AccountStore(accounts_db_path, accounts_db_key_path)
        try:
            account_items = store.list_accounts()
        except FileNotFoundError as exc:
            raise RuntimeError(
                "账号库还没有初始化。请先运行：\n"
                f"  python account_store.py --db \"{accounts_db_path}\" --key \"{accounts_db_key_path}\" init\n"
                f"  python account_store.py --db \"{accounts_db_path}\" --key \"{accounts_db_key_path}\" import-json \"{config_path}\""
            ) from exc
    elif accounts_source == "json":
        account_items = raw.get("accounts", [])
    else:
        raise ValueError("accounts_source 只支持 db 或 json")

    accounts = [
        _account_from_mapping(item, base, default_download_root)
        for item in account_items
        if isinstance(item, dict)
        and _clean_text(item.get("key"))
        and _to_bool(item.get("enabled"), default=True)
    ]
    if accounts_source == "db" and not accounts:
        raise RuntimeError("账号库中没有启用账号，请先导入或启用账号记录。")

    return Settings(
        config_path=config_path,
        accounts_source=accounts_source,
        accounts_db_path=accounts_db_path,
        accounts_db_key_path=accounts_db_key_path,
        login_url=_clean_text(raw.get("login_url")),
        chrome_executable_path=_clean_text(raw.get("chrome_executable_path")),
        data_root=_resolve_path(configured_data_root, base, DEFAULT_SHARED_DATA_ROOT),
        log_dir=_resolve_path(raw.get("log_dir"), base, "./logs"),
        screenshot_dir=_resolve_path(raw.get("screenshot_dir"), base, "./logs/screenshots"),
        headless=bool(raw.get("headless", False)),
        download_timeout_sec=int(raw.get("download_timeout_sec") or 300),
        task_timeout_sec=int(raw.get("task_timeout_sec") or 600),
        poll_interval_sec=float(raw.get("poll_interval_sec") or 2),
        accounts=accounts,
        direct_urls={k: _clean_text(v) for k, v in raw.get("direct_urls", {}).items()},
        selectors=raw.get("selectors", {}),
        cleanup=raw.get("cleanup", {}),
    )


def _account_from_mapping(item: dict[str, Any], base: Path, default_download_root: Path) -> Account:
    key = _clean_text(item.get("key"))
    account_download_dir = _resolve_path(item.get("download_dir"), base, f"./downloads/{key}")
    return Account(
        key=key,
        name=_clean_text(item.get("name")) or key,
        username=_clean_text(item.get("username")),
        password=_clean_text(item.get("password")),
        port=int(item.get("port") or 0),
        profile_dir=_resolve_path(item.get("profile_dir"), base, f"./browser_profiles/{key}"),
        download_dir=account_download_dir if item.get("download_dir") else default_download_root / key,
        supplier_names=_as_list(item.get("supplier_names")),
        tasks=[normalize_task_name(v) for v in _as_list(item.get("tasks"))],
        note=_clean_text(item.get("note")),
        xpath_vars=_as_dict(item.get("xpath_vars")),
        selector_overrides=_as_dict(item.get("selector_overrides")),
        enabled=_to_bool(item.get("enabled"), default=True),
    )


def normalize_task_name(value: str) -> str:
    text = _clean_text(value)
    if not text:
        return text
    return TASK_ALIASES.get(text, text)


def selected_tasks(task_names: Iterable[str] | None) -> list[str]:
    values = [normalize_task_name(v) for v in (task_names or []) if _clean_text(v)]
    if not values or "all" in values:
        return list(TASKS)
    unknown = [v for v in values if v not in TASKS]
    if unknown:
        raise KeyError(f"未知猫超任务: {', '.join(unknown)}")
    return values


class MaochaoRPA:
    def __init__(
        self,
        settings: Settings,
        manual_login: bool = False,
        headless: bool | None = None,
    ):
        self.settings = settings
        self.manual_login = manual_login
        self.headless = settings.headless if headless is None else headless
        self.sync_playwright, self.PlaywrightTimeoutError = _require_playwright()
        self._active_account: Account | None = None
        self._active_selectors: dict[str, Any] = settings.selectors
        self.last_run_paused = False
        self._current_supplier: SupplierRef | None = None
        self._active_operator_name = ""
        self._item_ids_by_supplier: dict[tuple[str, str], list[str]] = {}
        self._handlers: dict[str, Callable[[Any, Account], list[RunResult]]] = {
            "realtime-inventory": self._task_realtime_inventory,
            "pincang-detail": self._task_pincang_detail,
            "system-order": self._task_system_order,
            "po-list": self._task_po_list,
            "channel-goods": self._task_channel_goods,
            "transfer-order": self._task_transfer_order,
        }

    def run(
        self,
        tasks: list[str],
        account_keys: list[str] | None = None,
        force_account_tasks: bool = False,
        should_pause: Callable[[], bool] | None = None,
        skip_completed: set[tuple[str, str, str]] | None = None,
        suppliers: list[dict[str, Any]] | SupplierRef | None = None,
        use_current_supplier: bool = False,
        operator_name: str = "",
        item_ids_by_supplier: dict[tuple[str, str], list[str]] | None = None,
    ) -> list[RunResult]:
        self._active_operator_name = _clean_text(operator_name)
        self._item_ids_by_supplier = item_ids_by_supplier or {}
        self._ensure_dirs()
        selected_accounts = self._selected_accounts(account_keys)
        results: list[RunResult] = []
        self.last_run_paused = False
        skip_completed = skip_completed or set()
        requested_suppliers = self._normalize_supplier_refs(suppliers)

        with self.sync_playwright() as p:
            for account in selected_accounts:
                # 新口径：每个运营已分配供应商都要把本次勾选的任务跑完。
                # 默认是任务 1-6；不再按账号库里的旧 tasks 拆成库存号/通用号。
                if requested_suppliers or use_current_supplier or force_account_tasks:
                    account_tasks = list(tasks)
                else:
                    account_tasks = [task for task in tasks if task in account.tasks]
                if not account_tasks:
                    continue

                if should_pause and should_pause():
                    self.last_run_paused = True
                    break

                print(f"[猫超] 接管账号: {account.name} ({account.key})")
                self._ensure_account_dirs(account)
                self._ensure_chrome(account)
                account_started = datetime.now().isoformat(timespec="seconds")
                browser = None
                page = None
                self._current_supplier = None
                try:
                    browser = self._connect_over_cdp(p, account)
                    context = browser.contexts[0] if browser.contexts else browser.new_context()
                    page = self._prefer_business_page(context)
                    self._set_download_behavior(context, page, account.download_dir)
                    self._ensure_browser_window(page)
                    self._set_active_account(account)
                    self._login_or_reuse_session(page, account)

                    account_suppliers = [
                        item for item in requested_suppliers
                        if not item.account_key or item.account_key == account.key
                    ]
                    if use_current_supplier and not account_suppliers:
                        current = self._current_header_supplier(page)
                        if current is None or is_placeholder_supplier(current.supplier_id, current.supplier_name):
                            raise RuntimeError("无法读取右上角当前真实供应商。请先同步供应商清单，不要停在「全部」。")
                        account_suppliers = [current]
                    if not account_suppliers:
                        raise RuntimeError("未选择运营已分配的供应商。请先在 Web 中为运营勾选负责供应商，再执行任务。")

                    for supplier in account_suppliers:
                        if should_pause and should_pause():
                            self.last_run_paused = True
                            break
                        if is_placeholder_supplier(supplier.supplier_id, supplier.supplier_name):
                            results.append(
                                RunResult(
                                    task="__supplier__",
                                    title=f"切换供应商 {supplier.supplier_name or supplier.supplier_id}",
                                    account=account.key,
                                    status="failed",
                                    error="「全部」不是真实供应商，不能作为本次执行对象。请重新同步右上角并勾选具体供应商。",
                                    started_at=datetime.now().isoformat(timespec="seconds"),
                                    finished_at=datetime.now().isoformat(timespec="seconds"),
                                    supplier_id=supplier.supplier_id,
                                    supplier_name=supplier.supplier_name,
                                )
                            )
                            continue
                        remaining_tasks = [
                            task for task in account_tasks
                            if (account.key, supplier.supplier_id, task) not in skip_completed
                        ]
                        if not remaining_tasks:
                            continue
                        supplier_started = datetime.now().isoformat(timespec="seconds")
                        try:
                            print(
                                f"[猫超] 切换供应商: {supplier.supplier_name or supplier.supplier_id} "
                                f"id={supplier.supplier_id}，将执行 {len(remaining_tasks)} 个任务"
                            )
                            self._switch_header_supplier(page, supplier)
                            self._current_supplier = supplier
                            for task_key in remaining_tasks:
                                if should_pause and should_pause():
                                    self.last_run_paused = True
                                    break
                                started = datetime.now().isoformat(timespec="seconds")
                                try:
                                    print(
                                        f"[猫超] 开始: {supplier.supplier_name or supplier.supplier_id} / "
                                        f"{TASKS[task_key]['title']} / {account.name}"
                                    )
                                    self._ensure_browser_window(page)
                                    task_results = self._stamp_supplier(self._handlers[task_key](page, account), supplier)
                                    results.extend(task_results)
                                except Exception as exc:
                                    shot = self._screenshot(page, f"{task_key}_{account.key}_{_slug(supplier.supplier_id)}_failed")
                                    results.append(
                                        RunResult(
                                            task=task_key,
                                            title=TASKS[task_key]["title"],
                                            account=account.key,
                                            status="failed",
                                            screenshot=str(shot),
                                            error=f"{exc}\n{traceback.format_exc(limit=4)}",
                                            started_at=started,
                                            finished_at=datetime.now().isoformat(timespec="seconds"),
                                            supplier_id=supplier.supplier_id,
                                            supplier_name=supplier.supplier_name,
                                        )
                                    )
                                    print(f"[猫超] 失败: {TASKS[task_key]['title']} / {supplier.supplier_name} -> {exc}")
                        except Exception as exc:
                            shot = self._capture_error_screenshot(page, f"{account.key}_{_slug(supplier.supplier_id)}_switch_error")
                            results.append(
                                RunResult(
                                    task="__supplier__",
                                    title=f"切换供应商 {supplier.supplier_name or supplier.supplier_id}",
                                    account=account.key,
                                    status="failed",
                                    screenshot=shot,
                                    error=f"{exc}\n{traceback.format_exc(limit=4)}",
                                    started_at=supplier_started,
                                    finished_at=datetime.now().isoformat(timespec="seconds"),
                                    supplier_id=supplier.supplier_id,
                                    supplier_name=supplier.supplier_name,
                                )
                            )
                            print(f"[猫超] 供应商切换失败: {supplier.supplier_name or supplier.supplier_id} -> {exc}")
                    if self.last_run_paused:
                        break
                except Exception as exc:
                    shot = self._capture_error_screenshot(page, f"{account.key}_account_error")
                    results.append(
                        RunResult(
                            task="__account__",
                            title=f"{account.name} 账户阶段",
                            account=account.key,
                            status="failed",
                            screenshot=shot,
                            error=f"{exc}\n{traceback.format_exc(limit=4)}",
                            started_at=account_started,
                            finished_at=datetime.now().isoformat(timespec="seconds"),
                        )
                    )
                    print(f"[猫超] 账户阶段失败: {account.name} -> {exc}")
                finally:
                    self._current_supplier = None
                    # 这里不主动 close 已接管的浏览器，避免把账号资料目录里的登录态一并关掉。
                    pass

        self._write_manifest(results)
        return results

    def sync_header_suppliers(self, account_keys: list[str] | None = None) -> list[dict[str, Any]]:
        self._ensure_dirs()
        selected_accounts = self._selected_accounts(account_keys)
        payload: list[dict[str, Any]] = []
        with self.sync_playwright() as p:
            for account in selected_accounts:
                print(f"[猫超] 同步供应商: {account.name} ({account.key})")
                self._ensure_account_dirs(account)
                self._ensure_chrome(account)
                browser = None
                page = None
                try:
                    browser = self._connect_over_cdp(p, account)
                    context = browser.contexts[0] if browser.contexts else browser.new_context()
                    page = self._prefer_business_page(context)
                    self._set_download_behavior(context, page, account.download_dir)
                    self._ensure_browser_window(page)
                    self._set_active_account(account)
                    suppliers = self._login_or_reuse_session(page, account, harvest_second_suppliers=True)
                    if not suppliers:
                        suppliers = self._discover_account_suppliers(page)
                    print(f"[猫超] 账号可见供应商 {len(suppliers)} 个")
                    for item in suppliers:
                        print(f"[猫超]   {item.supplier_id} {item.supplier_name}")
                        payload.append(
                            {
                                "account_key": account.key,
                                "supplier_id": item.supplier_id,
                                "supplier_name": item.supplier_name,
                            }
                        )
                except Exception:
                    shot = self._capture_error_screenshot(page, f"{account.key}_sync_suppliers_error")
                    if shot:
                        print(f"[猫超] 已保存供应商同步错误截图: {shot}")
                    raise
                finally:
                    pass
        return payload

    def login_only(self, account_keys: list[str] | None = None) -> None:
        self._ensure_dirs()
        selected_accounts = self._selected_accounts(account_keys)
        with self.sync_playwright() as p:
            for account in selected_accounts:
                print(f"[猫超] 打开账号浏览器用于首次登录: {account.name} ({account.key})")
                self._ensure_account_dirs(account)
                self._ensure_chrome(account)
                browser = None
                page = None
                try:
                    browser = self._connect_over_cdp(p, account)
                    context = browser.contexts[0] if browser.contexts else browser.new_context()
                    page = self._prefer_business_page(context)
                    self._set_download_behavior(context, page, account.download_dir)
                    self._ensure_browser_window(page)
                    self._set_active_account(account)
                    self._login_or_reuse_session(page, account, allow_login_navigation=True)
                    input()
                except Exception as exc:
                    shot = self._capture_error_screenshot(page, f"{account.key}_login_error")
                    print(f"[猫超] 首次登录失败: {account.name} -> {exc}")
                    if shot:
                        print(f"[猫超] 已保存登录错误截图: {shot}")
                    raise
                finally:
                    # 同样不主动关闭，保持 profile 与登录态。
                    pass

    def _selected_accounts(self, account_keys: list[str] | None) -> list[Account]:
        if not account_keys:
            return self.settings.accounts
        wanted = set(account_keys)
        accounts = [account for account in self.settings.accounts if account.key in wanted]
        missing = sorted(wanted - {account.key for account in accounts})
        if missing:
            raise KeyError(f"配置中找不到账号: {', '.join(missing)}")
        return accounts

    def _ensure_dirs(self) -> None:
        try:
            self.settings.data_root.mkdir(parents=True, exist_ok=True)
        except OSError:
            if not str(self.settings.data_root).startswith(("\\", "//")):
                raise
        self.settings.log_dir.mkdir(parents=True, exist_ok=True)
        self.settings.screenshot_dir.mkdir(parents=True, exist_ok=True)

    def _ensure_account_dirs(self, account: Account) -> None:
        account.profile_dir.mkdir(parents=True, exist_ok=True)
        account.download_dir.mkdir(parents=True, exist_ok=True)
        for path in self._account_data_dirs(account):
            try:
                path.mkdir(parents=True, exist_ok=True)
            except OSError:
                if not str(path).startswith(("\\", "//")):
                    raise

    def _account_data_dirs(self, account: Account) -> tuple[Path, Path]:
        today = date.today().strftime("%Y%m%d")
        person = _slug(self._active_operator_name or account.name or account.key)
        root = self.settings.data_root / person / today
        supplier = self._current_supplier
        if supplier is not None:
            supplier_slug = _slug(supplier.supplier_name or supplier.supplier_id)
            if supplier_slug:
                root = root / supplier_slug
        return root / "raw", root / "cleaned"

    def _normalize_supplier_refs(self, suppliers: list[dict[str, Any]] | SupplierRef | None) -> list[SupplierRef]:
        if suppliers is None:
            return []
        if isinstance(suppliers, SupplierRef):
            return [] if is_placeholder_supplier(suppliers.supplier_id, suppliers.supplier_name) else [suppliers]
        rows: list[SupplierRef] = []
        seen: set[tuple[str, str]] = set()
        for item in suppliers:
            if isinstance(item, SupplierRef):
                supplier = item
            elif isinstance(item, dict):
                supplier_id = _clean_text(item.get("supplier_id"))
                supplier_name = _clean_text(item.get("supplier_name"))
                if not supplier_id and not supplier_name:
                    continue
                if not supplier_id:
                    supplier_id = f"name:{supplier_name}"
                if is_placeholder_supplier(supplier_id, supplier_name):
                    continue
                supplier = SupplierRef(
                    supplier_id=supplier_id,
                    supplier_name=supplier_name,
                    account_key=_clean_text(item.get("account_key")),
                )
            else:
                continue
            if is_placeholder_supplier(supplier.supplier_id, supplier.supplier_name):
                continue
            key = (supplier.account_key, supplier.supplier_id)
            if key in seen:
                continue
            seen.add(key)
            rows.append(supplier)
        return rows

    def _stamp_supplier(self, results: list[RunResult], supplier: SupplierRef | None = None) -> list[RunResult]:
        current = supplier or self._current_supplier
        if current is None:
            return results
        for item in results:
            if not item.supplier_id:
                item.supplier_id = current.supplier_id
            if not item.supplier_name:
                item.supplier_name = current.supplier_name
        return results

    def _supplier_prefix(self) -> str:
        if self._current_supplier is None:
            return ""
        return _slug(self._current_supplier.supplier_name or self._current_supplier.supplier_id)

    def _set_active_account(self, account: Account) -> None:
        self._active_account = account
        context = account_context(
            {
                "key": account.key,
                "name": account.name,
                "username": account.username,
                "password": account.password,
                "port": account.port,
                "profile_dir": str(account.profile_dir),
                "download_dir": str(account.download_dir),
                "supplier_names": account.supplier_names,
                "tasks": account.tasks,
                "note": account.note,
                "xpath_vars": account.xpath_vars,
                "selector_overrides": account.selector_overrides,
                "enabled": account.enabled,
            }
        )
        rendered = render_tree(self.settings.selectors, context)
        overrides = render_tree(account.selector_overrides, context)
        self._active_selectors = deep_merge(rendered, overrides)
        unresolved = unresolved_templates(self._active_selectors)
        if unresolved:
            print(f"[猫超] 提醒：账号 {account.key} 的部分 XPath 变量未展开: {unresolved}")

    def _ensure_chrome(self, account: Account) -> None:
        if account.port <= 0:
            raise RuntimeError(f"账号 {account.key} 缺少有效 browser port")
        if self._port_open(account.port):
            print(f"[猫超] 端口 {account.port} 已打开，直接接管。")
            return
        chrome_path = self.settings.chrome_executable_path
        if not chrome_path or not Path(chrome_path).exists():
            raise RuntimeError(f"找不到 Chrome: {chrome_path or '未配置'}")

        chrome_args = [
            chrome_path,
            "--remote-debugging-address=127.0.0.1",
            f"--remote-debugging-port={account.port}",
            f"--user-data-dir={account.profile_dir}",
            "--no-first-run",
            "--no-default-browser-check",
            "--disable-popup-blocking",
            "about:blank",
        ]
        if self.headless:
            chrome_args.insert(-1, "--headless=new")
            cmd = chrome_args
        else:
            app_bundle = self._chrome_app_bundle(Path(chrome_path))
            cmd = ["open", "-na", str(app_bundle), "--args", *chrome_args[1:]] if app_bundle else chrome_args
        print(f"[猫超] 启动 Chrome: {account.name} / port={account.port}")
        subprocess.Popen(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        deadline = time.time() + 60
        while time.time() < deadline:
            if self._port_open(account.port):
                return
            time.sleep(0.3)
        raise RuntimeError(f"Chrome 启动后端口未打开: {account.port}")

    def _connect_over_cdp(self, playwright: Any, account: Account) -> Any:
        url = f"http://127.0.0.1:{account.port}"
        print(f"[猫超] 连接 Chrome CDP: port={account.port}")
        browser = playwright.chromium.connect_over_cdp(
            url,
            no_defaults=True,
            is_local=True,
            timeout=15000,
        )
        print(f"[猫超] 已连接 Chrome CDP: port={account.port}")
        return browser

    def _chrome_app_bundle(self, chrome_path: Path) -> Path | None:
        for parent in chrome_path.parents:
            if parent.suffix == ".app":
                return parent
        return None

    def _port_open(self, port: int) -> bool:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.settimeout(0.3)
            return sock.connect_ex(("127.0.0.1", port)) == 0

    def _set_download_behavior(self, context: Any, page: Any, download_dir: Path) -> None:
        download_dir.mkdir(parents=True, exist_ok=True)
        try:
            session = context.new_cdp_session(page)
            try:
                session.send(
                    "Browser.setDownloadBehavior",
                    {"behavior": "allow", "downloadPath": str(download_dir), "eventsEnabled": True},
                )
            except Exception:
                session.send("Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": str(download_dir)})
        except Exception as exc:
            print(f"[猫超] 下载目录 CDP 设置失败，将使用浏览器默认目录: {exc}")

    def _ensure_browser_window(self, page: Any) -> None:
        try:
            session = page.context.new_cdp_session(page)
            info = session.send("Browser.getWindowForTarget")
            window_id = info["windowId"]
            bounds = info.get("bounds") or {}
            state = str(bounds.get("windowState") or "normal")
            width = int(bounds.get("width") or 0)
            height = int(bounds.get("height") or 0)
            if state == "minimized":
                session.send(
                    "Browser.setWindowBounds",
                    {"windowId": window_id, "bounds": {"windowState": "normal"}},
                )
                print("[猫超] 已恢复最小化的 Chrome 窗口")
            if width < 1200 or height < 700 or state == "minimized":
                session.send(
                    "Browser.setWindowBounds",
                    {
                        "windowId": window_id,
                        "bounds": {
                            "windowState": "normal",
                            "width": 1440,
                            "height": 960,
                            "left": 40,
                            "top": 40,
                        },
                    },
                )
                print(f"[猫超] 已把 Chrome 窗口调到 1440x960（原 {width}x{height} {state}）")
            try:
                page.bring_to_front()
            except Exception:
                pass
        except Exception as exc:
            print(f"[猫超] 调整 Chrome 窗口失败: {exc}")

    def _visible_frame_url_contains(self, page: Any, needle: str) -> bool:
        if not needle:
            return False
        for src in self._visible_iframe_srcs(page):
            if needle in src:
                return True
        url = str(getattr(page, "url", "") or "")
        return needle in url

    def _wait_frame_url_contains(self, page: Any, needle: str, timeout_ms: int = 8000) -> bool:
        deadline = time.time() + timeout_ms / 1000
        while time.time() < deadline:
            if self._visible_frame_url_contains(page, needle):
                return True
            time.sleep(0.3)
        print(f"[猫超] 等待目标 iframe 超时: {needle}")
        return False

    def _wait_toolbar_title(self, page: Any, title: str, timeout_ms: int = 8000) -> bool:
        target = _clean_text(title)
        script = """
        (title) => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          return Array.from(document.querySelectorAll('.comp-toolbar-title-text'))
            .some((el) => textOf(el).includes(title));
        }
        """
        deadline = time.time() + timeout_ms / 1000
        while time.time() < deadline:
            for scope in self._iter_scopes(page):
                try:
                    if scope.evaluate(script, target):
                        return True
                except Exception:
                    continue
            time.sleep(0.3)
        return False

    def _login_or_reuse_session(
        self,
        page: Any,
        account: Account,
        harvest_second_suppliers: bool = False,
        allow_login_navigation: bool = False,
    ) -> list[SupplierRef]:
        if not self.settings.login_url:
            raise RuntimeError("config 缺少 login_url")

        if self._page_looks_logged_in(page):
            self._dismiss_blocking_popups(page)
            print("[猫超] 检测到现有登录态，跳过登录页。")
            return []

        if self._merchant_selector_visible(page):
            print("[猫超] 当前就在选择商家账号页，直接读取二级供应商。")
            harvested = self._handle_merchant_selector(page, harvest=harvest_second_suppliers)
            if not self._wait_business_home(page, 30000):
                raise RuntimeError("登录未完成：未进入商家主页，请检查验证码/滑块/登录态。")
            self._dismiss_blocking_popups(page)
            print("[猫超] 登录态可用，继续执行。")
            return harvested

        if self._wait_business_home(page, 3000):
            self._dismiss_blocking_popups(page)
            print("[猫超] 检测到现有登录态，跳过登录页。")
            return []

        current_url = str(getattr(page, "url", "") or "")
        login_page = bool(re.search(r"/login(?:[/?#]|$)", current_url))
        auth_callback = "oauth_sign" in current_url or "havanalogin.taobao.com" in current_url
        on_login = login_page or auth_callback
        if not on_login and "txcs.tmall.com" not in current_url:
            try:
                page.goto("https://web.txcs.tmall.com/", wait_until="domcontentloaded")
                self._wait_quiet(page, 8000)
                if self._wait_business_home(page, 8000):
                    self._dismiss_blocking_popups(page)
                    print("[猫超] 已复用现有登录态，跳过登录页。")
                    return []
            except Exception as exc:
                print(f"[猫超] 复用现有登录态失败，改走登录页: {exc}")
            current_url = str(getattr(page, "url", "") or "")
            login_page = bool(re.search(r"/login(?:[/?#]|$)", current_url))
            auth_callback = "oauth_sign" in current_url or "havanalogin.taobao.com" in current_url
            on_login = login_page or auth_callback
        if not on_login and not allow_login_navigation:
            print(f"[猫超] 未识别工作台登录态，且当前不是登录页，不打开登录、不填写密码: {current_url[:120]}")
            raise RuntimeError("未检测到商家工作台登录态。已中止，避免覆盖现有 Chrome 登录态。")

        if not login_page:
            page.goto(self.settings.login_url, wait_until="domcontentloaded")
            self._wait_quiet(page, 8000)

        login_scope = self._login_form_scope(page, timeout=5000)

        if login_scope is None and self._manual_login_wait_needed(page):
            raise RuntimeError(self._manual_login_required_message(page))

        if login_scope:
            if account.username and account.password:
                print(
                    f"[猫超] 检测到登录页，准备自动填写: {account.name} "
                    f"(账号长度={len(account.username)}, 密码长度={len(account.password)})"
                )
                self._scope_fill_any(login_scope, self._login_selector_candidates("login.username_input"), account.username, "账号")
                self._scope_fill_any(login_scope, self._login_selector_candidates("login.password_input"), account.password, "密码")
                self._verify_login_fields(login_scope, account.username, account.password)
                self._write_login_diagnostic(page, login_scope, account, "before_submit")
                self._wait_quiet(page, 1000)
                self._dismiss_blocking_popups(page)
                self._scope_click_any(login_scope, self._login_selector_candidates("login.login_button"), "登录", timeout=30000)
                try:
                    self._wait_login_transition(page, 60000)
                except Exception as exc:
                    if self.manual_login and sys.stdin.isatty():
                        print("[猫超] 自动登录未完成，请在当前浏览器完成登录后回到终端按 Enter。")
                        input()
                        self._wait_login_transition(page, 120000)
                    elif self._manual_login_wait_needed(page):
                        self._write_login_diagnostic(page, login_scope, account, "verification_required")
                        raise RuntimeError(self._manual_login_required_message(page, exc)) from exc
                    else:
                        self._write_login_diagnostic(page, login_scope, account, "submit_failed")
                        raise RuntimeError(self._login_failure_message(page, exc)) from exc
            else:
                print(f"[猫超] 检测到登录页，但账号 {account.key} 未配置账号密码。请人工登录后回车。")
                if not sys.stdin.isatty():
                    raise RuntimeError("登录页出现但 Worker 无法等待人工输入，已中止以免卡死。")
                input()

        if self.manual_login:
            if not sys.stdin.isatty():
                print("[猫超] Worker 无终端，跳过人工确认等待。")
            else:
                print("[猫超] 如有验证码/扫码/手机确认，请处理完成后回到终端按 Enter。")
                input()
                self._wait_quiet(page, 5000)

        harvested = self._handle_merchant_selector(page, harvest=harvest_second_suppliers)
        if not self._wait_business_home(page, 30000):
            raise RuntimeError(self._login_failure_message(page))
        self._dismiss_blocking_popups(page)
        print("[猫超] 登录态可用，继续执行。")
        return harvested

    def _wait_login_transition(self, page: Any, timeout_ms: int) -> None:
        deadline = time.time() + timeout_ms / 1000
        while time.time() < deadline:
            if self._frame_with_selectors(page, ("merchant.enter_button",), timeout=500):
                return
            if self._wait_business_home(page, 500):
                return
            time.sleep(1)
        raise RuntimeError("登录提交后未进入商家选择页或商家主页。")

    def _login_selector_candidates(self, selector_key: str) -> list[str]:
        configured = self._selector_optional(selector_key)
        fallbacks = {
            "login.username_input": [
                "#fm-login-id",
                "input[name='fm-login-id']",
                "input[placeholder*='账号']",
                "input[placeholder*='手机']",
                "input[placeholder*='邮箱']",
            ],
            "login.password_input": [
                "#fm-login-password",
                "input[name='fm-login-password']",
                "input[type='password']",
                "input[placeholder*='密码']",
            ],
            "login.login_button": [
                "button.fm-submit",
                "button.password-login",
                "button:has-text('登录')",
                "[role='button']:has-text('登录')",
            ],
        }.get(selector_key, [])
        candidates: list[str] = []
        for item in [*fallbacks, configured]:
            if item and item not in candidates:
                candidates.append(item)
        return candidates

    def _login_form_scope(self, page: Any, timeout: int = 5000) -> Any | None:
        groups = [
            self._login_selector_candidates("login.username_input"),
            self._login_selector_candidates("login.password_input"),
            self._login_selector_candidates("login.login_button"),
        ]
        deadline = time.time() + timeout / 1000
        while time.time() < deadline:
            visible_scopes: list[Any] = [page]
            fallback_scopes: list[Any] = []
            for frame in getattr(page, "frames", []) or []:
                try:
                    if frame is getattr(page, "main_frame", None):
                        continue
                    if self._frame_is_displayed(frame):
                        visible_scopes.append(frame)
                    else:
                        fallback_scopes.append(frame)
                except Exception:
                    continue
            matches: list[tuple[int, float, Any]] = []
            for scope in [*visible_scopes, *fallback_scopes]:
                try:
                    if all(self._first_visible_in_scope(scope, selector_group, timeout=200) is not None for selector_group in groups):
                        url = str(getattr(scope, "url", "") or "")
                        preferred = 0 if "havanalogin.taobao.com/mini_login.htm" in url else 1
                        top = 0.0
                        if scope is not page:
                            try:
                                box = scope.frame_element().bounding_box()
                                top = float((box or {}).get("y") or 0)
                            except Exception:
                                pass
                        matches.append((preferred, top, scope))
                except Exception:
                    continue
            if matches:
                matches.sort(key=lambda item: (item[0], item[1]))
                return matches[0][2]
            time.sleep(0.2)
        return None

    def _first_visible_in_scope(self, scope: Any, selectors: list[str], timeout: int = 1000) -> Any | None:
        for selector in selectors:
            locator = self._scoped_visible_locator(scope, selector, timeout=timeout)
            if locator is not None:
                return locator
        return None

    def _scope_fill_any(self, scope: Any, selectors: list[str], value: str, label: str, timeout: int = 10000) -> None:
        locator = self._first_visible_in_scope(scope, selectors, timeout=timeout)
        if locator is None:
            raise RuntimeError(f"找不到{label}: login selector candidates")
        try:
            locator.click(timeout=min(timeout, 2000))
            locator.fill(value, timeout=min(timeout, 3000))
            if locator.input_value(timeout=800) == value:
                print(f"[猫超] {label}填写验证成功: 长度={len(value)}")
                return
        except Exception:
            pass
        if self._set_input_value(locator, value):
            print(f"[猫超] {label}填写验证成功(JS): 长度={len(value)}")
            return
        raise RuntimeError(f"{label}填写后未生效: login selector candidates")

    def _login_field_state(self, scope: Any) -> dict[str, Any]:
        state: dict[str, Any] = {
            "frame_url": str(getattr(scope, "url", "") or ""),
            "username_present": False,
            "username_length": 0,
            "password_present": False,
            "password_length": 0,
            "login_button_present": False,
        }
        for key, present_key, length_key in (
            ("login.username_input", "username_present", "username_length"),
            ("login.password_input", "password_present", "password_length"),
        ):
            locator = self._first_visible_in_scope(
                scope,
                self._login_selector_candidates(key),
                timeout=800,
            )
            if locator is None:
                continue
            state[present_key] = True
            try:
                state[length_key] = len(locator.input_value(timeout=800) or "")
            except Exception:
                state[length_key] = -1
        state["login_button_present"] = (
            self._first_visible_in_scope(
                scope,
                self._login_selector_candidates("login.login_button"),
                timeout=800,
            )
            is not None
        )
        return state

    def _verify_login_fields(self, scope: Any, username: str, password: str) -> None:
        state = self._login_field_state(scope)
        if (
            not state["username_present"]
            or not state["password_present"]
            or state["username_length"] != len(username)
            or state["password_length"] != len(password)
        ):
            raise RuntimeError(
                "自动填写失败：登录表单未生效 "
                f"(账号长度={state['username_length']}, 密码长度={state['password_length']}, "
                f"账号框存在={state['username_present']}, 密码框存在={state['password_present']})"
            )
        print(
            "[猫超] 登录表单提交前验证通过: "
            f"账号长度={state['username_length']}, 密码长度={state['password_length']}, "
            f"登录按钮存在={state['login_button_present']}"
        )

    def _write_login_diagnostic(
        self,
        page: Any,
        scope: Any,
        account: Account,
        phase: str,
    ) -> None:
        try:
            payload = self._login_field_state(scope)
            payload["page_url"] = str(getattr(page, "url", "") or "")
            payload["account_key"] = account.key
            payload["phase"] = phase
            directory = self.settings.log_dir / "login_diagnostics"
            directory.mkdir(parents=True, exist_ok=True)
            path = directory / (
                f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{_slug(account.key)}_{phase}.json"
            )
            path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"[猫超] 已写入脱敏登录诊断: {path}")
        except Exception as exc:
            print(f"[猫超] 写入登录诊断失败: {exc}")

    def _scope_click_any(self, scope: Any, selectors: list[str], label: str, timeout: int = 10000) -> None:
        locator = self._first_visible_in_scope(scope, selectors, timeout=timeout)
        if locator is None:
            raise RuntimeError(f"找不到{label}: login selector candidates")
        try:
            locator.click(timeout=min(timeout, 5000))
            return
        except Exception:
            if not self._js_click(locator):
                raise RuntimeError(f"点击{label}失败: login selector candidates")

    def _wait_business_home(self, page: Any, timeout_ms: int) -> bool:
        deadline = time.time() + timeout_ms / 1000
        stable_since: float | None = None
        while time.time() < deadline:
            if self._page_looks_logged_in(page):
                if stable_since is None:
                    stable_since = time.time()
                elif time.time() - stable_since >= 1.0:
                    return True
            else:
                stable_since = None
            time.sleep(0.3)
        return False

    def _page_looks_logged_in(self, page: Any) -> bool:
        url = str(getattr(page, "url", "") or "")
        if re.search(r"/login(?:[/?#]|$)", url):
            return False

        blocker_script = """
        () => {
          const href = location.href || '';
          if (/\\/login(?:[/?#]|$)/.test(href)) return true;
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.display !== 'none' &&
              style.visibility !== 'hidden' &&
              Number(style.opacity || 1) > 0;
          };
          if (Array.from(document.querySelectorAll('input[type=password]')).some(visible)) return true;
          const text = ((document.body && document.body.innerText) || '').slice(0, 1200);
          return /WELCOME|向右滑动验证|滑动验证|请完成验证|安全验证|选择商家账号|进入商家/.test(text);
        }
        """
        for scope in self._iter_scopes(page):
            try:
                if scope.evaluate(blocker_script):
                    return False
            except Exception:
                continue

        script = """
        () => {
          const href = location.href || '';
          const hasPassword = !!document.querySelector('input[type=password]');
          if (hasPassword) return false;
          const supplier = document.querySelector('.current-supplier-name');
          const supplierText = ((supplier && supplier.innerText) || '').replace(/\\s+/g, ' ').trim();
          if (supplierText && supplierText !== '全部') return true;
          const text = ((document.body && document.body.innerText) || '').slice(0, 800);
          if (/WELCOME|向右滑动验证|滑动验证|请完成验证|安全验证|选择商家账号|进入商家/.test(text)) return false;
          if (/补货单|供应链AI工作台|采购单列表|库存分析|实时库存/.test(text)) return true;
          return /txcs\\.tmall\\.com/.test(href)
            && !/\\/login(?:[/?#]|$)/.test(href)
            && /(purchase_order|inventory_realtime_search|purchase_transfer_order|merchandise_channel_store|ai_tj_inventory_3)/.test(href)
            && text.length >= 80;
        }
        """
        for scope in self._iter_scopes(page):
            try:
                if scope.evaluate(script):
                    return True
            except Exception:
                continue
        return "txcs.tmall.com" in url and "/login" not in url and "purchase_order" in url

    def _login_verification_visible(self, page: Any) -> bool:
        script = """
        () => {
          const text = ((document.body && document.body.innerText) || '').slice(0, 1200);
          return /向右滑动验证|滑动验证|请完成验证|安全验证/.test(text)
            || !!document.querySelector(
              '[class*="slider"], [class*="slide"], [class*="verify"], [id*="slider"], [id*="verify"]'
            );
        }
        """
        for scope in self._iter_scopes(page):
            try:
                if scope.evaluate(script):
                    return True
            except Exception:
                continue
        return False

    def _manual_login_wait_needed(self, page: Any) -> bool:
        if self.headless:
            return False
        url = str(getattr(page, "url", "") or "")
        return self._login_verification_visible(page) or self._login_form_visible(page) or "/login" in url

    def _manual_login_required_message(self, page: Any, cause: Exception | None = None) -> str:
        url = str(getattr(page, "url", "") or "")
        detail = f"（当前页面: {url[:180]}）" if url else ""
        message = f"登录需要人工处理：RPA 浏览器仍停留在登录页/人工验证页{detail}。请到 Win 主机上的 RPA Chrome 窗口完成登录或滑动验证后重试。"
        if cause:
            return f"{message} 原始原因：{cause}"
        return message

    def _login_failure_message(self, page: Any, cause: Exception | None = None) -> str:
        url = str(getattr(page, "url", "") or "")
        if self._login_verification_visible(page):
            state = "仍停留在登录页的人工验证状态"
        elif "/login" in url:
            state = "仍停留在登录页，人工验证或登录确认未完成"
        elif self._login_form_visible(page):
            state = "仍停留在登录页，登录提交未完成"
        elif self._merchant_selector_visible(page):
            state = "已到商家选择页，但未识别到可进入商家的控件"
        else:
            state = "未识别到商家主页"
        detail = f"（当前页面: {url[:180]}）" if url else ""
        if cause:
            return f"登录流程失败：{state}{detail}。原始原因：{cause}"
        return f"登录流程失败：{state}{detail}。"

    def _handle_merchant_selector(self, page: Any, harvest: bool = False) -> list[SupplierRef]:
        merchant_scope = self._frame_with_selectors(
            page,
            ("merchant.enter_button",),
            timeout=2500,
        )
        if merchant_scope is None and self._selector_visible(page, "merchant.enter_button", timeout=1500):
            merchant_scope = page
        if merchant_scope is None:
            return []
        print("[猫超] 检测到选择商家账号中间页。")
        layout = "三字段(含商家类型)" if self._merchant_type_visible(page, merchant_scope) else "两字段(租户+二级供应商)"
        print(f"[猫超] 选择商家页布局: {layout}")
        if self._merchant_type_visible(page, merchant_scope):
            self._select_merchant_type(merchant_scope, page)
            self._wait_quiet(page, 1500)
        collected: list[SupplierRef] = []
        if harvest:
            collected = self._collect_second_suppliers(page, merchant_scope)
            print(f"[猫超] 登录页二级供应商 {len(collected)} 个")
        self._select_second_supplier(merchant_scope, page)
        self._scope_click(merchant_scope, "merchant.enter_button", "进入商家")
        try:
            page.wait_for_url(re.compile(r"^https://web\.txcs\.tmall\.com/(?:\?|$)"), timeout=15000)
        except Exception:
            pass
        self._wait_quiet(page, 10000)
        return collected

    def _merchant_type_visible(self, page: Any, scope: Any | None = None) -> bool:
        search_scope = scope or page
        for selector in (
            "xpath=//*[contains(normalize-space(.), '商家类型')]",
            "button:has-text(\"品牌商\")",
            "[role=\"button\"]:has-text(\"品牌商\")",
            "span:has-text(\"品牌商\")",
        ):
            try:
                if self._scoped_visible_locator(search_scope, selector, timeout=600) is not None:
                    return True
            except Exception:
                pass
            if self._visible_locator(page, selector, "商家类型", timeout=400) is not None:
                return True
        return False

    def _select_merchant_type(self, scope: Any, page: Any) -> None:
        configured = self._selector_optional("merchant.goods_supplier_radio")
        candidates = [
            configured,
            "button:has-text(\"商品供应商\")",
            "[role=\"button\"]:has-text(\"商品供应商\")",
            "label:has-text(\"商品供应商\")",
            "span:has-text(\"商品供应商\")",
            "xpath=//*[contains(normalize-space(.), '商家类型')]/following::*[normalize-space(.)='商品供应商'][1]",
            "xpath=//*[contains(normalize-space(.), '商品供应商') and not(contains(normalize-space(.), '品牌商'))][1]",
        ]
        for selector in [item for item in candidates if item]:
            try:
                locator = self._scoped_visible_locator(scope, selector, timeout=1200)
                if locator is None:
                    locator = self._visible_locator(page, selector, "商品供应商", timeout=1200)
                if locator is None:
                    continue
                locator.click(timeout=1200)
                self._wait_quiet(page, 500)
                print("[猫超] 商家类型已选择: 商品供应商")
                return
            except Exception:
                continue
        print("[猫超] 商家类型可见但未点到「商品供应商」，继续用当前选项。")

    def _select_second_supplier(self, scope: Any, page: Any) -> None:
        option_candidates = [
            self._selector_optional("merchant.second_supplier_first_option"),
            ".next-overlay-wrapper:visible [role=\"option\"]",
            ".next-overlay-wrapper:visible .next-menu-item",
            ".next-overlay-wrapper:visible .menu-item",
            "[role=\"listbox\"]:visible [role=\"option\"]",
        ]
        if self._open_second_supplier_dropdown(page, scope):
            if self._click_first_dropdown_option(page, option_candidates):
                print("[猫超] 二级供应商已选择。")
                return
            try:
                page.keyboard.press("ArrowDown")
                page.keyboard.press("Enter")
                self._wait_quiet(page, 500)
                print("[猫超] 二级供应商已通过键盘选择。")
                return
            except Exception:
                pass
        print("[猫超] 二级供应商未自动选择，将尝试直接进入商家。")

    def _click_first_dropdown_option(self, page: Any, selectors: list[str]) -> bool:
        skip_texts = {"", "请选择", "无数据", "暂无数据"}
        for selector in [item for item in selectors if item]:
            pw_selector = _pw_selector(selector)
            deadline = time.time() + 2
            while time.time() < deadline:
                for scope in self._iter_scopes(page):
                    try:
                        locator = scope.locator(pw_selector)
                        count = min(locator.count(), 50)
                    except Exception:
                        continue
                    for idx in range(count):
                        item = locator.nth(idx)
                        try:
                            if not item.is_visible(timeout=150):
                                continue
                            text = _clean_text(item.inner_text(timeout=300))
                            if text in skip_texts:
                                continue
                            item.click(timeout=1500)
                            self._wait_quiet(page, 500)
                            return True
                        except Exception:
                            continue
                time.sleep(0.2)
        return False

    def _merchant_selector_visible(self, page: Any, timeout: int = 1200) -> bool:
        if self._visible_locator(page, "button:has-text(\"进入商家\")", "进入商家", timeout=timeout) is not None:
            return True
        if timeout >= 800 and self._selector_visible(page, "merchant.enter_button", timeout=timeout):
            return True
        return self._visible_locator(
            page,
            "xpath=//*[contains(normalize-space(.), '选择商家账号')]",
            "选择商家账号",
            timeout=min(timeout, 600),
        ) is not None

    def _login_form_visible(self, page: Any) -> bool:
        if self._merchant_selector_visible(page):
            return False
        return self._login_form_scope(page, timeout=800) is not None

    def _second_supplier_dropdown_selectors(self) -> list[str]:
        configured = self._selector_optional("merchant.second_supplier_dropdown")
        return [
            "xpath=//*[contains(normalize-space(.), '二级供应商')]/ancestor::*[contains(@class, 'next-form-item')][1]//*[contains(@class, 'next-select-wrapper') or contains(@class, 'next-select')][1]",
            "xpath=//*[contains(normalize-space(.), '二级供应商')]/following::*[contains(@class, 'next-select-wrapper')][1]",
            configured or "",
        ]

    def _discover_account_suppliers(self, page: Any) -> list[SupplierRef]:
        if not self._merchant_selector_visible(page):
            if self._wait_business_home(page, 1000):
                self._reopen_merchant_selector(page)
        if self._merchant_selector_visible(page):
            items = self._handle_merchant_selector(page, harvest=True)
            if items:
                print(f"[猫超] 已从登录页二级供应商同步 {len(items)} 个")
                return items
            print("[猫超] 登录页二级供应商下拉为空")
        print("[猫超] 未读到登录页二级供应商，回退右上角清单")
        return self._discover_header_suppliers(page)

    def _reopen_merchant_selector(self, page: Any) -> bool:
        self._dismiss_blocking_popups(page)
        for text in ("切换商家", "选择商家", "切换账号"):
            if self._click_exact_control(page, text, timeout=1200) or self._click_text(
                page, text, timeout=800, optional=True
            ):
                self._wait_quiet(page, 2500)
                if self._merchant_selector_visible(page):
                    print(f"[猫超] 已打开选择商家页: {text}")
                    return True
        try:
            page.goto(self.settings.login_url, wait_until="domcontentloaded")
            self._wait_quiet(page, 5000)
        except Exception as exc:
            print(f"[猫超] 打开登录页以读取二级供应商失败: {exc}")
            return False
        if self._merchant_selector_visible(page):
            print("[猫超] 已从登录入口进入选择商家页")
            return True
        if self._login_form_visible(page):
            print("[猫超] 登录页出现密码框，立即返回工作台，不重新登录")
            try:
                page.goto("https://web.txcs.tmall.com/", wait_until="domcontentloaded")
                self._wait_quiet(page, 4000)
            except Exception:
                pass
            return False
        return False

    def _open_second_supplier_dropdown(self, page: Any, scope: Any | None = None) -> bool:
        search_scope = scope or page
        for selector in [item for item in self._second_supplier_dropdown_selectors() if item]:
            try:
                dropdown = self._scoped_visible_locator(search_scope, selector, timeout=1500)
                if dropdown is None:
                    dropdown = self._visible_locator(page, selector, "二级供应商", timeout=1500)
                if dropdown is None:
                    continue
                dropdown.click(timeout=1500)
                self._wait_quiet(page, 800)
                return True
            except Exception:
                continue
        return False

    def _collect_second_suppliers(self, page: Any, scope: Any | None = None) -> list[SupplierRef]:
        if not self._open_second_supplier_dropdown(page, scope):
            print("[猫超] 打不开二级供应商下拉")
            return []
        skip_texts = {"", "请选择", "无数据", "暂无数据", "全部"}
        script = """
        () => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const nodes = document.querySelectorAll(
            '.next-overlay-wrapper [role="option"], .next-overlay-wrapper .next-menu-item, .next-overlay-wrapper .menu-item, [role="listbox"] [role="option"]'
          );
          const items = [];
          const seen = new Set();
          for (const el of nodes) {
            const text = textOf(el);
            if (!text || seen.has(text)) continue;
            seen.add(text);
            items.push(text);
          }
          return items;
        }
        """
        names: list[str] = []
        deadline = time.time() + 6
        while time.time() < deadline:
            batch: list[str] = []
            for item_scope in self._iter_scopes(page):
                try:
                    batch.extend(item_scope.evaluate(script) or [])
                except Exception:
                    continue
            for name in batch:
                if name not in names:
                    names.append(name)
            self._scroll_header_supplier_overlay(page)
            time.sleep(0.35)
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        rows: list[SupplierRef] = []
        seen: set[str] = set()
        for name in names:
            text = _clean_text(name)
            if text in skip_texts or is_placeholder_supplier(f"name:{text}", text):
                continue
            parsed = self._parse_ascp_supplier(text)
            if parsed is None:
                parsed = SupplierRef(supplier_id=f"name:{text}", supplier_name=text)
            key = parsed.supplier_id or parsed.supplier_name
            if not key or key in seen:
                continue
            seen.add(key)
            rows.append(parsed)
        return rows

    def _header_supplier_skip_pattern(self) -> re.Pattern[str]:
        return re.compile(r"^(文件|消息|通知|帮助|设置|退出|下载|首页|更多|升级新版|搜索)")

    def _header_chrome_label(self, text: str) -> bool:
        label = _clean_text((text or "").split("\n")[0])
        return bool(self._header_supplier_skip_pattern().search(label)) or bool(
            re.search(r"消息|通知|钉钉|升级新版", label)
        )

    def _prefer_business_page(self, context: Any) -> Any:
        pages = list(getattr(context, "pages", []) or [])
        for page in pages:
            try:
                if self._merchant_selector_visible(page, timeout=500):
                    print("[猫超] 接管选择商家账号页签")
                    try:
                        page.bring_to_front()
                    except Exception:
                        pass
                    return page
            except Exception:
                continue
        for page in pages:
            url = str(getattr(page, "url", "") or "")
            if "txcs.tmall.com" in url:
                return page
        if pages:
            return pages[0]
        return context.new_page()

    def _parse_ascp_supplier(self, text: str) -> SupplierRef | None:
        raw = _clean_text(re.sub(r"^Hi[,，]\s*", "", text or ""))
        if not raw:
            return None
        match = re.match(r"^(\d{6,})-(.+)$", raw)
        if match:
            supplier_id, supplier_name = match.group(1), _clean_text(match.group(2))
            if is_placeholder_supplier(supplier_id, supplier_name):
                return None
            return SupplierRef(supplier_id=supplier_id, supplier_name=supplier_name or supplier_id)
        if is_placeholder_supplier(f"name:{raw}", raw) or self._header_chrome_label(raw):
            return None
        return SupplierRef(supplier_id=f"name:{raw}", supplier_name=raw)

    def _collect_ascp_header_suppliers(self, page: Any) -> list[SupplierRef]:
        script = """
        () => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const root = document.querySelector('.current-supplier-name')
            ? document.querySelector('.current-supplier-name').closest('.ascp-frame-dropdown')
            : document.querySelector('.ascp-frame-dropdown.ascp-header-dropdown-item');
          if (!root) return [];
          const nodes = root.querySelectorAll('.supplier-name, .ascp-frame-menu-item');
          const items = [];
          const seen = new Set();
          for (const el of nodes) {
            const name = textOf(el);
            if (!name || seen.has(name)) continue;
            seen.add(name);
            items.push(name);
          }
          return items;
        }
        """
        names: list[str] = []
        for scope in self._iter_scopes(page):
            try:
                names.extend(scope.evaluate(script) or [])
            except Exception:
                continue
        rows: list[SupplierRef] = []
        seen: set[str] = set()
        for name in names:
            parsed = self._parse_ascp_supplier(name)
            if parsed is None or parsed.supplier_id.startswith("name:") or parsed.supplier_id in seen:
                continue
            seen.add(parsed.supplier_id)
            rows.append(parsed)
        return rows

    def _discover_header_suppliers(self, page: Any) -> list[SupplierRef]:
        opened = self._open_header_supplier_dropdown(page)
        items = self._collect_ascp_header_suppliers(page)
        if not items:
            items = self._wait_header_supplier_options(page, timeout_ms=5000)
        current = self._current_header_supplier_text(page)
        real = [item for item in items if not is_placeholder_supplier(item.supplier_id, item.supplier_name)]
        if not real and current and not is_placeholder_supplier(f"name:{current}", current):
            parsed = self._parse_ascp_supplier(current)
            print(f"[猫超] 下拉未列出完整清单，先登记当前供应商: {current}")
            real = [parsed] if parsed is not None else [SupplierRef(supplier_id=f"name:{current}", supplier_name=current)]
        if not real:
            raise RuntimeError(
                "未能读取右上角真实供应商。请确认已登录，且右上角是具体公司名而不是搜索框里的「全部」。"
            )
        if opened:
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            self._wait_quiet(page, 400)
        print(f"[猫超] 当前右上角供应商: {current or '未知'}，清单 {len(real)} 个")
        return real

    def _switch_header_supplier(self, page: Any, supplier: SupplierRef) -> SupplierRef:
        self._dismiss_notification_center(page)
        self._dismiss_blocking_popups(page)
        if is_placeholder_supplier(supplier.supplier_id, supplier.supplier_name):
            raise RuntimeError("不能切换到「全部」。请选择具体供应商后再执行任务。")
        current = self._current_header_supplier(page)
        if current and self._header_supplier_matches(supplier, current):
            print(f"[猫超] 右上角已是目标供应商: {current.supplier_name} id={current.supplier_id}")
            return current
        print("[猫超] 切供应商前打开工作台首页")
        try:
            page.goto("https://web.txcs.tmall.com/", wait_until="domcontentloaded")
            self._wait_quiet(page, 3500)
            self._dismiss_blocking_popups(page)
        except Exception as exc:
            print(f"[猫超] 打开工作台首页失败: {exc}")
            self._goto_workbench_home(page)
        items = self._collect_ascp_header_suppliers(page)
        if not items:
            self._open_header_supplier_dropdown(page)
            items = self._collect_ascp_header_suppliers(page) or self._wait_header_supplier_options(page, timeout_ms=5000)
        target = self._match_header_supplier(supplier, items)
        if target is None:
            if not items:
                print(
                    f"[猫超] 新工作台没有右上角供应商下拉，沿用当前登录会话继续: "
                    f"{supplier.supplier_name} id={supplier.supplier_id}"
                )
                return supplier
            raise RuntimeError(
                f"右上角供应商下拉中找不到: id={supplier.supplier_id or '-'} name={supplier.supplier_name or '-'}"
            )
        if not self._click_header_supplier_option(page, target):
            raise RuntimeError(f"点击右上角供应商失败: {target.supplier_name}")
        self._wait_quiet(page, 2500)
        confirmed = self._wait_header_supplier(page, target, timeout_ms=15000)
        if confirmed is None and self._header_display_matches(page, target):
            print(f"[猫超] 已切换右上角供应商: {target.supplier_name} id={target.supplier_id}（按显示名确认）")
            return target
        if confirmed is None:
            raise RuntimeError(
                f"右上角供应商切换未确认成功: id={target.supplier_id} name={target.supplier_name} "
                f"当前显示={self._current_header_supplier_text(page) or '未知'}"
            )
        print(f"[猫超] 已切换右上角供应商: {confirmed.supplier_name} id={confirmed.supplier_id}")
        self._wait_quiet(page, 4000)
        return confirmed

    def _header_display_matches(self, page: Any, expected: SupplierRef) -> bool:
        text = self._current_header_supplier_text(page)
        name = expected.supplier_name or ""
        if not text or not name or len(text) < 12:
            return False
        text_norm = self._normalize_supplier_text(text)
        name_norm = self._normalize_supplier_text(name)
        return bool(text_norm and name_norm) and (
            name_norm.startswith(text_norm)
            or text_norm in name_norm
            or self._supplier_name_score(name, text) >= 0.78
        )

    def _current_header_supplier(self, page: Any) -> SupplierRef | None:
        identified = self._current_ascp_supplier(page)
        if identified is not None:
            return identified
        text = self._current_header_supplier_text(page)
        if not text:
            return None
        return SupplierRef(supplier_id=f"name:{text}", supplier_name=text)

    def _current_ascp_supplier(self, page: Any) -> SupplierRef | None:
        script = """
        () => {
          const root = document.querySelector('.current-supplier-name')
            ? document.querySelector('.current-supplier-name').closest('.ascp-frame-dropdown')
            : null;
          if (!root) return { current: '', hi: '' };
          const current = (root.querySelector('.current-supplier-name')?.innerText || '').replace(/\\s+/g, ' ').trim();
          const first = (root.querySelector('.ascp-frame-menu-item')?.innerText || '').replace(/\\s+/g, ' ').trim();
          return { current, hi: first };
        }
        """
        for scope in self._iter_scopes(page):
            try:
                payload = scope.evaluate(script) or {}
            except Exception:
                continue
            parsed = self._parse_ascp_supplier(str(payload.get("hi") or ""))
            if parsed is not None and not parsed.supplier_id.startswith("name:"):
                return parsed
            current = _clean_text(payload.get("current"))
            if current:
                items = self._collect_ascp_header_suppliers(page)
                hits = [
                    item
                    for item in items
                    if item.supplier_name == current
                    or item.supplier_name.startswith(current)
                ]
                if len(hits) == 1:
                    return hits[0]
        return None

    def _wait_header_supplier(self, page: Any, expected: SupplierRef, timeout_ms: int = 15000) -> SupplierRef | None:
        deadline = time.time() + timeout_ms / 1000
        last: SupplierRef | None = None
        while time.time() < deadline:
            last = self._current_header_supplier(page)
            if last and self._header_supplier_matches(expected, last):
                return last
            time.sleep(0.4)
        return last if last and self._header_supplier_matches(expected, last) else None

    def _header_supplier_matches(self, expected: SupplierRef, actual: SupplierRef) -> bool:
        expected_id = _clean_text(expected.supplier_id)
        actual_id = _clean_text(actual.supplier_id)
        if expected_id and not expected_id.startswith("name:"):
            return bool(actual_id) and actual_id == expected_id
        if expected.supplier_name and actual.supplier_name:
            return self._normalize_supplier_text(expected.supplier_name) == self._normalize_supplier_text(actual.supplier_name)
        if expected_id.startswith("name:") and actual.supplier_name:
            return self._normalize_supplier_text(expected_id[5:]) == self._normalize_supplier_text(actual.supplier_name)
        return False

    def _match_header_supplier(self, expected: SupplierRef, items: list[SupplierRef]) -> SupplierRef | None:
        if expected.supplier_id:
            for item in items:
                if item.supplier_id and item.supplier_id == expected.supplier_id:
                    return item
        if expected.supplier_name:
            best: SupplierRef | None = None
            best_score = 0.0
            for item in items:
                score = self._supplier_name_score(expected.supplier_name, item.supplier_name)
                if score > best_score:
                    best_score = score
                    best = item
            if best is not None and best_score >= 0.78:
                return best
        return None

    def _open_header_supplier_dropdown(self, page: Any) -> bool:
        self._dismiss_notification_center(page)
        if self._header_supplier_overlay_open(page):
            print("[猫超] 右上角供应商下拉已展开")
            return True
        script = """
        () => {
          const name = document.querySelector('.current-supplier-name');
          if (!name) return 'missing';
          const trig = name.closest('.trigger') || name.closest('.ascp-frame-dropdown') || name;
          trig.dispatchEvent(new MouseEvent('mouseenter', {bubbles: true}));
          trig.dispatchEvent(new MouseEvent('mouseover', {bubbles: true, view: window}));
          trig.dispatchEvent(new MouseEvent('mousedown', {bubbles: true, cancelable: true, view: window}));
          trig.dispatchEvent(new MouseEvent('mouseup', {bubbles: true, cancelable: true, view: window}));
          trig.dispatchEvent(new MouseEvent('click', {bubbles: true, cancelable: true, view: window}));
          return 'clicked';
        }
        """
        for scope in self._iter_scopes(page):
            try:
                hit = scope.evaluate(script)
            except Exception:
                continue
            if hit == "clicked":
                if self._wait_header_supplier_options(page, timeout_ms=3500):
                    print("[猫超] 已打开右上角供应商下拉")
                    return True
                if self._header_supplier_overlay_open(page):
                    print("[猫超] 已打开右上角供应商下拉")
                    return True
        if self._open_header_supplier_by_right_items(page):
            print("[猫超] 已通过右上角区域打开供应商下拉")
            return True
        return False

    def _header_supplier_overlay_open(self, page: Any) -> bool:
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 8 && rect.height > 8 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const nodes = document.querySelectorAll(
            'li.ascp-frame-menu-item, .ascp-frame-dropdown .supplier-name, .next-overlay-wrapper.opened [role="option"]'
          );
          let count = 0;
          for (const el of nodes) {
            if (visible(el)) count += 1;
          }
          return count >= 2;
        }
        """
        for scope in self._iter_scopes(page):
            try:
                if scope.evaluate(script):
                    return True
            except Exception:
                continue
        return False

    def _click_and_wait_supplier_overlay(self, page: Any, locator: Any) -> bool:
        try:
            locator.click(timeout=1500)
        except Exception:
            try:
                locator.click(timeout=1500, force=True)
            except Exception:
                return False
        items = self._wait_header_supplier_options(page, timeout_ms=4000)
        if items:
            return True
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        self._wait_quiet(page, 200)
        return False

    def _open_header_supplier_by_right_items(self, page: Any) -> bool:
        try:
            locators = page.locator("header .header-right-item, header [class*='header-right'] > *")
            count = min(locators.count(), 20)
        except Exception:
            return False
        for idx in range(count):
            item = locators.nth(idx)
            try:
                if not item.is_visible(timeout=200):
                    continue
                text = _clean_text(item.inner_text(timeout=300))
                label = text.split("\n")[0] if text else ""
                if not label or self._header_chrome_label(label) or len(label) > 40:
                    continue
                if self._click_and_wait_supplier_overlay(page, item):
                    return True
            except Exception:
                continue
        current = self._current_header_supplier_text(page)
        if current and not is_placeholder_supplier(f"name:{current}", current):
            locator = self._visible_locator(
                page,
                f"header :text(\"{current[:6]}\")",
                "当前供应商名",
                timeout=1500,
            )
            if locator is not None and self._click_and_wait_supplier_overlay(page, locator):
                return True
        return False

    def _wait_header_supplier_options(self, page: Any, timeout_ms: int = 8000) -> list[SupplierRef]:
        deadline = time.time() + timeout_ms / 1000
        last: list[SupplierRef] = []
        while time.time() < deadline:
            last = self._header_supplier_option_items(page, include_placeholders=True)
            real = [item for item in last if not is_placeholder_supplier(item.supplier_id, item.supplier_name)]
            if real:
                return real
            self._scroll_header_supplier_overlay(page)
            time.sleep(0.35)
        return [item for item in last if not is_placeholder_supplier(item.supplier_id, item.supplier_name)]

    def _scroll_header_supplier_overlay(self, page: Any) -> None:
        script = """
        () => {
          const panels = document.querySelectorAll(
            '.next-overlay-wrapper .next-select-menu, .next-overlay-wrapper [role="listbox"], [role="listbox"]'
          );
          for (const panel of panels) {
            const rect = panel.getBoundingClientRect();
            if (rect.height > 0) panel.scrollTop += Math.max(160, rect.height - 20);
          }
        }
        """
        for scope in self._iter_scopes(page):
            try:
                scope.evaluate(script)
            except Exception:
                continue

    def _header_supplier_option_items(self, page: Any, include_placeholders: bool = False) -> list[SupplierRef]:
        configured = self._selector_optional("header_supplier.option")
        payload: list[dict[str, Any]] = []
        script = """
            (configuredSelector) => {
              const visible = (el) => {
                if (!el) return false;
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return rect.width > 0 && rect.height > 0 &&
                  style.visibility !== 'hidden' && style.display !== 'none' &&
                  Number(style.opacity || 1) > 0;
              };
              const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
              const chrome = /^(文件|消息|通知|帮助|设置|退出|下载|首页|更多|升级新版|搜索|请选择|无数据|暂无数据)/;
              const selectors = [
                configuredSelector,
                'li.ascp-frame-menu-item .supplier-name',
                'li.ascp-frame-menu-item',
                '.next-overlay-wrapper [role="option"]',
                '.next-overlay-wrapper .next-menu-item',
                '.next-overlay-wrapper .menu-item',
                '.next-overlay-wrapper li',
                '[role="listbox"] [role="option"]',
                '.header-supplier-panel li',
                '.shop-list-item'
              ].filter(Boolean);
              const items = [];
              const seen = new Set();
              for (const selector of selectors) {
                for (const el of document.querySelectorAll(selector)) {
                  if (!visible(el)) continue;
                  const name = textOf(el);
                  if (!name || name.length > 120 || chrome.test(name)) continue;
                  const id = el.getAttribute('data-id')
                    || el.getAttribute('data-value')
                    || el.getAttribute('data-key')
                    || el.getAttribute('value')
                    || (el.dataset && (el.dataset.id || el.dataset.value || el.dataset.key))
                    || '';
                  const key = String(id || '') + '|' + name;
                  if (seen.has(key)) continue;
                  seen.add(key);
                  items.push({ supplier_id: String(id || ''), supplier_name: name });
                }
              }
              return items;
            }
        """
        for scope in self._iter_scopes(page):
            try:
                payload.extend(scope.evaluate(script, configured) or [])
            except Exception:
                continue
        rows: list[SupplierRef] = []
        seen: set[str] = set()
        for item in payload:
            name = _clean_text(item.get("supplier_name"))
            supplier_id = _clean_text(item.get("supplier_id")) or f"name:{name}"
            if not name or supplier_id in seen:
                continue
            if not include_placeholders and is_placeholder_supplier(supplier_id, name):
                continue
            seen.add(supplier_id)
            rows.append(SupplierRef(supplier_id=supplier_id, supplier_name=name))
        return rows

    def _click_header_supplier_option(self, page: Any, supplier: SupplierRef) -> bool:
        self._dismiss_notification_center(page)
        if not self._header_supplier_overlay_open(page):
            if not self._open_header_supplier_dropdown(page):
                print("[猫超] 右上角供应商下拉未展开")
        unique = ""
        name = supplier.supplier_name or ""
        if "--" in name:
            candidate = name.split("--", 1)[-1].strip()
            if len(self._normalize_supplier_text(candidate)) >= 4:
                unique = candidate
        payload = {
            "id": supplier.supplier_id if supplier.supplier_id and not supplier.supplier_id.startswith("name:") else "",
            "name": name,
            "unique": unique,
        }
        script = """
        ({id, name, unique}) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 8 && rect.height > 8 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const normalize = (value) => String(value || '').replace(/[\\s\\-_/／·—–,，.。()（）\\[\\]【】]+/g, '');
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const mouseClick = (el) => {
            el.scrollIntoView({block: 'nearest'});
            const rect = el.getBoundingClientRect();
            const opts = {
              bubbles: true, cancelable: true, view: window,
              clientX: rect.left + rect.width / 2,
              clientY: rect.top + rect.height / 2,
            };
            for (const type of ['mouseenter', 'mouseover', 'mousedown', 'mouseup', 'click']) {
              el.dispatchEvent(new MouseEvent(type, opts));
            }
          };
          const nodes = Array.from(document.querySelectorAll(
            'li.ascp-frame-menu-item, li.ascp-frame-menu-item .supplier-name, .ascp-frame-dropdown .supplier-name, .next-overlay-wrapper [role="option"], .next-overlay-wrapper .next-menu-item'
          ));
          const labels = [];
          let best = null;
          let bestLabel = '';
          let bestScore = 0;
          for (const el of nodes) {
            const isVisible = visible(el);
            const text = textOf(el);
            if (!text || text.startsWith('Hi,') || /退出登录|注销/.test(text) || text.length > 120) continue;
            if (isVisible) labels.push(text.slice(0, 60));
            const dataId = el.getAttribute('data-id') || el.getAttribute('data-value') || '';
            const textNorm = normalize(text);
            const nameNorm = normalize(name);
            const uniqueNorm = normalize(unique);
            let score = 0;
            if (id && (text.includes(id) || dataId === id)) score += 100;
            if (uniqueNorm && textNorm.includes(uniqueNorm)) score += 40;
            if (nameNorm && (textNorm.includes(nameNorm) || nameNorm.includes(textNorm))) score += 80;
            if (name && text.includes(name)) score += 20;
            if (isVisible) score += 5;
            if (score > bestScore) {
              best = el.closest('li') || el;
              bestLabel = text;
              bestScore = score;
            }
          }
          if (!best || bestScore < 20) {
            return {ok: false, labels: labels.slice(0, 8), score: bestScore};
          }
          mouseClick(best.querySelector('.supplier-name') || best);
          return {ok: true, label: bestLabel, score: bestScore, labels: labels.slice(0, 8)};
        }
        """
        for scope in self._iter_scopes(page):
            try:
                result = scope.evaluate(script, payload)
            except Exception:
                continue
            if not result:
                continue
            if result.get("ok"):
                print(f"[猫超] 已点选供应商: {result.get('label')} score={result.get('score')}")
                return True
            labels = result.get("labels") or []
            if labels:
                print(f"[猫超] 供应商下拉可见项: {' / '.join(str(x) for x in labels)}")
        if self._playwright_click_supplier_option(page, supplier):
            return True
        return False

    def _playwright_click_supplier_option(self, page: Any, supplier: SupplierRef) -> bool:
        needles = []
        if supplier.supplier_id and not supplier.supplier_id.startswith("name:"):
            needles.append(supplier.supplier_id)
        name = supplier.supplier_name or ""
        if "--" in name:
            candidate = name.split("--", 1)[-1].strip()
            if len(self._normalize_supplier_text(candidate)) >= 4:
                needles.append(candidate)
        if name:
            needles.append(name)
        for scope in self._iter_scopes(page):
            for needle in needles:
                if not needle:
                    continue
                try:
                    locator = scope.locator(
                        "li.ascp-frame-menu-item, .supplier-name, [role='option'], .next-menu-item"
                    ).filter(has_text=needle)
                    if locator.count() == 0:
                        continue
                    item = locator.first
                    if not self._element_is_displayed(item):
                        continue
                    item.click(timeout=2000)
                    print(f"[猫超] 已鼠标点选供应商: {needle}")
                    return True
                except Exception:
                    continue
        return False

    def _current_header_supplier_text(self, page: Any) -> str:
        for selector in (
            self._selector_optional("header_supplier.current"),
            ".current-supplier-name",
        ):
            if not selector:
                continue
            locator = self._visible_locator(page, selector, "当前供应商", timeout=1200)
            if locator is None:
                continue
            try:
                text = _clean_text(locator.inner_text(timeout=500)).split("\n")[0]
                if text and not is_placeholder_supplier(f"name:{text}", text) and not self._header_chrome_label(text):
                    return text
            except Exception:
                continue
        script = """
            () => {
              const current = document.querySelector('.current-supplier-name');
              if (current) {
                const text = (current.innerText || current.textContent || '').replace(/\\s+/g, ' ').trim();
                if (text) return text;
              }
              const chrome = /^(文件|消息|通知|帮助|设置|退出|下载|首页|更多|升级新版|搜索|全部)$/;
              const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
              const nodes = document.querySelectorAll(
                'header .next-select-values, header .next-select-trigger, header .header-right-item, header [class*="header-right"] > *'
              );
              const hits = [];
              for (const el of nodes) {
                if (el.closest('.search-wrapper')) continue;
                const rect = el.getBoundingClientRect();
                if (!(rect.width > 0 && rect.height > 0 && rect.top >= 0 && rect.top < 90)) continue;
                const text = textOf(el).split(/\\n/)[0].replace(/\\s+/g, ' ').trim();
                if (!text || text.length > 40 || chrome.test(text) || /消息|通知|钉钉|升级新版/.test(text)) continue;
                hits.push(text);
              }
              const preferred = hits.find((text) => /公司|贸易|有限|寄售/.test(text));
              return preferred || hits[0] || '';
            }
        """
        for scope in self._iter_scopes(page):
            try:
                text = _clean_text(scope.evaluate(script))
            except Exception:
                continue
            if text and not is_placeholder_supplier(f"name:{text}", text):
                return text
        return ""

    def _task_realtime_inventory(self, page: Any, account: Account) -> list[RunResult]:
        """任务 1：只处理当前这一个「运营已分配」供应商，导出一张实时库存表。

        选谁不由任务 1 决定，也不扫描实时库存页全部供应商。
        外层已经按运营分配名单切到右上角供应商 A，这里只对 A 查一次、导一张。
        """
        supplier = self._current_supplier
        if supplier is None:
            raise RuntimeError("任务 1 缺少运营已分配的供应商。请先选择运营负责的供应商，再导出实时库存。")

        self._open_task_page(
            page,
            "realtime-inventory",
            (
                "realtime.menu_inventory",
                "realtime.menu_inventory_query",
            ),
        )
        self._select_current_realtime_supplier(page, supplier)
        self._wait_quiet(page, 1500)
        self._click(page, "realtime.query_button", "查询")
        result_count = self._wait_realtime_inventory_result_count(page, timeout_ms=12000)
        return [self._export_realtime_supplier(page, account, supplier.supplier_name or supplier.supplier_id, result_count)]

    def _select_current_realtime_supplier(self, page: Any, supplier: SupplierRef) -> None:
        if not supplier.supplier_id and not supplier.supplier_name:
            raise RuntimeError("任务 1 缺少供应商 ID，不能用页面第一项兜底。")
        if not self._selector_visible(page, "realtime.supplier_field", timeout=1500):
            print(
                f"[猫超] 实时库存页未见供应商名称筛选项，沿用已确认的右上角供应商: "
                f"{supplier.supplier_name or supplier.supplier_id}"
            )
            return
        self._select_realtime_supplier(page, supplier)
        print(f"[猫超] 实时库存已按同一供应商确认: {supplier.supplier_name or supplier.supplier_id}")

    def _task_pincang_detail(self, page: Any, account: Account) -> list[RunResult]:
        # 新工作台没有旧「天机」顶栏。三个账号同一套页面：
        # iframe ai_tj_inventory_3 → 页签「品仓明细」→ 表格下载图标 →「导出货品明细」。
        last_exc: Exception | None = None
        for attempt in range(1, 3):
            try:
                self._open_pincang_page(page)
                self._click_page_tab(page, "品仓明细", sibling_hints=["概览指标", "诊断建议", "库龄分析", "效期分析"])
                self._wait_quiet(page, 2000)
                if self._wait_pincang_download_icon(page, timeout_ms=20000):
                    break
                if self._pincang_detail_has_no_data(page):
                    return [self._no_data_result("pincang-detail", account, "品仓明细表无数据，未生成下载文件")]
                raise RuntimeError("已打开库存分析，但品仓明细表下载图标未出现")
            except Exception as exc:
                last_exc = exc
                print(f"[猫超] 品仓第 {attempt} 次进入失败: {exc}")
                self._goto_workbench_home(page)
        else:
            raise RuntimeError(str(last_exc) if last_exc else "打开品仓明细失败")
        if not self._click_toolbar_export(
            page,
            option_texts=["导出货品明细", "导出明细", "导出列表"],
            allow_direct=True,
            toolbar_title="品仓明细表",
            file_task_key="pincang-detail",
        ):
            raise RuntimeError("品仓明细表导出未生成新文件任务")
        return [self._download_and_clean(page, "pincang-detail", account)]

    def _open_pincang_page(self, page: Any) -> None:
        self._dismiss_notification_center(page)
        self._dismiss_blocking_popups(page)
        direct_url = (
            self.settings.direct_urls.get("pincang-detail", "")
            or "https://web.txcs.tmall.com/?frameUrl=https%3A%2F%2Fweb.txcs.tmall.com%2Fpages%2Fchaoshi%2Fai_tj_inventory_3"
        )
        last_exc: Exception | None = None
        for attempt in range(1, 3):
            try:
                if direct_url:
                    print(f"[猫超] 打开库存分析直达页 ({attempt}/2)")
                    page.goto(direct_url, wait_until="domcontentloaded")
                    self._wait_quiet(page, 8000)
                else:
                    self._goto_workbench_home(page)
                    if not (self._click_sidebar_link(page, "库存分析") or self._search_and_open_menu(page, "库存分析")):
                        raise RuntimeError("侧栏/搜索打不开「库存分析」")
                    self._wait_quiet(page, 5000)
                self._dismiss_blocking_popups(page)
                if self._wait_frame_url_contains(page, "ai_tj_inventory_3", timeout_ms=10000):
                    if self._wait_toolbar_title(page, "概览指标", timeout_ms=8000) or self._pincang_tabs_ready(page):
                        return
                raise RuntimeError("库存分析页未加载出页签")
            except Exception as exc:
                last_exc = exc
                print(f"[猫超] 库存分析第 {attempt} 次打开失败: {exc}")
                self._goto_workbench_home(page)
                if self._click_sidebar_link(page, "库存分析") or self._search_and_open_menu(page, "库存分析"):
                    self._wait_quiet(page, 5000)
                    if self._wait_frame_url_contains(page, "ai_tj_inventory_3", timeout_ms=8000):
                        if self._wait_toolbar_title(page, "概览指标", timeout_ms=5000) or self._pincang_tabs_ready(page):
                            return
                        print("[猫超] 库存分析 iframe 已出现，但页签仍未加载，继续重试")
        raise RuntimeError(f"打开库存分析失败: {last_exc}")

    def _wait_pincang_download_icon(self, page: Any, timeout_ms: int = 15000) -> bool:
        script = """
        () => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const visible = (el) => {
            if (!el) return false;
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 8 && rect.height > 8 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const titles = Array.from(document.querySelectorAll('.comp-toolbar-title-text'));
          const hit = titles.find((el) => ['品仓明细表', '品仓明细'].includes(textOf(el)));
          const downloadSelector = [
            '.toolbar-func-download',
            '.toolbar-func-export',
            '[data-tip*="下载"]',
            '[data-tip*="导出"]',
            '[title*="下载"]',
            '[title*="导出"]',
            '[aria-label*="下载"]',
            '[aria-label*="导出"]',
            '.toolbar-gei-export-button-wrapper',
            '[class*="download"]',
            '[class*="export"]'
          ].join(',');
          const actionable = (el) => {
            if (!visible(el)) return false;
            const rect = el.getBoundingClientRect();
            const small = rect.width <= 260 && rect.height <= 90;
            return small && (
              el.matches('button, a, span, i, [role="button"]') ||
              !!el.closest('.comp-toolbar, [class*="toolbar"]')
            );
          };
          const toolbars = [];
          if (hit) {
            hit.scrollIntoView({block: 'center', inline: 'nearest'});
            const toolbar = hit.closest('.comp-toolbar');
            if (toolbar) toolbars.push(toolbar);
          } else {
            for (const toolbar of document.querySelectorAll('.comp-toolbar, [class*="toolbar"]')) {
              if (visible(toolbar)) toolbars.push(toolbar);
            }
          }
          if (!toolbars.length) return hit ? 'no-toolbar' : 'no-title';
          for (const toolbar of toolbars) {
            const btn = Array.from(toolbar.querySelectorAll(downloadSelector)).find(actionable);
            if (btn) return 'ok';
          }
          const loose = Array.from(document.querySelectorAll(downloadSelector)).find(actionable);
          if (!loose) return 'no-btn';
          return 'ok';
        }
        """
        deadline = time.time() + timeout_ms / 1000
        last = ""
        self._dismiss_notification_center(page)
        self._dismiss_blocking_popups(page)
        while time.time() < deadline:
            for scope in self._iter_scopes(page):
                try:
                    last = scope.evaluate(script) or ""
                except Exception:
                    continue
                if last == "ok":
                    print("[猫超] 品仓明细表下载图标已出现")
                    return True
            time.sleep(0.4)
        diag = self._pincang_detail_diagnostic(page)
        if diag:
            print(f"[猫超] 品仓下载图标诊断: {diag}")
        print(f"[猫超] 等待品仓下载图标超时: {last or 'unknown'}")
        return False

    def _pincang_detail_diagnostic(self, page: Any) -> str:
        script = """
        () => {
          const visible = (el) => {
            if (!el) return false;
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el?.innerText || el?.textContent || '').replace(/\\s+/g, ' ').trim();
          const labelsOf = (selector) => Array.from(document.querySelectorAll(selector))
            .filter(visible)
            .map((el) => textOf(el) || el.getAttribute('title') || el.getAttribute('aria-label') || el.getAttribute('data-tip') || el.className || '')
            .map((text) => String(text).replace(/\\s+/g, ' ').trim())
            .filter(Boolean)
            .slice(0, 20);
          const body = textOf(document.body);
          return {
            url: location.href,
            hasPincang: body.includes('品仓明细'),
            empty: /暂无数据|无数据|没有数据|暂无符合条件的数据|当前查询无数据|共\\s*0\\s*项/.test(body),
            countText: (body.match(/共\\s*[0-9,]+\\s*项/) || [''])[0],
            rowCount: Array.from(document.querySelectorAll('.next-table-body .next-table-row, tbody tr.next-table-row, .river-table tbody tr')).filter(visible).length,
            tabs: labelsOf('[role="tab"], .next-tabs-tab, li.next-tabs-tab'),
            toolbarTitles: labelsOf('.comp-toolbar-title-text'),
            downloadLike: labelsOf('button, a, span, div, i, [role="button"], [title], [aria-label], [data-tip]')
              .filter((text) => /下载|导出|download|export|toolbar|品仓/.test(text))
              .slice(0, 15),
            bodySlice: body.slice(0, 180),
          };
        }
        """
        items: list[dict[str, Any]] = []
        for scope in self._iter_scopes(page):
            try:
                data = scope.evaluate(script) or {}
            except Exception:
                continue
            url = str(data.get("url") or "")
            if data.get("hasPincang") or "ai_tj_inventory_3" in url:
                items.append(data)
        if not items:
            return ""
        return json.dumps(items[:3], ensure_ascii=False)

    def _pincang_detail_has_no_data(self, page: Any) -> bool:
        script = """
        () => {
          const visible = (el) => {
            if (!el) return false;
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const tabs = Array.from(document.querySelectorAll('[role="tab"], .next-tabs-tab, li.next-tabs-tab'))
            .map(textOf);
          const body = textOf(document.body);
          const toolbarTitles = Array.from(document.querySelectorAll('.comp-toolbar-title-text'))
            .filter(visible)
            .map(textOf);
          const ready = tabs.some((text) => text.includes('品仓明细'))
            || toolbarTitles.some((text) => text.includes('品仓明细'))
            || body.includes('品仓明细');
          if (!ready) return {ready: false};
          const rows = Array.from(document.querySelectorAll(
            '.next-table-body .next-table-row, tbody tr.next-table-row, .river-table tbody tr'
          )).filter(visible).length;
          const count0 = /共\\s*0\\s*项/.test(body);
          const empty = /暂无数据|无数据|没有数据|暂无符合条件的数据|当前查询无数据/.test(body);
          const headerOnly = body.includes('SKUID') && body.includes('货品ID') && body.includes('诊断类型');
          const hasDownload = Array.from(document.querySelectorAll(
            '.toolbar-func-download, .toolbar-func-export, [data-tip*="下载"], [data-tip*="导出"], [title*="下载"], [title*="导出"], [aria-label*="下载"], [aria-label*="导出"], .toolbar-gei-export-button-wrapper, [class*="download"], [class*="export"]'
          )).some((el) => {
            if (!visible(el)) return false;
            const rect = el.getBoundingClientRect();
            const small = rect.width <= 260 && rect.height <= 90;
            return small && (
              el.matches('button, a, span, i, [role="button"]') ||
              !!el.closest('.comp-toolbar, [class*="toolbar"]')
            );
          });
          return {ready, rows, count0, empty, headerOnly, hasDownload};
        }
        """
        for scope in self._iter_scopes(page):
            try:
                result = scope.evaluate(script) or {}
            except Exception:
                continue
            if not result.get("ready"):
                continue
            rows = int(result.get("rows") or 0)
            if rows == 0 and (result.get("count0") or result.get("empty") or (result.get("headerOnly") and not result.get("hasDownload"))):
                print("[猫超] 品仓明细表为空，跳过导出")
                return True
        return False

    def _pincang_tabs_ready(self, page: Any) -> bool:
        script = """
        () => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const tabs = Array.from(document.querySelectorAll('[role="tab"], .next-tabs-tab'));
          return tabs.some((el) => textOf(el) === '品仓明细');
        }
        """
        for scope in self._iter_scopes(page):
            try:
                if scope.evaluate(script):
                    return True
            except Exception:
                continue
        return False

    def _task_system_order(self, page: Any, account: Account) -> list[RunResult]:
        self._open_purchase_replenishment(page, task_key="system-order", force=True)
        statuses = self._list_config("system_order.statuses", ["待供应商确认"])
        self._select_first_purchase_status(page, statuses)
        self._click(page, "purchase.query_button", "查询")
        self._wait_quiet(page, 5000)
        self._click(page, "system_order.import_button", "导入")
        self._click(page, "system_order.import_confirm_option", "导入确认")
        self._wait_quiet(page, 2000)
        self._snapshot_file_task_ids(page)
        self._click(page, "system_order.dialog_export_data", "导出数据", timeout=5000)
        return [self._download_and_clean(page, "system-order", account)]

    def _task_po_list(self, page: Any, account: Account) -> list[RunResult]:
        self._open_purchase_replenishment(page, task_key="po-list", force=True)
        self._expand_more_filters(page)
        self._fill_last_two_months(page)
        self._select_po_list_statuses(page)
        self._click(page, "purchase.query_button", "查询")
        self._wait_quiet(page, 5000)
        if self._po_list_is_empty(page):
            if self._clear_po_list_statuses(page):
                print("[猫超] 补货单固定状态筛选为 0，已清空采购单状态并重查")
                self._click(page, "purchase.query_button", "查询")
                self._wait_quiet(page, 5000)
            if self._po_list_is_empty(page):
                return [self._no_data_result("po-list", account, "补货单列表无数据，未生成下载文件")]
            print("[猫超] 补货单清空采购单状态后查到数据，继续导出")
        self._unclick_current_page_only_if_present(page)
        self._export_po_list(page)
        return [self._download_and_clean(page, "po-list", account)]

    def _po_list_is_empty(self, page: Any) -> bool:
        count = self._visible_result_count(page)
        if count is not None:
            print(f"[猫超] 补货单查询结果: 共 {count} 项")
            return count == 0
        if self._page_has_no_items(page):
            print("[猫超] 补货单查询结果: 共 0 项")
            return True
        has_row = self._visible_locator(
            page,
            ".next-table-body .next-table-row, tbody tr.next-table-row, .next-table tbody tr",
            "补货单行",
            timeout=1500,
        ) is not None
        print(f"[猫超] 补货单查询结果: {'有表格行' if has_row else '未见表格行'}")
        return not has_row

    def _export_po_list(self, page: Any) -> None:
        if not self._click_toolbar_export(
            page,
            option_texts=["导出明细"],
            allow_direct=False,
            file_task_key="po-list",
        ):
            raise RuntimeError("补货单列表导出未生成新文件任务")

    def _select_po_list_statuses(self, page: Any) -> None:
        requested = self._list_config(
            "po_list.statuses",
            ["待供应商预约", "供应商已确认", "待收货", "待部分收货"],
        )
        aliases = {
            "待供应商预约": ["待供应商预约"],
            "供应商已确认": ["供应商已确认"],
            "待收货": ["待收货"],
            "待部分收货": ["待部分收货", "部分收货"],
            "部分收货": ["部分收货", "待部分收货"],
        }
        selected: list[str] = []
        missing: list[str] = []
        self._click(page, "purchase.po_status_field", "采购单状态")
        self._wait_quiet(page, 600)
        for status in requested:
            hit = ""
            for candidate in aliases.get(status, [status]):
                hit = self._click_select_option_exact(page, candidate)
                if hit:
                    print(f"[猫超] 采购单状态已选择: {hit}")
                    selected.append(hit)
                    break
            if not hit:
                missing.append(status)
                print(f"[猫超] 采购单状态未选中: {status}")
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        if missing:
            print(f"[猫超] 采购单状态未选全: {missing}；已选 {selected}")
        if not selected:
            raise RuntimeError(f"采购单状态下拉中一个也没选到: {requested}")
        print(f"[猫超] 采购单状态已按原文多选: {' / '.join(selected)}")

    def _clear_po_list_statuses(self, page: Any) -> bool:
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const fields = Array.from(document.querySelectorAll(
            '.next-select-trigger.next-select-multiple'
          )).filter(visible);
          const field = fields.find((el) => {
            const input = el.querySelector(
              "input[groupname*='purchase.order.bizStatus']"
            );
            const text = (el.innerText || el.textContent || '').replace(/\\s+/g, ' ');
            return !!input || /待供应商预约|供应商已确认|待收货|部分收货/.test(text);
          });
          if (!field) {
            return {count: 0, candidates: fields.length};
          }
          let count = 0;
          for (let attempt = 0; attempt < 20; attempt += 1) {
            const tags = Array.from(field.querySelectorAll(
              '.next-tag-close-btn'
            )).filter(visible);
            if (!tags.length) break;
            tags[tags.length - 1].click();
            count += 1;
          }
          return {count, candidates: fields.length};
        }
        """
        for scope in self._iter_scopes(page):
            try:
                result = scope.evaluate(script) or {}
            except Exception:
                continue
            cleared = int(result.get("count") or 0)
            if cleared:
                try:
                    page.keyboard.press("Escape")
                except Exception:
                    pass
                print(f"[猫超] 已清空采购单状态 {cleared} 项")
                return True
        print("[猫超] 未找到可清空的采购单状态字段")
        return False

    def _task_channel_goods(self, page: Any, account: Account) -> list[RunResult]:
        self._open_task_page(page, "channel-goods", (
            "channel_goods.menu_goods",
            "channel_goods.menu_channel_goods",
        ))
        supplier = self._current_supplier
        supplier_id = supplier.supplier_id if supplier is not None else ""
        item_ids = self._item_ids_by_supplier.get((account.key, supplier_id), [])
        batches = [item_ids[index:index + 30] for index in range(0, len(item_ids), 30)] or [[]]
        results: list[RunResult] = []
        try:
            for batch_index, batch in enumerate(batches, start=1):
                if batch:
                    self._set_channel_goods_item_ids(page, batch)
                    print(f"[猫超] 库位明细货品 ID 分批: {batch_index}/{len(batches)}，{len(batch)} 个")
                self._click_optional(page, "channel_goods.filter_button", "查询")
                self._wait_quiet(page, 3000)
                self._click_toolbar_export(
                    page,
                    option_texts=["导出明细", "导出全部", "导出当前页"],
                    allow_direct=True,
                )
                try:
                    results.append(self._download_and_clean(
                        page,
                        "channel-goods",
                        account,
                        prefix_extra=f"{self._supplier_prefix()}_batch{batch_index:02d}" if len(batches) > 1 else "",
                        note=f"商品→渠道货品→货品ID筛选→导出（{batch_index}/{len(batches)}）" if batch else "商品→渠道货品→查询→全量导出",
                    ))
                except RuntimeError as exc:
                    if self._is_null_download_error(exc) and batch:
                        print(f"[猫超] 库位明细第 {batch_index} 批无数据，继续下一批")
                        continue
                    raise
        except Exception as exc:
            self._remove_partial_cleaned_files(results)
            if self._is_null_download_error(exc) and not results:
                return [self._no_data_result(
                    "channel-goods",
                    account,
                    f"10、库位明细 平台未生成下载文件，已跳过: {exc}",
                )]
            raise
        if not results:
            return [self._no_data_result("channel-goods", account, "10、库位明细所有货品 ID 分批均无数据")]
        if len(results) == 1:
            return results
        return [self._merge_channel_goods_results(results, account)]

    def _set_channel_goods_item_ids(self, page: Any, item_ids: list[str]) -> None:
        value = ",".join(str(item_id).strip() for item_id in item_ids if str(item_id).strip())
        if not value:
            return
        script = """
        (value) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.display !== 'none' && style.visibility !== 'hidden';
          };
          const fields = Array.from(document.querySelectorAll('input, textarea')).filter(visible);
          const field = fields.find((el) => {
            const box = el.closest('.next-form-item, .form-item, [class*="formItem"], [class*="form-item"]');
            const text = (box?.innerText || el.parentElement?.innerText || '').replace(/\\s+/g, ' ');
            const hint = `${el.getAttribute('placeholder') || ''} ${el.getAttribute('name') || ''}`;
            return /货品\\s*ID/i.test(`${text} ${hint}`);
          });
          if (!field) return false;
          const proto = field.tagName === 'TEXTAREA' ? HTMLTextAreaElement.prototype : HTMLInputElement.prototype;
          const setter = Object.getOwnPropertyDescriptor(proto, 'value')?.set;
          if (setter) setter.call(field, value); else field.value = value;
          field.dispatchEvent(new Event('input', {bubbles: true}));
          field.dispatchEvent(new Event('change', {bubbles: true}));
          field.blur();
          return true;
        }
        """
        for scope in self._iter_scopes(page):
            try:
                if scope.evaluate(script, value):
                    return
            except Exception:
                continue
        raise RuntimeError("库位明细页面未找到货品 ID 筛选框")

    def _remove_partial_cleaned_files(self, results: list[RunResult]) -> None:
        for result in results:
            path = Path(result.cleaned_file) if result.cleaned_file else None
            if path is not None and path.is_file():
                try:
                    path.unlink()
                except OSError:
                    pass

    def _merge_channel_goods_results(self, results: list[RunResult], account: Account) -> RunResult:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError as exc:
            self._remove_partial_cleaned_files(results)
            raise RuntimeError("合并库位明细需要 openpyxl") from exc
        source_paths = [Path(item.cleaned_file) for item in results if item.cleaned_file]
        if len(source_paths) != len(results) or any(not path.is_file() for path in source_paths):
            self._remove_partial_cleaned_files(results)
            raise RuntimeError("库位明细分批文件不完整，不生成合并文件")
        merged = Workbook()
        merged.remove(merged.active)
        seen_by_sheet: dict[str, set[tuple[Any, ...]]] = {}
        try:
            for source_path in source_paths:
                source = load_workbook(source_path, read_only=True, data_only=True)
                for source_sheet in source.worksheets:
                    if source_sheet.title not in merged.sheetnames:
                        target = merged.create_sheet(source_sheet.title)
                        header = [cell.value for cell in source_sheet[1]] if source_sheet.max_row else []
                        if header:
                            target.append(header)
                        seen_by_sheet[source_sheet.title] = set()
                    target = merged[source_sheet.title]
                    seen = seen_by_sheet[source_sheet.title]
                    for values in source_sheet.iter_rows(min_row=2, values_only=True):
                        key = tuple(values)
                        if not any(value not in (None, "") for value in key) or key in seen:
                            continue
                        seen.add(key)
                        target.append(list(values))
                source.close()
            if not merged.sheetnames:
                raise RuntimeError("库位明细分批文件中没有可合并的工作表")
            _, cleaned_dir = self._account_data_dirs(account)
            cleaned_dir.mkdir(parents=True, exist_ok=True)
            target_path = self._unique_path(cleaned_dir / f"{TASKS['channel-goods']['prefix']}_{self._supplier_prefix()}_merged.xlsx")
            merged.save(target_path)
        except Exception:
            self._remove_partial_cleaned_files(results)
            raise
        self._remove_partial_cleaned_files(results)
        final = results[0]
        final.cleaned_file = str(target_path)
        final.note = f"货品 ID 分 {len(results)} 批导出并合并，已去除重复数据"
        final.finished_at = datetime.now().isoformat(timespec="seconds")
        return final

    def _task_transfer_order(self, page: Any, account: Account) -> list[RunResult]:
        self._open_task_page(page, "transfer-order", (
            "purchase.menu_purchase",
            "transfer_order.menu_transfer_order",
        ), force=True)
        self._reset_transfer_filters(page)
        count = self._visible_result_count(page)
        if count is not None:
            print(f"[猫超] 调拨单查询结果: 共 {count} 项")
            if count == 0:
                return [self._no_data_result("transfer-order", account, "调拨单无数据，未生成下载文件")]
        elif self._page_has_no_items(page):
            return [self._no_data_result("transfer-order", account, "调拨单无数据，未生成下载文件")]
        export_created = self._click_toolbar_export(
            page,
            option_texts=["导出货品明细", "导出明细", "调拨明细数据导出"],
            allow_direct=True,
            file_task_key="transfer-order",
            file_task_timeout_sec=25,
        )
        if not export_created:
            if self._page_has_no_items(page):
                return [self._no_data_result("transfer-order", account, "调拨单无数据，未生成下载文件")]
            raise RuntimeError("调拨单导出未生成新文件任务")
        try:
            return [self._download_and_clean(page, "transfer-order", account)]
        except RuntimeError as exc:
            if self._is_null_download_error(exc):
                return [self._no_data_result("transfer-order", account, f"调拨单平台未生成下载文件，已跳过: {exc}")]
            raise

    def _open_purchase_replenishment(self, page: Any, task_key: str = "system-order", force: bool = False) -> None:
        self._open_task_page(
            page,
            task_key,
            (
                "purchase.menu_purchase",
                "purchase.menu_replenishment_order",
            ),
            force=force,
        )

    def _open_task_page(self, page: Any, task_key: str, menu_selectors: tuple[str, ...], force: bool = False) -> None:
        self._dismiss_notification_center(page)
        self._dismiss_blocking_popups(page)
        frame_hint = TASK_FRAME_HINTS.get(task_key, "")
        if force:
            direct_url = self.settings.direct_urls.get(task_key, "")
            if direct_url:
                print(f"[猫超] 强制重开直达 URL: {TASKS[task_key]['title']}")
                page.goto(direct_url, wait_until="domcontentloaded")
                self._wait_quiet(page, 8000)
                self._dismiss_blocking_popups(page)
                if not frame_hint or self._wait_frame_url_contains(page, frame_hint, timeout_ms=8000):
                    return
        if frame_hint and self._visible_frame_url_contains(page, frame_hint):
            print(f"[猫超] 已在目标页: {TASKS[task_key]['title']}")
            return
        leaf = self._menu_text_hint(menu_selectors[-1]) if menu_selectors else ""
        if leaf and (self._click_sidebar_link(page, leaf) or self._search_and_open_menu(page, leaf)):
            print(f"[猫超] 已打开侧栏/搜索菜单: {leaf}")
            self._wait_quiet(page, 5000)
            self._dismiss_blocking_popups(page)
            if not frame_hint or self._wait_frame_url_contains(page, frame_hint, timeout_ms=8000):
                return
        direct_url = self.settings.direct_urls.get(task_key, "")
        if direct_url:
            print(f"[猫超] 打开直达 URL: {TASKS[task_key]['title']}")
            page.goto(direct_url, wait_until="domcontentloaded")
            self._wait_quiet(page, 8000)
            self._dismiss_blocking_popups(page)
            if not frame_hint or self._wait_frame_url_contains(page, frame_hint, timeout_ms=8000):
                return
        if menu_selectors and not self._menu_selector_visible(page, menu_selectors[0]):
            print(f"[猫超] 当前页找不到 {menu_selectors[0]}，先回工作台首页再切模块")
            self._goto_workbench_home(page)
            if leaf and (self._click_sidebar_link(page, leaf) or self._search_and_open_menu(page, leaf)):
                print(f"[猫超] 回首页后已打开: {leaf}")
                self._wait_quiet(page, 5000)
                self._dismiss_blocking_popups(page)
                return
        for idx, selector_key in enumerate(menu_selectors):
            self._dismiss_blocking_popups(page)
            try:
                self._reveal_top_menu(page)
                self._click_menu(page, selector_key)
                self._wait_quiet(page, 5000)
            except Exception as exc:
                if idx == 0:
                    print(f"[猫超] 顶部菜单未点到 {selector_key}，回首页后重试: {exc}")
                    self._goto_workbench_home(page)
                    self._reveal_top_menu(page)
                    try:
                        self._click_menu(page, selector_key)
                        self._wait_quiet(page, 5000)
                        continue
                    except Exception:
                        pass
                if idx + 1 < len(menu_selectors):
                    next_selector = self._selector_optional(menu_selectors[idx + 1])
                    if next_selector and self._visible_locator(page, next_selector, menu_selectors[idx + 1], timeout=1000):
                        continue
                raise RuntimeError(f"打开{TASKS[task_key]['title']}失败: {selector_key}: {exc}") from exc

    def _menu_text_hint(self, selector_key: str) -> str:
        return {
            "channel_goods.menu_goods": "商品",
            "channel_goods.menu_channel_goods": "渠道货品",
            "purchase.menu_purchase": "采购",
            "purchase.menu_replenishment_order": "补货单",
            "pincang.menu_tianji": "天机",
            "pincang.tab_pincang_detail": "品仓明细",
            "pincang.menu_inventory_analysis": "库存分析",
            "transfer_order.menu_transfer_order": "调拨单",
            "realtime.menu_inventory": "库存",
            "realtime.menu_inventory_query": "库存查询",
        }.get(selector_key, "")

    def _click_sidebar_link(self, page: Any, text: str) -> bool:
        target = _clean_text(text)
        if not target:
            return False
        script = """
        (target) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const inChrome = (el) => !!el.closest('header, .ascp-frame-header, .header-right, .search-wrapper');
          const clickTarget = (el) => el.closest(
            'a, button, [role="link"], [role="menuitem"], .next-menu-item, li, [class*="menu-item"], [class*="nav-item"]'
          ) || el;
          const nodes = Array.from(document.querySelectorAll(
            '.sidebar a, .sidebar button, .sidebar span, .sidebar div, .sidebar li, .sidebar-wrap a, .sidebar-wrap button, .sidebar-wrap span, .sidebar-wrap div, .sidebar-wrap li, aside a, aside button, aside span, aside div, aside li, .next-shell-aside a, .next-shell-aside button, .next-shell-aside span, .next-shell-aside div, .next-shell-aside li, [class*="side"] a, [class*="side"] button, [class*="side"] span, [class*="side"] div, [class*="side"] li'
          ));
          let best = null;
          let bestScore = 1e9;
          for (const el of nodes) {
            if (!visible(el) || inChrome(el)) continue;
            const label = textOf(el);
            if (label !== target) continue;
            const targetEl = clickTarget(el);
            const rect = targetEl.getBoundingClientRect();
            if (rect.left > 360 || rect.top < 40) continue;
            let score = el.querySelectorAll('*').length;
            score += rect.left;
            if (score < bestScore) {
              best = targetEl;
              bestScore = score;
            }
          }
          if (!best) return '';
          best.click();
          return target;
        }
        """
        for scope in self._iter_scopes(page):
            try:
                hit = scope.evaluate(script, target)
            except Exception:
                continue
            if hit:
                print(f"[猫超] 已点击侧栏: {target}")
                return True
        return False

    def _search_and_open_menu(self, page: Any, text: str) -> bool:
        target = _clean_text(text)
        if not target:
            return False
        script = """
        (target) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const inputs = Array.from(document.querySelectorAll('input, textarea'));
          const box = inputs.find((el) => {
            if (!visible(el)) return false;
            const ph = (el.getAttribute('placeholder') || '') + (el.getAttribute('aria-label') || '');
            return /搜索/.test(ph);
          });
          if (!box) return 'missing';
          box.focus();
          const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
          setter.call(box, '');
          box.dispatchEvent(new Event('input', { bubbles: true }));
          setter.call(box, target);
          box.dispatchEvent(new Event('input', { bubbles: true }));
          box.dispatchEvent(new KeyboardEvent('keydown', { key: 'Enter', bubbles: true }));
          return 'typed';
        }
        """
        typed = False
        for scope in self._iter_scopes(page):
            try:
                result = scope.evaluate(script, target)
            except Exception:
                continue
            if result == "typed":
                typed = True
                break
        if not typed:
            return False
        self._wait_quiet(page, 1500)
        if self._js_click_matching_text(page, [target], overlay_only=True, exact=True):
            print(f"[猫超] 已从搜索结果打开: {target}")
            return True
        try:
            page.keyboard.press("Enter")
            self._wait_quiet(page, 1200)
        except Exception:
            pass
        if self._js_click_matching_text(page, [target], overlay_only=True, exact=True):
            print(f"[猫超] 已从搜索结果打开: {target}")
            return True
        if self._js_click_matching_text(page, [target], overlay_only=False, exact=True):
            print(f"[猫超] 已从搜索结果打开: {target}")
            return True
        return False

    def _top_nav_spec(self, selector_key: str) -> tuple[str, str] | None:
        return {
            "channel_goods.menu_goods": ("745311", "商品"),
            "pincang.menu_tianji": ("745317", "天机"),
            "purchase.menu_purchase": ("745314", "采购"),
            "realtime.menu_inventory": ("745316", "库存"),
        }.get(selector_key)

    def _top_nav_selector(self, selector_key: str) -> str:
        spec = self._top_nav_spec(selector_key)
        if spec is None:
            return ""
        data_id, _ = spec
        return f'a.nav-item[data-id="{data_id}"], .nav-more a[data-id="{data_id}"], li.auto-more a[data-id="{data_id}"]'

    def _menu_selector_visible(self, page: Any, selector_key: str) -> bool:
        nav = self._top_nav_selector(selector_key)
        if nav:
            return self._visible_locator(page, nav, selector_key, timeout=800) is not None
        if self._selector_visible(page, selector_key, timeout=1500):
            return True
        hint = self._menu_text_hint(selector_key)
        if not hint:
            return False
        return self._visible_locator(page, f"a.nav-item:has-text(\"{hint}\")", hint, timeout=800) is not None

    def _click_menu(self, page: Any, selector_key: str) -> None:
        hint = self._menu_text_hint(selector_key)
        if hint and self._click_sidebar_link(page, hint):
            print(f"[猫超] 已点击侧栏菜单: {hint}")
            return
        spec = self._top_nav_spec(selector_key)
        if spec is not None:
            data_id, label = spec
            self._reveal_top_menu(page)
            locator = self._visible_locator(page, self._top_nav_selector(selector_key), label, timeout=4000)
            if locator is None:
                self._reveal_top_menu(page)
                locator = self._visible_locator(page, self._top_nav_selector(selector_key), label, timeout=4000)
            if locator is None:
                if self._search_and_open_menu(page, label):
                    print(f"[猫超] 已搜索打开菜单: {label}")
                    return
                raise RuntimeError(f"找不到顶部菜单 {label} (data-id={data_id})")
            if not self._js_click(locator):
                raise RuntimeError(f"点击顶部菜单失败 {label} (data-id={data_id})")
            print(f"[猫超] 已点击顶部菜单: {label} ({data_id})")
            return
        if selector_key == "pincang.tab_pincang_detail" or hint == "品仓明细":
            self._click_page_tab(page, hint or "品仓明细", sibling_hints=["概览指标", "诊断建议", "库龄分析"])
            return
        if hint and self._click_exact_control(page, hint, timeout=2500):
            print(f"[猫超] 已按文字点击菜单: {hint}")
            return
        try:
            self._click(page, selector_key, selector_key)
            return
        except Exception:
            if hint and self._click_text(page, hint, timeout=4000, optional=True):
                print(f"[猫超] 已按包含文字点击菜单: {hint}")
                return
            raise

    def _click_page_tab(self, page: Any, text: str, sibling_hints: list[str] | None = None) -> None:
        target = _clean_text(text)
        hints = [_clean_text(item) for item in (sibling_hints or []) if _clean_text(item)]
        tab_selector = "[role='tab'], .next-tabs-tab, li.next-tabs-tab"
        script = """
        ({target, hints}) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const active = (el) => {
            const names = `${el.className || ''} ${(el.parentElement && el.parentElement.className) || ''}`;
            return el.getAttribute('aria-selected') === 'true'
              || /(^|\\s)(active|selected)(\\s|$)/.test(names)
              || names.includes('next-tabs-tab-active');
          };
          const nodes = Array.from(document.querySelectorAll('[role="tab"], .next-tabs-tab, li.next-tabs-tab'));
          let best = null;
          let bestScore = -1;
          for (const el of nodes) {
            if (!visible(el) || textOf(el) !== target) continue;
            const rect = el.getBoundingClientRect();
            const area = rect.width * rect.height;
            const siblings = Array.from(el.parentElement ? el.parentElement.children : []);
            const siblingText = siblings.map(textOf).join(' ');
            const inTabRow = hints.some((hint) => siblingText.includes(hint));
            let score = area;
            if (inTabRow) score += 100000;
            if (el.classList.contains('active') || el.getAttribute('aria-selected') === 'true') score += 10;
            if (score > bestScore) {
              best = el;
              bestScore = score;
            }
          }
          if (!best) return {ok: false};
          return {ok: true, index: nodes.indexOf(best), active: active(best), score: bestScore};
        }
        """
        for attempt in range(1, 4):
            clicked = False
            for scope in self._iter_scopes(page):
                try:
                    result = scope.evaluate(script, {"target": target, "hints": hints})
                except Exception:
                    continue
                if not result or not result.get("ok"):
                    continue
                if result.get("active"):
                    print(f"[猫超] 页签已激活: {target}")
                    return
                index = int(result["index"])
                try:
                    locator = scope.locator(tab_selector).nth(index)
                    locator.scroll_into_view_if_needed(timeout=1500)
                    locator.click(timeout=2500, force=attempt > 1)
                    clicked = True
                except Exception:
                    try:
                        clicked = bool(
                            scope.evaluate(
                                """
                                (index) => {
                                  const tabs = Array.from(document.querySelectorAll(
                                    '[role="tab"], .next-tabs-tab, li.next-tabs-tab'
                                  ));
                                  const tab = tabs[index];
                                  if (!tab) return false;
                                  const target = tab.querySelector('.next-tabs-tab-inner, [class*="tab-inner"]') || tab;
                                  target.scrollIntoView({block: 'center', inline: 'nearest'});
                                  const rect = target.getBoundingClientRect();
                                  const options = {
                                    bubbles: true, cancelable: true, view: window,
                                    clientX: rect.left + rect.width / 2,
                                    clientY: rect.top + rect.height / 2,
                                  };
                                  for (const type of ['pointerdown', 'mousedown', 'pointerup', 'mouseup', 'click']) {
                                    target.dispatchEvent(new MouseEvent(type, options));
                                  }
                                  return true;
                                }
                                """,
                                index,
                            )
                        )
                    except Exception:
                        clicked = False
                if clicked:
                    break
            if not clicked:
                continue
            deadline = time.time() + 4
            while time.time() < deadline:
                for scope in self._iter_scopes(page):
                    try:
                        result = scope.evaluate(script, {"target": target, "hints": hints})
                    except Exception:
                        continue
                    if result and result.get("active"):
                        self._wait_quiet(page, 800)
                        print(f"[猫超] 已切换页签: {target}")
                        return
                time.sleep(0.25)
            print(f"[猫超] 页签未激活，重试切换 ({attempt}/3): {target}")
        raise RuntimeError(f"页签未切换至目标状态: {target}")

    def _overlay_item_selectors(self) -> tuple[str, ...]:
        return (
            ".next-overlay-wrapper.opened .next-menu-item",
            ".next-overlay-wrapper:visible .next-menu-item",
            ".next-overlay-wrapper:visible [role='menuitem']",
            ".next-overlay-wrapper:visible [role='option']",
            ".next-overlay-wrapper:visible li",
            ".next-menu:visible .next-menu-item",
            "[role='menu']:visible [role='menuitem']",
            ".next-balloon:visible .next-menu-item",
            ".next-overlay-inner:visible .next-menu-item",
        )

    def _visible_overlay_labels(self, page: Any) -> list[str]:
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const labels = [];
          const seen = new Set();
          const nodes = document.querySelectorAll(
            '.next-overlay-wrapper.opened .next-menu-item, .next-overlay-wrapper .next-menu-item, .next-menu .next-menu-item, [role="menu"] [role="menuitem"], [role="listbox"] [role="option"]'
          );
          for (const el of nodes) {
            if (!visible(el)) continue;
            const label = textOf(el);
            if (!label || seen.has(label) || label.length > 40) continue;
            seen.add(label);
            labels.push(label);
          }
          return labels;
        }
        """
        labels: list[str] = []
        seen: set[str] = set()
        for scope in self._iter_scopes(page):
            try:
                values = scope.evaluate(script) or []
            except Exception:
                continue
            for label in values:
                text = _clean_text(label)
                if not text or text in seen:
                    continue
                seen.add(text)
                labels.append(text)
        return labels

    def _click_overlay_option(self, page: Any, option_texts: list[str], timeout: int = 5000) -> str:
        deadline = time.time() + timeout / 1000
        while time.time() < deadline:
            hit = self._js_click_matching_text(page, option_texts, overlay_only=True, exact=True)
            if hit:
                self._wait_quiet(page, 400)
                return hit
            time.sleep(0.2)
        return ""

    def _confirm_export_dialog(self, page: Any) -> None:
        self._wait_quiet(page, 800)
        labels = self._visible_overlay_labels(page)
        if labels:
            print(f"[猫超] 导出后弹层: {' / '.join(labels[:8])}")
        toasts = self._visible_toast_texts(page)
        if toasts:
            print(f"[猫超] 导出后提示: {' / '.join(toasts[:6])}")
        overlay_text = self._opened_overlay_text(page)
        if overlay_text:
            print(f"[猫超] 导出后遮罩文本: {overlay_text[:180]}")
        for text in ("确定", "确认", "开始导出", "知道了"):
            if self._click_exact_control(page, text, timeout=1200, overlay_only=True):
                print(f"[猫超] 已确认导出弹窗: {text}")
                return

    def _opened_overlay_text(self, page: Any) -> str:
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 40 && rect.height > 20 &&
              style.visibility !== 'hidden' && style.display !== 'none';
          };
          const nodes = Array.from(document.querySelectorAll(
            '.next-overlay-wrapper.opened, .next-dialog, .next-message, [role="dialog"]'
          )).filter(visible);
          return nodes.map((el) => (el.innerText || '').replace(/\\s+/g, ' ').trim())
            .filter((text) => text && text.length < 200)
            .slice(0, 4)
            .join(' || ');
        }
        """
        texts: list[str] = []
        for scope in self._iter_scopes(page):
            try:
                text = _clean_text(scope.evaluate(script) or "")
            except Exception:
                continue
            if text:
                texts.append(text)
        return " | ".join(texts[:3])

    def _visible_toast_texts(self, page: Any) -> list[str]:
        script = """
        () => {
          const nodes = Array.from(document.querySelectorAll(
            '.next-message, .next-notice, [role="alert"], .next-balloon-content, .next-feedback'
          ));
          return nodes.map((el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim())
            .filter((text) => text && text.length < 80);
        }
        """
        texts: list[str] = []
        seen: set[str] = set()
        for scope in self._iter_scopes(page):
            try:
                values = scope.evaluate(script) or []
            except Exception:
                continue
            for value in values:
                text = _clean_text(value)
                if not text or text in seen:
                    continue
                seen.add(text)
                texts.append(text)
        return texts

    def _click_gei_download_icon(self, page: Any, toolbar_title: str = "") -> bool:
        target = _clean_text(toolbar_title)
        script = """
        (toolbarTitle) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const inChrome = (el) => !!el.closest('header, .ascp-frame-header, .header-right, .river-header');
          const titles = Array.from(document.querySelectorAll('.comp-toolbar-title-text'));
          const needles = toolbarTitle ? [toolbarTitle, '品仓明细表', '品仓明细'] : [];
          const hit = titles.find((el) => ['品仓明细表', '品仓明细'].includes(textOf(el)))
            || (needles.length ? titles.find((el) => needles.some((n) => textOf(el).includes(n))) : null);
          const downloadSelector = [
            '.toolbar-func-download',
            '.toolbar-func-export',
            '[data-tip*="下载"]',
            '[data-tip*="导出"]',
            '[title*="下载"]',
            '[title*="导出"]',
            '[aria-label*="下载"]',
            '[aria-label*="导出"]',
            '.toolbar-gei-export-button-wrapper',
            '[class*="download"]',
            '[class*="export"]'
          ].join(',');
          const actionable = (el) => {
            if (!visible(el) || inChrome(el)) return false;
            const rect = el.getBoundingClientRect();
            const small = rect.width <= 260 && rect.height <= 90;
            return small && (
              el.matches('button, a, span, i, [role="button"]') ||
              !!el.closest('.comp-toolbar, [class*="toolbar"]')
            );
          };
          const clickTarget = (node, forceCenter = false) => {
            let target = node;
            for (let i = 0; target && i < 8; i++, target = target.parentElement) {
              if (actionable(target)) break;
            }
            if (!target || !actionable(target)) {
              if (!forceCenter) return false;
              target = node;
              for (let i = 0; target && i < 8; i++, target = target.parentElement) {
                if (visible(target) && !inChrome(target)) break;
              }
              if (!target || !visible(target) || inChrome(target)) return false;
            }
            target.scrollIntoView({block: 'center', inline: 'nearest'});
            const rect = target.getBoundingClientRect();
            const options = {
              bubbles: true, cancelable: true, view: window,
              clientX: rect.left + rect.width / 2,
              clientY: rect.top + rect.height / 2,
            };
            for (const type of ['pointerdown', 'mousedown', 'pointerup', 'mouseup']) {
              target.dispatchEvent(new MouseEvent(type, options));
            }
            target.click();
            return true;
          };
          const pincangTitle = (el) => {
            const text = textOf(el);
            return /^品仓明细表(?:\\s*[（(]\\s*[0-9,]+\\s*[）)])?/.test(text)
              && text.length < 220;
          };
          const sectionTitle = Array.from(document.querySelectorAll('*'))
            .filter((el) => visible(el) && pincangTitle(el))
            .sort((a, b) => {
              const ar = a.getBoundingClientRect();
              const br = b.getBoundingClientRect();
              return (ar.width * ar.height) - (br.width * br.height);
            })[0] || null;
          if (sectionTitle) {
            const titleRect = sectionTitle.getBoundingClientRect();
            let root = sectionTitle;
            for (let depth = 0; root && depth < 12; depth++, root = root.parentElement) {
              const controls = Array.from(root.querySelectorAll(downloadSelector))
                .filter(actionable);
              if (!controls.length) continue;
              const ranked = controls.map((btn) => {
                const rect = btn.getBoundingClientRect();
                const verticalDistance = Math.abs(rect.top - titleRect.top);
                let score = 100 - Math.min(verticalDistance, 100);
                if (rect.top >= titleRect.top - 120) score += 25;
                if (rect.left >= titleRect.left) score += 10;
                if (btn.matches('.toolbar-func-download, .toolbar-gei-export-button-wrapper')) score += 20;
                return {btn, score};
              }).sort((a, b) => b.score - a.score);
              if (ranked.length && clickTarget(ranked[0].btn)) {
                return 'pincang-section';
              }
            }
          }
          const pincangXPath = '/html/body/div[7]/div[2]/div/div/div/div[2]/div[7]/div/div[3]/div/div/div[1]/div[2]/div/div/span/span/div/span/svg/use';
          try {
            const result = document.evaluate(
              pincangXPath,
              document,
              null,
              XPathResult.FIRST_ORDERED_NODE_TYPE,
              null
            );
            const node = result.singleNodeValue;
            if (node && clickTarget(node, true)) return 'pincang-xpath';
          } catch (error) {
            // The absolute XPath is a last-resort anchor and may not exist in every layout.
          }
          const toolbars = [];
          if (hit) {
            hit.scrollIntoView({block: 'center', inline: 'nearest'});
            const toolbar = hit.closest('.comp-toolbar');
            if (toolbar) toolbars.push(toolbar);
          } else {
            for (const el of document.querySelectorAll('.comp-toolbar')) {
              if (visible(el)) toolbars.push(el);
            }
          }
          for (const toolbar of toolbars) {
            const btn = Array.from(toolbar.querySelectorAll(downloadSelector)).find(actionable);
            if (!btn) continue;
            if (!clickTarget(btn)) continue;
            return textOf(toolbar.querySelector('.comp-toolbar-title-text') || toolbar).slice(0, 30) || 'download-icon';
          }
          const candidates = [];
          for (const btn of document.querySelectorAll(downloadSelector)) {
            if (!actionable(btn)) continue;
            const rect = btn.getBoundingClientRect();
            const root = btn.closest('.comp-toolbar, [class*="toolbar"], [class*="table"], [class*="Table"], [class*="comp-container"]');
            const rootText = textOf(root || btn.parentElement || btn);
            let score = 0;
            if (toolbarTitle && rootText.includes(toolbarTitle)) score += 80;
            if (rootText.includes('品仓明细表') || rootText.includes('品仓明细')) score += 80;
            const ident = `${btn.className || ''} ${btn.getAttribute('data-tip') || ''} ${btn.getAttribute('title') || ''} ${btn.getAttribute('aria-label') || ''}`;
            if (/toolbar-func-download|icon-download|下载/.test(ident)) score += 35;
            if (/toolbar|export|download|下载|导出/.test(ident)) score += 20;
            score += Math.max(rect.left, 0) / Math.max(window.innerWidth, 1);
            candidates.push({btn, score, label: rootText.slice(0, 30) || textOf(btn).slice(0, 30) || 'download-icon'});
          }
          candidates.sort((a, b) => b.score - a.score);
          if (candidates.length) {
            const best = candidates[0];
            if (clickTarget(best.btn)) return best.label;
          }
          return '';
        }
        """
        for scope in self._iter_scopes(page):
            try:
                hit = scope.evaluate(script, target)
            except Exception:
                continue
            if hit:
                print(f"[猫超] 已点击表格下载图标: {hit}")
                return True
        return False

    def _click_toolbar_export_button(self, page: Any, prefer_arrow: bool) -> bool:
        script = """
        (preferArrow) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const inChrome = (el) => !!el.closest('header, .ascp-frame-header, .header-right, .river-header');
          const nodes = Array.from(document.querySelectorAll('button, [role="button"], .next-btn, .next-menu-btn'));
          const scored = [];
          for (const el of nodes) {
            if (!visible(el) || inChrome(el)) continue;
            const text = textOf(el);
            if (text !== '导出') continue;
            let score = 10;
            let ancestor = el.parentElement;
            for (let i = 0; i < 8 && ancestor; i++) {
              if (/共\\s*[0-9,]+\\s*项/.test(textOf(ancestor))) {
                score += 30;
                break;
              }
              ancestor = ancestor.parentElement;
            }
            if (el.closest('.next-table-wrapper, .next-table, .river-page, .next-box')) score += 8;
            score += el.getBoundingClientRect().x / Math.max(window.innerWidth, 1);
            scored.push(el);
            el.__exportScore = score;
          }
          scored.sort((a, b) => (b.__exportScore || 0) - (a.__exportScore || 0));
          if (!scored.length) return {ok: false};
          const best = scored[0];
          const ownArrow = best.querySelector('.next-icon-arrow-down, .next-icon-arrow-down-filling');
          if (ownArrow) {
            best.click();
            return {ok: true, via: 'button-with-icon'};
          }
          const sibling = best.nextElementSibling;
          if (preferArrow && sibling && visible(sibling) && (
            sibling.matches('button, a, [role="button"]') ||
            sibling.querySelector('.next-icon-arrow-down, .next-icon-arrow-down-filling')
          )) {
            sibling.click();
            return {ok: true, via: 'next-sibling-arrow'};
          }
          best.click();
          return {ok: true, via: 'button'};
        }
        """
        for scope in self._iter_scopes(page):
            try:
                result = scope.evaluate(script, prefer_arrow)
            except Exception:
                continue
            if result and result.get("ok"):
                print(f"[猫超] 已点击表格工具栏导出: {result.get('via')}")
                return True
        return False

    def _click_export_menu_item(self, page: Any, option_texts: list[str], native_first: bool = False) -> str:
        needles = [_clean_text(text) for text in option_texts if _clean_text(text)]
        if not needles:
            return ""
        pw_hit = self._playwright_click_export_menu(page, option_texts, native_first=native_first)
        if pw_hit:
            return pw_hit
        inspect_script = """
        (needles) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            const inView = rect.bottom > 0 && rect.right > 0 &&
              rect.top < window.innerHeight && rect.left < window.innerWidth;
            return inView && rect.width > 8 && rect.height > 8 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const ownText = (el) => Array.from(el.childNodes)
            .filter((n) => n.nodeType === 3)
            .map((n) => (n.textContent || '').replace(/\\s+/g, ' ').trim())
            .filter(Boolean)
            .join('');
          const nodes = Array.from(document.querySelectorAll(
            '.next-overlay-wrapper.opened .next-menu-item, .next-overlay-wrapper.opened [role="menuitem"], .next-menu .next-menu-item, [role="menu"] [role="menuitem"]'
          ));
          const scored = [];
          for (const el of nodes) {
            if (!visible(el)) continue;
            const label = ownText(el) || textOf(el);
            if (!label || label.length > 20) continue;
            const exact = needles.some((n) => label === n);
            const loose = needles.some((n) => label.includes(n));
            if (!exact && !loose) continue;
            const rect = el.getBoundingClientRect();
            const opened = !!(el.closest('.next-overlay-wrapper.opened'));
            scored.push({
              label,
              exact,
              opened,
              area: Math.max(rect.width, 0) * Math.max(rect.height, 0),
              top: Math.round(rect.top),
              left: Math.round(rect.left),
            });
          }
          scored.sort((a, b) => (Number(b.opened) - Number(a.opened)) || (Number(b.exact) - Number(a.exact)) || (b.area - a.area));
          return scored[0] || null;
        }
        """
        click_script = """
        (needles) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            const inView = rect.bottom > 0 && rect.right > 0 &&
              rect.top < window.innerHeight && rect.left < window.innerWidth;
            return inView && rect.width > 8 && rect.height > 8 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const ownText = (el) => Array.from(el.childNodes)
            .filter((n) => n.nodeType === 3)
            .map((n) => (n.textContent || '').replace(/\\s+/g, ' ').trim())
            .filter(Boolean)
            .join('');
          const mouseClick = (el) => {
            const target = el.querySelector('.next-menu-item-text, a, button') || el;
            target.scrollIntoView({block: 'nearest'});
            const rect = target.getBoundingClientRect();
            const opts = {
              bubbles: true, cancelable: true, view: window,
              clientX: rect.left + rect.width / 2,
              clientY: rect.top + rect.height / 2,
            };
            for (const type of ['mouseenter', 'mouseover', 'mousedown', 'mouseup', 'click']) {
              target.dispatchEvent(new MouseEvent(type, opts));
            }
            if (typeof target.click === 'function') target.click();
            if (target !== el && typeof el.click === 'function') el.click();
          };
          const nodes = Array.from(document.querySelectorAll(
            '.next-overlay-wrapper.opened .next-menu-item, .next-overlay-wrapper.opened [role="menuitem"], .next-menu .next-menu-item, [role="menu"] [role="menuitem"]'
          ));
          const scored = [];
          for (const el of nodes) {
            if (!visible(el)) continue;
            const label = ownText(el) || textOf(el);
            if (!label || label.length > 20) continue;
            const exact = needles.some((n) => label === n);
            const loose = needles.some((n) => label.includes(n));
            if (!exact && !loose) continue;
            const rect = el.getBoundingClientRect();
            const opened = !!(el.closest('.next-overlay-wrapper.opened'));
            scored.push({el, label, exact, opened, area: Math.max(rect.width, 0) * Math.max(rect.height, 0)});
          }
          scored.sort((a, b) => (Number(b.opened) - Number(a.opened)) || (Number(b.exact) - Number(a.exact)) || (b.area - a.area));
          if (!scored.length) return '';
          mouseClick(scored[0].el);
          return scored[0].label;
        }
        """
        best = None
        best_scope = None
        for scope in self._iter_scopes(page):
            try:
                info = scope.evaluate(inspect_script, needles)
            except Exception:
                continue
            if not info:
                continue
            score = (int(bool(info.get("opened"))), int(bool(info.get("exact"))), float(info.get("area") or 0))
            best_score = (
                int(bool((best or {}).get("opened"))),
                int(bool((best or {}).get("exact"))),
                float((best or {}).get("area") or 0),
            ) if best else (-1, -1, -1)
            if score > best_score:
                best = info
                best_scope = scope
        if not best or best_scope is None:
            return ""
        url = str(getattr(best_scope, "url", "") or "")[:90]
        print(
            f"[猫超] 导出菜单可见项: {best.get('label')} area={best.get('area')} "
            f"opened={best.get('opened')} @{int(best.get('left') or 0)},{int(best.get('top') or 0)} {url}"
        )
        label = str(best.get("label") or "")
        try:
            hit = best_scope.evaluate(click_script, needles)
        except Exception as exc:
            print(f"[猫超] 脚本点击导出菜单失败: {exc}")
            hit = ""
        if hit:
            print(f"[猫超] 已脚本点击导出菜单: {hit}")
            return str(hit)
        if label:
            try:
                loc = best_scope.get_by_text(label, exact=True)
                print(f"[猫超] get_by_text({label}) count={loc.count()}")
                loc.first.click(timeout=2500, delay=80)
                print(f"[猫超] 已鼠标点击导出菜单: {label}")
                return label
            except Exception as exc:
                print(f"[猫超] get_by_text点击失败: {exc}")
                try:
                    loc.first.click(timeout=2000, force=True)
                    print(f"[猫超] 已强制鼠标点击导出菜单: {label}")
                    return label
                except Exception as exc2:
                    print(f"[猫超] 强制点击也失败: {exc2}")
        pw_hit = self._playwright_click_export_menu(page, [label] + option_texts)
        if pw_hit:
            return pw_hit
        try:
            hit = best_scope.evaluate(click_script, needles)
        except Exception:
            hit = ""
        return str(hit or "")

    def _playwright_click_export_menu(
        self,
        page: Any,
        option_texts: list[str],
        native_first: bool = False,
    ) -> str:
        for scope in self._iter_scopes(page):
            for text in option_texts:
                target = _clean_text(text)
                if not target:
                    continue
                locators = []
                try:
                    locators.append(scope.locator(".next-overlay-wrapper.opened .next-menu-item").filter(has_text=target))
                except Exception:
                    pass
                try:
                    locators.append(scope.locator("[role='listbox'] [role='option']").filter(has_text=target))
                except Exception:
                    pass
                try:
                    locators.append(scope.get_by_role("menuitem", name=target, exact=True))
                except Exception:
                    pass
                try:
                    locators.append(scope.get_by_role("option", name=target, exact=True))
                except Exception:
                    pass
                for locator in locators:
                    try:
                        count = min(locator.count(), 8)
                        if count == 0:
                            continue
                        for idx in range(count):
                            item = locator.nth(idx)
                            if not self._element_is_displayed(item):
                                continue
                            try:
                                item.scroll_into_view_if_needed(timeout=1000)
                            except Exception:
                                pass
                            try:
                                box = item.bounding_box()
                            except Exception:
                                box = None
                            mask_state = self._disable_search_mask_pointer_events(scope)
                            try:
                                try:
                                    item.click(timeout=2500, delay=100)
                                    click_kind = "原生" if native_first else "去遮罩原生"
                                    print(f"[猫超] 已{click_kind}点击导出菜单: {target}")
                                    return target
                                except Exception as exc:
                                    print(f"[猫超] 原生点击导出菜单失败，改坐标点击: {exc}")
                                if box:
                                    page.mouse.click(
                                        box["x"] + box["width"] / 2,
                                        box["y"] + box["height"] / 2,
                                        delay=100,
                                    )
                                    print(f"[猫超] 已坐标点击导出菜单: {target}")
                                else:
                                    item.click(timeout=2000, force=True)
                                    print(f"[猫超] 已强制点击导出菜单: {target}")
                                return target
                            finally:
                                self._restore_search_mask_pointer_events(scope, mask_state)
                    except Exception:
                        continue
        return ""

    def _disable_search_mask_pointer_events(self, scope: Any) -> list[dict[str, str]]:
        script = """
        () => {
          const changes = [];
          document.querySelectorAll('.search-mask').forEach((el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            if (rect.width <= 100 || rect.height <= 100 || style.pointerEvents === 'none') return;
            const index = changes.length;
            changes.push({
              pointerEvents: el.style.getPropertyValue('pointer-events'),
              priority: el.style.getPropertyPriority('pointer-events'),
            });
            el.dataset.maochaoRpaMaskIndex = String(index);
            el.style.setProperty('pointer-events', 'none', 'important');
          });
          return changes;
        }
        """
        try:
            return scope.evaluate(script) or []
        except Exception:
            return []

    def _restore_search_mask_pointer_events(self, scope: Any, changes: list[dict[str, str]]) -> None:
        if not changes:
            return
        script = """
        (changes) => {
          document.querySelectorAll('.search-mask[data-maochao-rpa-mask-index]').forEach((el) => {
            const index = Number(el.dataset.maochaoRpaMaskIndex);
            const previous = changes[index];
            if (!previous) return;
            if (previous.pointerEvents) {
              el.style.setProperty('pointer-events', previous.pointerEvents, previous.priority || '');
            } else {
              el.style.removeProperty('pointer-events');
            }
            delete el.dataset.maochaoRpaMaskIndex;
          });
        }
        """
        try:
            scope.evaluate(script, changes)
        except Exception:
            pass

    def _has_new_file_task(self, page: Any, task_key: str = "", quiet: bool = False) -> bool:
        before = getattr(self, "_pre_export_file_task_ids", set()) or set()
        now = set()
        row_texts: dict[str, str] = {}
        script = """
        () => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const rowKey = (row) => {
            const id = row.id || row.getAttribute('data-id') || row.getAttribute('data-row-key') || '';
            if (id) return id;
            return `text:${textOf(row).slice(0, 220)}`;
          };
          const roots = Array.from(document.querySelectorAll(
            '.river-notification-center_notification, .notification-center, .notification-drawer-container, #notification-center'
          ));
          if (!roots.length) roots.push(document);
          const selectors = [
            'li[id^="fileTask"]',
            '[id^="fileTask"]',
            '[id*="fileTask"]',
            '[role="row"]',
            '.next-table-row',
            '.next-list-item',
            '.file-task-item',
            '[class*="fileTask"]',
            '[class*="file-task"]',
            '[class*="task-item"]'
          ];
          const values = new Set();
          for (const root of roots) {
            for (const selector of selectors) {
              for (const row of root.querySelectorAll(selector)) {
                const text = textOf(row);
                if (!text || text.length >= 1200 || !/导出|下载|文件|实时库存|PO明细|品仓|调拨|货品/.test(text)) continue;
                values.add(JSON.stringify({key: rowKey(row), text}));
              }
            }
          }
          return Array.from(values).map((item) => JSON.parse(item));
        }
        """
        for scope in self._iter_scopes(page):
            try:
                rows = scope.evaluate(script) or []
            except Exception:
                continue
            for row in rows:
                if not isinstance(row, dict):
                    continue
                key = str(row.get("key") or "")
                if not key:
                    continue
                now.add(key)
                row_texts[key] = str(row.get("text") or "")
        new_ids = now - before
        titles = self._list_file_task_titles(page)
        if not new_ids:
            if not quiet:
                print(f"[猫超] 点击后未见新文件任务，当前 {len(now)} 条")
                if titles:
                    print(f"[猫超] 当前文件任务: {' || '.join(titles[:6])}")
            return False
        if not quiet:
            print(f"[猫超] 点击后新增文件任务 {len(new_ids)} 条")
            if titles:
                print(f"[猫超] 当前文件任务: {' || '.join(titles[:8])}")
        if not task_key:
            return True
        needles = self._file_task_text_candidates(task_key, TASKS.get(task_key, {}).get("file_task_text", ""))
        if not needles:
            return True
        blob = " ".join(row_texts.get(key, "") for key in new_ids)
        return any(needle and needle in blob for needle in needles)

    def _wait_for_new_file_task(self, page: Any, task_key: str = "", timeout_sec: int = 10) -> bool:
        deadline = time.time() + timeout_sec
        while time.time() < deadline:
            last = (deadline - time.time()) <= 1.2
            if self._has_new_file_task(page, task_key, quiet=not last):
                print(f"[猫超] 已出现新文件任务{('：' + task_key) if task_key else ''}")
                return True
            time.sleep(0.8)
        print("[猫超] 等待新文件任务超时")
        return False

    def _click_toolbar_export(
        self,
        page: Any,
        option_texts: list[str] | None = None,
        allow_direct: bool = False,
        toolbar_title: str = "",
        file_task_key: str = "",
        file_task_timeout_sec: int = 10,
    ) -> bool:
        option_texts = [text for text in (option_texts or []) if _clean_text(text)]
        self._snapshot_file_task_ids(page)
        self._dismiss_notification_center(page)
        clicked = False
        if toolbar_title:
            clicked = self._click_gei_download_icon(page, toolbar_title)
        if not clicked:
            clicked = self._click_toolbar_export_button(page, prefer_arrow=bool(option_texts))
        if not clicked:
            clicked = self._click_gei_download_icon(page, toolbar_title)
        if not clicked:
            raise RuntimeError("找不到表格工具栏导出/下载按钮")
        self._wait_quiet(page, 800)
        labels = self._visible_overlay_labels(page)
        if labels:
            print(f"[猫超] 导出菜单项: {' / '.join(labels)}")
        if option_texts:
            chosen = self._click_export_menu_item(page, option_texts) or self._click_overlay_option(page, option_texts, timeout=2000)
            if chosen:
                print(f"[猫超] 已点击导出菜单: {chosen}")
                self._wait_quiet(page, 800)
                self._confirm_export_dialog(page)
                created = self._wait_for_new_file_task(
                    page,
                    file_task_key,
                    timeout_sec=file_task_timeout_sec,
                )
                if not created:
                    print("[猫超] 导出后未见新文件任务，关闭遮罩后重试一次")
                    self._dismiss_notification_center(page)
                    if not self._click_toolbar_export_button(page, prefer_arrow=True):
                        return False
                    self._wait_quiet(page, 800)
                    retry_labels = self._visible_overlay_labels(page)
                    if retry_labels:
                        print(f"[猫超] 重试导出菜单项: {' / '.join(retry_labels)}")
                    retry = self._click_export_menu_item(page, option_texts, native_first=True) or self._click_overlay_option(page, option_texts, timeout=3000)
                    if retry:
                        print(f"[猫超] 已重试导出菜单: {retry}")
                        self._wait_quiet(page, 800)
                        self._confirm_export_dialog(page)
                        created = self._wait_for_new_file_task(
                            page,
                            file_task_key,
                            timeout_sec=file_task_timeout_sec,
                        )
                return created
        if toolbar_title or allow_direct:
            if file_task_key and self._has_new_file_task(page, file_task_key):
                print("[猫超] 下载图标已直接生成文件任务，按直接导出处理")
                self._confirm_export_dialog(page)
                return True
            if allow_direct and not labels:
                print("[猫超] 导出菜单未出现，按直接导出处理")
                self._confirm_export_dialog(page)
                if file_task_key:
                    return self._wait_for_new_file_task(
                        page,
                        file_task_key,
                        timeout_sec=file_task_timeout_sec,
                    )
                return True
        if option_texts and not toolbar_title:
            print("[猫超] 导出菜单未展开，再点一次导出")
            self._click_toolbar_export_button(page, prefer_arrow=True)
            self._wait_quiet(page, 800)
            labels = self._visible_overlay_labels(page)
            if labels:
                print(f"[猫超] 导出菜单项: {' / '.join(labels)}")
            chosen = self._click_export_menu_item(page, option_texts) or self._click_overlay_option(page, option_texts, timeout=3000)
            if chosen:
                print(f"[猫超] 已点击导出菜单: {chosen}")
                self._wait_quiet(page, 800)
                self._confirm_export_dialog(page)
                return self._wait_for_new_file_task(
                    page,
                    file_task_key,
                    timeout_sec=file_task_timeout_sec,
                )
            raise RuntimeError(
                f"已点「导出」，但没有点到菜单项 {option_texts}；当前菜单: {labels or '无'}"
            )
        if option_texts:
            raise RuntimeError(
                f"已点「导出」，但没有点到菜单项 {option_texts}；当前菜单: {labels or '无'}"
            )
        self._confirm_export_dialog(page)
        if file_task_key:
            return self._wait_for_new_file_task(
                page,
                file_task_key,
                timeout_sec=file_task_timeout_sec,
            )
        return True

    def _overlay_has_option(self, labels: list[str], option_texts: list[str]) -> bool:
        targets = [_clean_text(text) for text in option_texts if _clean_text(text)]
        return any(any(target == label or target in label for target in targets) for label in labels)

    def _goto_workbench_home(self, page: Any) -> None:
        self._dismiss_blocking_popups(page)
        if self._click_exact_control(page, "首页", timeout=1500) or self._click_text(page, "首页", timeout=1500, optional=True):
            self._wait_quiet(page, 2500)
            self._dismiss_blocking_popups(page)
            print("[猫超] 已点击首页回到工作台")
            return
        try:
            page.goto("https://web.txcs.tmall.com/", wait_until="domcontentloaded")
            self._wait_quiet(page, 3000)
            self._dismiss_blocking_popups(page)
            print("[猫超] 已打开工作台首页")
        except Exception as exc:
            print(f"[猫超] 返回工作台首页失败: {exc}")

    def _reveal_top_menu(self, page: Any) -> None:
        for selector in ("li.auto-more", "a:has-text(\"更多\")", "button:has-text(\"更多\")"):
            try:
                more = page.locator(selector).first
                if not more.count() or not self._element_is_displayed(more):
                    continue
                self._js_click(more)
                self._wait_quiet(page, 500)
                return
            except Exception:
                continue

    def _dismiss_blocking_popups(self, page: Any) -> None:
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass

        if self._click_close_button_by_script(page):
            self._wait_quiet(page, 800)
            print("[猫超] 已关闭遮挡弹窗。")
            return

        close_selectors = (
            ".next-overlay-wrapper:visible .next-dialog-close",
            ".next-overlay-wrapper:visible .next-overlay-close",
            ".next-overlay-wrapper:visible [aria-label=\"Close\"]",
            ".next-overlay-wrapper:visible [aria-label=\"关闭\"]",
            ".next-overlay-wrapper:visible button:has-text(\"取消\")",
            ".next-overlay-wrapper:visible button:has-text(\"我知道了\")",
            ".next-overlay-wrapper:visible button:has-text(\"×\")",
            ".next-overlay-wrapper:visible [role=\"button\"]:has-text(\"×\")",
            ".next-dialog:visible .next-dialog-close",
            ".next-dialog:visible [aria-label=\"Close\"]",
            ".next-dialog:visible [aria-label=\"关闭\"]",
            ".next-dialog:visible button:has-text(\"取消\")",
            ".next-dialog:visible button:has-text(\"我知道了\")",
            ".next-dialog:visible [class*=\"close\"]",
            ".ant-modal-root:visible .ant-modal-close",
            ".ant-modal:visible .ant-modal-close",
            "[role=\"dialog\"]:visible [aria-label=\"Close\"]",
            "[role=\"dialog\"]:visible [aria-label=\"关闭\"]",
            "[role=\"dialog\"]:visible button:has-text(\"取消\")",
            "[role=\"dialog\"]:visible button:has-text(\"我知道了\")",
            "[role=\"dialog\"]:visible button:has-text(\"×\")",
            "[class*=\"modal\"]:visible [class*=\"close\"]",
            "[class*=\"dialog\"]:visible [class*=\"close\"]",
            "[class*=\"popup\"]:visible [class*=\"close\"]",
        )
        for selector in close_selectors:
            locator = self._quick_visible_locator(page, selector, timeout=120)
            if locator is None:
                continue
            try:
                if self._js_click(locator):
                    self._wait_quiet(page, 800)
                    print("[猫超] 已关闭遮挡弹窗。")
                    return
            except Exception:
                continue

    def _click_close_button_by_script(self, page: Any) -> bool:
        script = (
            """
            () => {
              const visible = (el) => {
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return rect.width > 0 && rect.height > 0 &&
                  style.visibility !== 'hidden' && style.display !== 'none' &&
                  Number(style.opacity || 1) > 0;
              };
              const zIndex = (el) => {
                const raw = window.getComputedStyle(el).zIndex;
                const value = Number.parseInt(raw, 10);
                return Number.isFinite(value) ? value : 0;
              };
              const candidates = [];
              for (const el of document.querySelectorAll('button,[role="button"],a,span,div,i,svg')) {
                if (!visible(el)) continue;
                const rect = el.getBoundingClientRect();
                if (rect.top < 110) continue;
                const text = (el.innerText || el.textContent || '').trim();
                const label = [
                  el.getAttribute('aria-label') || '',
                  el.getAttribute('title') || '',
                  typeof el.className === 'string' ? el.className : ''
                ].join(' ').toLowerCase();
                const looksClose =
                  text === '×' || text === 'x' || text === 'X' || text === '✕' ||
                  label.includes('close') || label.includes('关闭');
                if (!looksClose) continue;
                candidates.push({ el, score: zIndex(el) * 10000 + rect.top + rect.left / 1000 });
              }
              candidates.sort((a, b) => b.score - a.score);
              const target = candidates[0]?.el;
              if (!target) return false;
              target.click();
              return true;
            }
            """
        )
        for scope in self._iter_scopes(page):
            try:
                if scope.evaluate(script):
                    return True
            except Exception:
                continue
        return False

    def _quick_visible_locator(self, page: Any, selector: str, timeout: int = 120) -> Any | None:
        return self._visible_locator(page, selector, selector, timeout=timeout)

    def _select_realtime_supplier(self, page: Any, supplier: SupplierRef | str) -> None:
        target = self._as_supplier_ref(supplier)
        if target.supplier_id in {"__first__", "__all__", "__auto__"} or target.supplier_name in {"__first__", "__all__", "__auto__"}:
            raise RuntimeError("任务 1 禁止使用第一项/全部供应商兜底，必须按同一供应商 ID 选择。")
        if not target.supplier_id and not target.supplier_name:
            raise RuntimeError("任务 1 缺少供应商 ID，不能用页面第一项兜底。")
        self._dismiss_notification_center(page)
        self._open_realtime_supplier_dropdown(page)
        if self._click_realtime_supplier_option(page, target, timeout=4000):
            return
        raise RuntimeError(
            f"实时库存页找不到同一供应商，已失败且未使用第一项兜底: "
            f"id={target.supplier_id or '-'} name={target.supplier_name or '-'}"
        )

    def _as_supplier_ref(self, supplier: SupplierRef | str) -> SupplierRef:
        if isinstance(supplier, SupplierRef):
            target = supplier
        else:
            text = _clean_text(supplier)
            target = SupplierRef(supplier_id=text, supplier_name=text)
        name = target.supplier_name
        supplier_id = target.supplier_id
        if supplier_id.startswith("name:"):
            name = name or supplier_id[5:]
        return SupplierRef(
            supplier_id=supplier_id,
            supplier_name=name,
            account_key=target.account_key,
        )

    def _open_realtime_supplier_dropdown(self, page: Any, optional: bool = False) -> bool:
        self._dismiss_notification_center(page)
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        self._wait_quiet(page, 1000)

        clicked = False
        fallbacks = (
            self._selector_optional("realtime.supplier_field"),
            "xpath=//*[contains(normalize-space(.), '供应商名称')]/following::*[contains(@class, 'next-select')][1]",
            "xpath=//*[contains(normalize-space(.), '供应商名称')]/following::*[contains(@class, 'next-select-selector')][1]",
            "xpath=//div[contains(@class, 'next-form-item')][.//*[contains(normalize-space(.), '供应商名称')]]//*[contains(normalize-space(.), '请选择')][1]",
            "text=请选择",
        )
        for fallback in fallbacks:
            if not fallback:
                continue
            try:
                locator = self._visible_locator(page, fallback, "供应商名称", timeout=8000)
                if locator is None:
                    continue
                if self._js_click(locator):
                    clicked = True
                    break
            except Exception:
                continue
        if not clicked:
            if optional:
                return False
            raise RuntimeError("找不到供应商名称下拉")
        self._wait_quiet(page, 1000)
        return True

    def _dismiss_notification_center(self, page: Any) -> None:
        script = """
        () => {
          const overlays = Array.from(document.querySelectorAll('.river-notification-center_notification'));
          const live = overlays.find((overlay) => {
            const rect = overlay.getBoundingClientRect();
            const style = window.getComputedStyle(overlay);
            const opened = rect.width > 80 && rect.height > 80 &&
              style.display !== 'none' && style.visibility !== 'hidden' &&
              Number(style.opacity || 1) > 0;
            return opened && overlay.querySelector('li.next-tabs-tab, li[id^="fileTask"]');
          });
          if (live) {
            const icon = document.querySelector('i.ascp-frame-icon-taskalert, i.river-origin-notification-icon');
            if (icon) icon.click();
            live.style.display = 'none';
            live.style.pointerEvents = 'none';
            live.style.visibility = 'hidden';
            return 'toggled';
          }
          for (const selector of [
            '.notification-center-mask.show',
            '.notification-center-mask',
            '.notification-center',
            '.notification-drawer-container'
          ]) {
            document.querySelectorAll(selector).forEach((el) => {
              el.classList.remove('show');
              el.style.display = 'none';
              el.style.pointerEvents = 'none';
            });
          }
          document.body.style.overflow = '';
          return live ? 'open-no-icon' : 'closed';
        }
        """
        for scope in self._iter_scopes(page):
            try:
                result = scope.evaluate(script)
            except Exception:
                continue
            if result == "toggled":
                print("[猫超] 已关闭文件中心遮罩")
                self._wait_quiet(page, 300)
                return
        self._wait_quiet(page, 200)

    def _realtime_supplier_names(self, page: Any, account: Account) -> list[str]:
        configured = [_clean_text(item) for item in account.supplier_names if _clean_text(item)]
        discovered = self._discover_realtime_supplier_names(page)
        if configured and not all(item in {"__first__", "__all__", "__auto__"} for item in configured):
            if discovered:
                matched: list[str] = []
                for item in configured:
                    hit = self._match_supplier_name(item, discovered)
                    if hit and hit not in matched:
                        matched.append(hit)
                    else:
                        matched.append(item)
                print(f"[猫超] 实时库存供应商按账号配置匹配: 配置={len(configured)} / 命中={len(matched)}")
                return matched
            return configured

        if discovered:
            print(f"[猫超] 已自动发现实时库存供应商数: {len(discovered)}")
            return discovered

        if configured:
            return configured
        return []

    def _discover_realtime_supplier_names(self, page: Any) -> list[str]:
        def collect() -> list[str]:
            candidates: list[str] = []
            seen: set[str] = set()
            for _, text, _supplier_id in self._visible_realtime_supplier_items(page):
                if not text or text in seen:
                    continue
                if text in {"请选择", "全部"}:
                    continue
                seen.add(text)
                candidates.append(text)
            return candidates

        if not self._open_realtime_supplier_dropdown(page, optional=True):
            return []

        candidates = collect()
        if not candidates:
            self._wait_quiet(page, 1500)
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            if self._open_realtime_supplier_dropdown(page, optional=True):
                candidates = collect()
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        return candidates

    def _visible_realtime_supplier_items(self, page: Any) -> list[tuple[Any, str, str]]:
        selectors = (
            ".next-overlay-wrapper:visible .next-select-menu [role='option']",
            ".next-overlay-wrapper:visible .next-select-menu .next-menu-item",
            ".next-overlay-wrapper:visible [role='option']",
            ".next-select-menu:visible [role='option']",
            ".next-select-menu:visible .next-menu-item",
        )
        items: list[tuple[Any, str, str]] = []
        seen: set[str] = set()
        for scope in self._iter_scopes(page):
            for selector in selectors:
                try:
                    locator = scope.locator(selector)
                    count = min(locator.count(), 100)
                except Exception:
                    continue
                for idx in range(count):
                    item = locator.nth(idx)
                    try:
                        if not item.is_visible(timeout=200):
                            continue
                        text = _clean_text(item.inner_text(timeout=300))
                        supplier_id = _clean_text(
                            item.get_attribute("data-id")
                            or item.get_attribute("data-value")
                            or item.get_attribute("data-key")
                            or item.get_attribute("value")
                            or ""
                        )
                    except Exception:
                        continue
                    if not text:
                        continue
                    key = f"{supplier_id}|{text}"
                    if key in seen:
                        continue
                    seen.add(key)
                    items.append((item, text, supplier_id))
        return items

    def _click_realtime_supplier_option(self, page: Any, supplier: SupplierRef | str, timeout: int = 5000) -> bool:
        expected = self._as_supplier_ref(supplier)
        deadline = time.time() + timeout / 1000
        while time.time() < deadline:
            items = self._visible_realtime_supplier_items(page)
            for item, text, supplier_id in items:
                if (
                    expected.supplier_id
                    and not expected.supplier_id.startswith("name:")
                    and supplier_id
                    and supplier_id == expected.supplier_id
                ):
                    self._click_option_locator(item)
                    return True
            best_item = None
            best_score = 0.0
            needle = expected.supplier_name or (
                expected.supplier_id[5:] if expected.supplier_id.startswith("name:") else expected.supplier_id
            )
            for item, text, _supplier_id in items:
                score = self._supplier_name_score(needle, text)
                if score > best_score:
                    best_score = score
                    best_item = item
            if best_item is not None and best_score >= 0.78:
                self._click_option_locator(best_item)
                return True
            time.sleep(0.2)
        return False

    def _click_option_locator(self, item: Any) -> None:
        try:
            item.click(timeout=1000)
        except Exception:
            try:
                item.evaluate("(el) => el.click()")
            except Exception:
                item.click(timeout=1000, force=True)
        try:
            self._wait_quiet(item.page, 800)
        except Exception:
            time.sleep(0.8)

    def _match_supplier_name(self, expected: str, candidates: list[str]) -> str:
        expected_norm = self._normalize_supplier_text(expected)
        best_candidate = ""
        best_score = 0.0
        for candidate in candidates:
            score = self._supplier_name_score(expected_norm, candidate)
            if score > best_score:
                best_score = score
                best_candidate = candidate
        return best_candidate if best_score >= 0.78 else ""

    def _normalize_supplier_text(self, value: str) -> str:
        text = _clean_text(value)
        if not text:
            return ""
        text = re.sub(r"[\(（\[].*?[\)）\]]", "", text)
        text = re.sub(r"[\s\-_/／·—–,，.。]+", "", text)
        return text

    def _supplier_name_matches(self, expected: str, actual: str) -> bool:
        return self._supplier_name_score(expected, actual) >= 0.78

    def _supplier_name_score(self, expected: str, actual: str) -> float:
        expected_norm = self._normalize_supplier_text(expected)
        actual_norm = self._normalize_supplier_text(actual)
        if not expected_norm or not actual_norm:
            return 0.0
        if expected_norm == actual_norm:
            return 1.0
        score = SequenceMatcher(None, expected_norm, actual_norm).ratio()
        if expected_norm in actual_norm or actual_norm in expected_norm:
            score = max(score, 0.9)
        return score

    def _export_realtime_supplier(
        self,
        page: Any,
        account: Account,
        supplier: str,
        result_count: int | None = None,
    ) -> RunResult:
        started = datetime.now().isoformat(timespec="seconds")
        if result_count is None:
            result_count = self._wait_realtime_inventory_result_count(page, timeout_ms=3000)
        raw_dir, cleaned_dir = self._account_data_dirs(account)
        raw_dir.mkdir(parents=True, exist_ok=True)
        cleaned_dir.mkdir(parents=True, exist_ok=True)

        if result_count == 0 or self._realtime_inventory_has_zero_items(page):
            note = f"{supplier} 查询结果 0 项，已跳过导出"
            print(f"[猫超] 实时库存: {note}")
            finished = datetime.now().isoformat(timespec="seconds")
            return self._stamp_supplier([RunResult(
                task="realtime-inventory",
                title=TASKS["realtime-inventory"]["title"],
                account=account.key,
                status="ok",
                note=note,
                started_at=started,
                finished_at=finished,
            )])[0]

        export_created = self._click_toolbar_export(
            page,
            option_texts=["导出全部", "全部"],
            allow_direct=True,
            file_task_key="realtime-inventory",
            file_task_timeout_sec=25,
        )
        existing_file_task_ids = set(getattr(self, "_pre_export_file_task_ids", set()) or set())
        self._wait_quiet(page, 1500)

        explicit_no_data = self._page_has_no_items(page) or self._page_has_text(page, "没有数据需要导出", timeout=1500)
        if result_count == 0 or explicit_no_data or (result_count is None and self._page_has_text(page, "没有数据", timeout=1000)):
            note = f"{supplier} 已尝试后台导出，平台提示无数据"
            print(f"[猫超] 实时库存: {note}")
            finished = datetime.now().isoformat(timespec="seconds")
            return self._stamp_supplier([
                RunResult(
                    task="realtime-inventory",
                    title=TASKS["realtime-inventory"]["title"],
                    account=account.key,
                    status="ok",
                    note=note,
                    started_at=started,
                    finished_at=finished,
                )
            ])[0]
        if not export_created:
            raise RuntimeError("实时库存导出未生成新文件任务")

        try:
            raw_file = self._wait_and_click_task_download(
                page,
                account,
                raw_dir,
                "realtime-inventory",
                TASKS["realtime-inventory"]["file_task_text"],
                TASKS["realtime-inventory"]["prefix"],
                prefix_extra=self._supplier_prefix() or _slug(supplier),
                task_wait_timeout_sec=self.settings.task_timeout_sec,
                exclude_file_task_ids=existing_file_task_ids,
            )
        except RuntimeError as exc:
            if self._is_null_download_error(exc) and (
                result_count == 0
                or self._realtime_inventory_has_zero_items(page)
                or self._page_has_no_items(page)
            ):
                note = f"{supplier} 实时库存文件任务返回 null/无下载文件，已跳过: {exc}"
                print(f"[猫超] 实时库存: {note}")
                finished = datetime.now().isoformat(timespec="seconds")
                return self._stamp_supplier([RunResult(
                    task="realtime-inventory",
                    title=TASKS["realtime-inventory"]["title"],
                    account=account.key,
                    status="ok",
                    note=note,
                    started_at=started,
                    finished_at=finished,
                )])[0]
            raise RuntimeError(f"实时库存供应商 {supplier} 已发起后台导出，但未下载到文件: {exc}") from exc

        cleaned_file = self._clean_file("realtime-inventory", raw_file, cleaned_dir)
        self._dismiss_notification_center(page)
        finished = datetime.now().isoformat(timespec="seconds")
        count_note = f"，查询结果 {result_count} 项" if result_count is not None else ""
        note = f"{supplier} 已下载实时库存{count_note}"
        print(f"[猫超] 实时库存: {note} -> {cleaned_file}")
        return self._stamp_supplier([
            RunResult(
                task="realtime-inventory",
                title=TASKS["realtime-inventory"]["title"],
                account=account.key,
                status="ok",
                raw_file=str(raw_file),
                cleaned_file=str(cleaned_file),
                note=note,
                started_at=started,
                finished_at=finished,
            )
        ])[0]

    def _select_purchase_statuses(
        self,
        page: Any,
        statuses: list[str],
        optional_statuses: list[str] | None = None,
        strict: bool = True,
        field_selector_key: str = "purchase.po_status_field",
        field_label: str = "采购单状态",
    ) -> None:
        optional = set(optional_statuses or [])
        for status in statuses + list(optional):
            self._click(page, field_selector_key, field_label)
            self._wait_quiet(page, 600)
            ok = self._click_text(page, status, timeout=3000, optional=True, loose=True)
            if not ok and status not in optional:
                if strict:
                    raise RuntimeError(f"{field_label}下拉中找不到: {status}")
                print(f"[猫超] {field_label}未找到，已跳过: {status}")
            if not ok and status in optional:
                print(f"[猫超] 可选状态未找到，已跳过: {status}")
            self._wait_quiet(page, 500)
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass

    def _select_first_purchase_status(
        self,
        page: Any,
        statuses: list[str],
        field_selector_key: str = "purchase.po_status_field",
        field_label: str = "采购单状态",
    ) -> str:
        tried: list[str] = []
        for status in statuses:
            tried.append(status)
            self._click(page, field_selector_key, field_label)
            self._wait_quiet(page, 600)
            if self._click_select_option_exact(page, status):
                print(f"[猫超] {field_label}已选择: {status}")
                try:
                    page.keyboard.press("Escape")
                except Exception:
                    pass
                return status
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            self._wait_quiet(page, 300)
        raise RuntimeError(f"{field_label}下拉中找不到: {', '.join(tried)}")

    def _fill_last_two_months(self, page: Any) -> None:
        today = date.today()
        start = _months_ago(today, 2).strftime("%Y-%m-%d")
        end = today.strftime("%Y-%m-%d")
        labels = self._list_filter_labels(page)
        if labels:
            print(f"[猫超] 当前筛选字段: {' / '.join(labels[:16])}")
        if self._fill_form_date_range(page, "创建时间", start, end):
            print(f"[猫超] 创建时间已设为 {start} ~ {end}")
            return
        if self._selector_visible(page, "po_list.start_date_input", timeout=800):
            self._fill(page, "po_list.start_date_input", start, "创建开始时间")
            self._fill(page, "po_list.end_date_input", end, "创建结束时间")
            self._click_optional(page, "po_list.date_confirm_button", "时间确定")
            print(f"[猫超] 创建时间已按配置XPath设为 {start} ~ {end}")
            return
        print(f"[猫超] 新工作台补货单页没有「创建时间」，按页面默认日期范围继续。当前字段: {labels[:16]}")

    def _expand_more_filters(self, page: Any) -> None:
        if self._form_item_has_label(page, "创建时间"):
            return
        if self._click_more_filters(page):
            self._wait_quiet(page, 800)
        if not self._form_item_has_label(page, "创建时间"):
            self._click_optional(page, "po_list.more_button", "更多筛选")
            self._wait_quiet(page, 800)

    def _click_more_filters(self, page: Any) -> bool:
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const inChrome = (el) => !!el.closest('header, .ascp-frame-header, .header-right, .next-table, .next-table-toolbar, .comp-toolbar');
          const query = Array.from(document.querySelectorAll('button, a, span')).find((el) => visible(el) && textOf(el) === '查询');
          const root = (query && (query.closest('form') || query.closest('.next-form, .river-page'))) || document;
          const nodes = Array.from(root.querySelectorAll('button, a, span, div[role="button"]'));
          const hit = nodes.find((el) => {
            if (!visible(el) || inChrome(el)) return false;
            const text = textOf(el);
            return text === '更多' || text === '展开' || /^更多\\s*\\(/.test(text);
          });
          if (!hit) return '';
          hit.click();
          return textOf(hit);
        }
        """
        for scope in self._iter_scopes(page):
            try:
                hit = scope.evaluate(script)
            except Exception:
                continue
            if hit:
                print(f"[猫超] 已展开更多筛选: {hit}")
                return True
        return False

    def _form_item_has_label(self, page: Any, label: str) -> bool:
        target = _clean_text(label)
        script = """
        (target) => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
          };
          return Array.from(document.querySelectorAll('label, .next-form-item-label, .next-form-item, span, div')).some((el) => {
            if (!visible(el)) return false;
            const text = textOf(el);
            return text === target || text.startsWith(target);
          });
        }
        """
        for scope in self._iter_scopes(page):
            try:
                if scope.evaluate(script, target):
                    return True
            except Exception:
                continue
        return False

    def _list_filter_labels(self, page: Any) -> list[str]:
        script = """
        () => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
          };
          const labels = [];
          const seen = new Set();
          const nodes = Array.from(document.querySelectorAll('.next-form-item-label, label, .next-form-item'));
          for (const el of nodes) {
            if (!visible(el)) continue;
            let text = textOf(el.querySelector('.next-form-item-label, label') || el);
            text = text.split('\\n')[0].replace(/[:：].*$/, '').trim();
            if (!text || text.length > 12 || seen.has(text)) continue;
            if (/查询|重置|保存|更多|展开/.test(text)) continue;
            seen.add(text);
            labels.push(text);
          }
          return labels;
        }
        """
        for scope in self._iter_scopes(page):
            try:
                values = scope.evaluate(script) or []
            except Exception:
                continue
            if values:
                return [_clean_text(item) for item in values if _clean_text(item)]
        return []

    def _fill_form_date_range(self, page: Any, label: str, start: str, end: str) -> bool:
        target = _clean_text(label)
        script = """
        ({target, start, end}) => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
          };
          const setValue = (el, value) => {
            const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
            setter.call(el, value);
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
          };
          const labels = Array.from(document.querySelectorAll('label, .next-form-item-label, span, div')).filter((el) => {
            if (!visible(el)) return false;
            const text = textOf(el);
            return text === target || text.startsWith(target + ' ') || text === target + '：' || text === target + ':';
          });
          for (const lab of labels) {
            const item = lab.closest('.next-form-item') || lab.parentElement;
            if (!item) continue;
            const inputs = Array.from(item.querySelectorAll('input')).filter((el) => visible(el));
            if (!inputs.length) continue;
            setValue(inputs[0], start);
            if (inputs[1]) setValue(inputs[1], end);
            return `ok:${inputs.length}`;
          }
          const dateInputs = Array.from(document.querySelectorAll('input')).filter((el) => {
            if (!visible(el)) return false;
            const ph = (el.getAttribute('placeholder') || '') + ' ' + (el.getAttribute('aria-label') || '');
            return /开始|结束|日期|时间/.test(ph);
          });
          if (dateInputs.length >= 2) {
            setValue(dateInputs[0], start);
            setValue(dateInputs[1], end);
            return `ok-ph:${dateInputs.length}`;
          }
          return 'no-item';
        }
        """
        for scope in self._iter_scopes(page):
            try:
                result = scope.evaluate(script, {"target": target, "start": start, "end": end})
            except Exception:
                continue
            if result and str(result).startswith("ok:"):
                return True
        return False

    def _click_select_option_exact(self, page: Any, option_text: str) -> str:
        target = _clean_text(option_text)
        if not target:
            return ""
        script = """
        (target) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const ownText = (el) => Array.from(el.childNodes)
            .filter((n) => n.nodeType === 3)
            .map((n) => (n.textContent || '').replace(/\\s+/g, ' ').trim())
            .filter(Boolean)
            .join('');
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const nodes = Array.from(document.querySelectorAll(
            '.next-overlay-wrapper.opened .next-menu-item, .next-select-menu .next-menu-item, [role="listbox"] [role="option"], .next-menu-item, [role="option"]'
          ));
          const scored = [];
          for (const el of nodes) {
            if (!visible(el)) continue;
            const label = ownText(el) || textOf(el);
            if (label !== target) continue;
            const rect = el.getBoundingClientRect();
            scored.push({el, area: Math.max(rect.width, 0) * Math.max(rect.height, 0)});
          }
          scored.sort((a, b) => b.area - a.area);
          const hit = scored.find((item) => item.area > 20);
          if (!hit) return '';
          hit.el.click();
          return target;
        }
        """
        for scope in self._iter_scopes(page):
            try:
                hit = scope.evaluate(script, target)
            except Exception:
                continue
            if hit:
                return str(hit)
        return ""

    def _reset_transfer_filters(self, page: Any) -> None:
        before = self._read_visible_date_values(page)
        if before:
            print(f"[猫超] 调拨单打开时日期: {' / '.join(before)}")
        self._click_text(page, "重置", timeout=1500, optional=True)
        self._wait_quiet(page, 600)
        after_reset = self._read_visible_date_values(page)
        if after_reset:
            print(f"[猫超] 调拨单点重置后日期: {' / '.join(after_reset)}（页面默认，不是近十天筛选）")
        self._clear_labeled_date_range(page, "创建时间")
        self._clear_visible_date_inputs(page)
        self._wait_quiet(page, 400)
        cleared = self._read_visible_date_values(page)
        print(f"[猫超] 调拨单已清空创建时间: {(' / '.join(cleared)) if cleared else '空'}")
        if self._click_text(page, "查询", timeout=1500, optional=True):
            print("[猫超] 调拨单已按无时间条件查询")
            self._wait_quiet(page, 3000)

    def _read_visible_date_values(self, page: Any) -> list[str]:
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none';
          };
          return Array.from(document.querySelectorAll('input')).filter((el) => {
            if (!visible(el)) return false;
            const ph = (el.getAttribute('placeholder') || '') + ' ' + (el.getAttribute('aria-label') || '');
            return /开始|结束|日期|时间|YYYY|yyyy/.test(ph) || /\\d{4}-\\d{2}-\\d{2}/.test(el.value || '');
          }).map((el) => (el.value || el.getAttribute('placeholder') || '').trim()).filter(Boolean);
        }
        """
        values: list[str] = []
        seen: set[str] = set()
        for scope in self._iter_scopes(page):
            try:
                found = scope.evaluate(script) or []
            except Exception:
                continue
            for item in found:
                text = _clean_text(item)
                if text and text not in seen:
                    seen.add(text)
                    values.append(text)
        return values

    def _clear_labeled_date_range(self, page: Any, label: str) -> bool:
        target = _clean_text(label)
        script = """
        (target) => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none';
          };
          const setValue = (el, value) => {
            const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
            setter.call(el, value);
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
          };
          const labels = Array.from(document.querySelectorAll('label, .next-form-item-label, span, div')).filter((el) => {
            if (!visible(el)) return false;
            const text = textOf(el);
            return text === target || text === target + '：' || text === target + ':';
          });
          let cleared = 0;
          for (const lab of labels) {
            const item = lab.closest('.next-form-item') || lab.parentElement;
            if (!item) continue;
            for (const icon of item.querySelectorAll(
              '.next-icon-delete-filling, .next-input-clear-icon, .next-icon-close, [aria-label="清除"], [aria-label="清空"]'
            )) {
              if (visible(icon)) {
                icon.click();
                cleared += 1;
              }
            }
            for (const input of item.querySelectorAll('input')) {
              if (!visible(input)) continue;
              setValue(input, '');
              cleared += 1;
            }
          }
          return cleared;
        }
        """
        total = 0
        for scope in self._iter_scopes(page):
            try:
                total += int(scope.evaluate(script, target) or 0)
            except Exception:
                continue
        if total:
            print(f"[猫超] 已清空「{target}」日期控件 {total} 处")
        return total > 0

    def _clear_visible_date_inputs(self, page: Any) -> None:
        selectors = (
            "input[placeholder*='日期']",
            "input[placeholder*='时间']",
            "input[placeholder*='开始']",
            "input[placeholder*='结束']",
            "input[placeholder*='YYYY']",
            "input[placeholder*='yyyy']",
        )
        for selector in selectors:
            for scope in self._iter_scopes(page):
                try:
                    locator = scope.locator(selector)
                    count = min(locator.count(), 50)
                except Exception:
                    continue
                for idx in range(count):
                    item = locator.nth(idx)
                    try:
                        if item.is_visible(timeout=100):
                            self._set_input_value(item, "")
                    except Exception:
                        continue

    def _visible_result_count(self, page: Any) -> int | None:
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const readCount = (el) => {
            const text = (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
            const match = text.match(/共\\s*([0-9,]+)\\s*项/);
            if (!match) return null;
            const count = Number(match[1].replace(/,/g, ''));
            return Number.isFinite(count) ? count : null;
          };
          const counts = [];
          for (const selector of ['.next-pagination-total', '.river-title-total', '.next-pagination', '.next-table-footer']) {
            for (const el of document.querySelectorAll(selector)) {
              if (!visible(el)) continue;
              const count = readCount(el);
              if (count !== null) counts.push(count);
            }
          }
          if (counts.length) return counts;
          for (const el of document.querySelectorAll('body *')) {
            if (!visible(el)) continue;
            const text = (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
            if (!text || text.length > 40) continue;
            const count = readCount(el);
            if (count !== null) counts.push(count);
          }
          return counts;
        }
        """
        frame_counts: list[int] = []
        page_counts: list[int] = []
        for scope in self._iter_scopes(page):
            try:
                values = scope.evaluate(script) or []
            except Exception:
                continue
            url = str(getattr(scope, "url", "") or "")
            bucket = frame_counts if any(
                token in url for token in ("purchase_order_list", "purchase_transfer_order_list")
            ) else page_counts
            for value in values:
                try:
                    bucket.append(int(value))
                except (TypeError, ValueError):
                    continue
        if frame_counts:
            return frame_counts[0]
        if page_counts:
            return page_counts[0]
        return None

    def _page_has_no_items(self, page: Any) -> bool:
        return any(
            self._page_has_text(page, text, timeout=500)
            for text in (
                "共 0 项",
                "共0项",
                "暂无数据",
                "无数据",
                "没有数据",
                "没有数据需要导出",
                "暂无符合条件的数据",
                "当前查询无数据",
            )
        )

    def _wait_realtime_inventory_result_count(self, page: Any, timeout_ms: int = 12000) -> int | None:
        deadline = time.time() + timeout_ms / 1000
        zero_ready_at = time.time() + min(timeout_ms / 1000, 6)
        last_count: int | None = None
        stable_hits = 0
        while time.time() < deadline:
            self._wait_quiet(page, 1000)
            count = self._realtime_inventory_result_count(page)
            if count is None:
                time.sleep(0.5)
                continue
            if count == last_count:
                stable_hits += 1
            else:
                last_count = count
                stable_hits = 1
            if stable_hits >= 2 and (count > 0 or time.time() >= zero_ready_at):
                return count
            time.sleep(0.5)
        return last_count

    def _realtime_inventory_has_zero_items(self, page: Any) -> bool:
        count = self._realtime_inventory_result_count(page)
        if count == 0:
            return True
        return count is None and self._page_has_no_items(page)

    def _realtime_inventory_result_count(self, page: Any) -> int | None:
        counts: list[int] = []
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const readCount = (el) => {
            const text = (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
            const match = text.match(/共\\s*([0-9,]+)\\s*项/);
            if (!match) return null;
            const count = Number(match[1].replace(/,/g, ''));
            return Number.isFinite(count) ? count : null;
          };
          const counts = [];
          for (const selector of [
            '.inventory_realtime_search .river-title-total',
            '.inventory_realtime_search .next-pagination-total',
            '.spa_inventory_realtime_search_1 .river-title-total',
            '.spa_inventory_realtime_search_1 .next-pagination-total',
            '.river-table .river-title-total',
            '.river-table .next-pagination-total'
          ]) {
            for (const el of document.querySelectorAll(selector)) {
              if (!visible(el)) continue;
              const count = readCount(el);
              if (count !== null) counts.push(count);
            }
          }
          if (counts.length) return counts;
          for (const el of document.querySelectorAll('body *')) {
            if (!visible(el)) continue;
            const text = (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
            if (!text || text.length > 100) continue;
            const count = readCount(el);
            if (count !== null) counts.push(count);
          }
          return counts;
        }
        """
        for scope in self._iter_scopes(page):
            try:
                values = scope.evaluate(script)
            except Exception:
                continue
            for value in values or []:
                try:
                    counts.append(int(value))
                except Exception:
                    continue
        if not counts:
            return None
        return max(counts)

    def _no_data_result(self, task_key: str, account: Account, note: str) -> RunResult:
        print(f"[猫超] {note}")
        started = datetime.now().isoformat(timespec="seconds")
        finished = datetime.now().isoformat(timespec="seconds")
        return self._stamp_supplier([
            RunResult(
                task=task_key,
                title=TASKS[task_key]["title"],
                account=account.key,
                status="ok",
                note=note,
                started_at=started,
                finished_at=finished,
            )
        ])[0]

    def _unclick_current_page_only_if_present(self, page: Any) -> None:
        selector = self._selector_optional("po_list.current_page_only_checkbox")
        if not selector:
            return
        locator = self._visible_locator(page, selector, "只下载当前页", timeout=800)
        if locator is None:
            return
        try:
            checked = bool(locator.evaluate("(el) => !!(el.checked || el.getAttribute('aria-checked') === 'true')", timeout=800))
            if checked and self._js_click(locator):
                print("[猫超] 已取消勾选“只下载当前页”。")
        except Exception:
            pass

    def _download_and_clean(
        self,
        page: Any,
        task_key: str,
        account: Account,
        file_task_id_contains: str = "",
        prefix_extra: str = "",
        note: str = "",
        task_wait_timeout_sec: int | None = None,
    ) -> RunResult:
        task = TASKS[task_key]
        started = datetime.now().isoformat(timespec="seconds")
        raw_dir, cleaned_dir = self._account_data_dirs(account)
        raw_dir.mkdir(parents=True, exist_ok=True)
        cleaned_dir.mkdir(parents=True, exist_ok=True)
        extra = prefix_extra or self._supplier_prefix()
        existing_file_task_ids = set(getattr(self, "_pre_export_file_task_ids", set()) or set())
        print(f"[猫超] 下载时排除导出前文件任务 {len(existing_file_task_ids)} 条")

        raw_file = self._wait_and_click_task_download(
            page,
            account,
            raw_dir,
            task_key,
            task["file_task_text"],
            task["prefix"],
            file_task_id_contains=file_task_id_contains,
            prefix_extra=extra,
            task_wait_timeout_sec=task_wait_timeout_sec,
            exclude_file_task_ids=existing_file_task_ids,
        )
        cleaned_file = self._clean_file(task_key, raw_file, cleaned_dir)
        self._dismiss_notification_center(page)
        finished = datetime.now().isoformat(timespec="seconds")
        print(f"[猫超] 完成: {task['title']} -> {cleaned_file}")
        return self._stamp_supplier([
            RunResult(
                task=task_key,
                title=task["title"],
                account=account.key,
                status="ok",
                raw_file=str(raw_file),
                cleaned_file=str(cleaned_file),
                started_at=started,
                finished_at=finished,
                note=note,
            )
        ])[0]

    def _wait_and_click_task_download(
        self,
        page: Any,
        account: Account,
        raw_dir: Path,
        task_key: str,
        file_task_text: str,
        prefix: str,
        file_task_id_contains: str = "",
        prefix_extra: str = "",
        task_wait_timeout_sec: int | None = None,
        exclude_file_task_ids: set[str] | None = None,
    ) -> Path:
        wait_started = time.time()
        file_task_texts = self._file_task_text_candidates(task_key, file_task_text)
        print(f"[猫超] 等待文件任务完成: {' / '.join(file_task_texts)}")
        self._file_center_probed = False
        js_clicked = False
        wait_timeout = task_wait_timeout_sec if task_wait_timeout_sec is not None else 90
        deadline = time.time() + wait_timeout
        polls = 0
        before = self._download_snapshot(account.download_dir)
        while time.time() < deadline and not js_clicked:
            polls += 1
            if polls == 1:
                self._dismiss_notification_center(page)
                self._wait_quiet(page, 250)
            self._open_file_notification_center(page)
            if self._click_file_task_download_js(page, file_task_texts, exclude_file_task_ids):
                js_clicked = True
                break
            failure = self._file_task_failure_detail(page, file_task_texts, exclude_file_task_ids)
            if failure:
                raise RuntimeError(f"文件任务后台失败: {failure}")
            titles = self._list_file_task_titles(page)
            if titles:
                print(f"[猫超] 文件中心 {len(titles)} 条，最新: {titles[0][:80]}")
            time.sleep(self.settings.poll_interval_sec)
        click_target = None
        if click_target is None and not js_clicked:
            existing = self._latest_matching_download(account.download_dir, task_key, wait_started)
            if existing is not None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                name_parts = [prefix, timestamp]
                if prefix_extra:
                    name_parts.append(prefix_extra)
                local_prefix = "_".join(name_parts)
                target = self._unique_path(raw_dir / f"{local_prefix}_{existing.name}")
                if existing.resolve() != target.resolve():
                    shutil.copy2(existing, target)
                return target
            titles = self._list_file_task_titles(page)
            if titles:
                print(f"[猫超] 当前文件任务: {' || '.join(titles[:8])}")
            raise RuntimeError(f"等待文件任务下载按钮超时: {' / '.join(file_task_texts)}")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name_parts = [prefix, timestamp]
        if prefix_extra:
            name_parts.append(prefix_extra)
        local_prefix = "_".join(name_parts)

        if js_clicked or click_target is None:
            downloaded = self._latest_matching_download(account.download_dir, task_key, wait_started)
            if downloaded is None:
                downloaded = self._wait_new_download(account.download_dir, before, timeout_sec=60)
            target = self._unique_path(raw_dir / f"{local_prefix}_{downloaded.name}")
            if downloaded.resolve() != target.resolve():
                shutil.copy2(downloaded, target)
            return target

        try:
            download_event_timeout = min(self.settings.download_timeout_sec, 10) * 1000
            with page.expect_download(timeout=download_event_timeout) as download_info:
                try:
                    click_target.evaluate("(el) => el.click()")
                except Exception:
                    click_target.click(timeout=5000)
            download = download_info.value
            suggested = _slug(download.suggested_filename or f"{task_key}.download")
            target = self._unique_path(raw_dir / f"{local_prefix}_{suggested}")
            download.save_as(str(target))
            if target.stat().st_size > 0:
                return target
            print("[猫超] Playwright 保存的下载文件为空，改用浏览器下载目录文件。")
            try:
                target.unlink()
            except Exception:
                pass
            downloaded = self._wait_new_download(account.download_dir, before)
            target = self._unique_path(raw_dir / f"{local_prefix}_{downloaded.name}")
            if downloaded.resolve() != target.resolve():
                shutil.copy2(downloaded, target)
            return target
        except Exception as exc:
            print(f"[猫超] Playwright download 事件未捕获，改用目录轮询: {exc}")
            try:
                try:
                    click_target.evaluate("(el) => el.click()")
                except Exception:
                    click_target.click(timeout=5000)
            except Exception:
                pass
            downloaded = self._wait_new_download(account.download_dir, before)
            target = self._unique_path(raw_dir / f"{local_prefix}_{downloaded.name}")
            if downloaded.resolve() != target.resolve():
                shutil.copy2(downloaded, target)
            return target

    def _file_download_locator(self, page: Any, file_task_texts: Iterable[str], file_task_id_contains: str = "") -> Any:
        tasks = page.locator("li[id^='fileTask']")
        if file_task_id_contains:
            tasks = page.locator(f'li[id*="{file_task_id_contains}"]')
        needles = [_clean_text(text) for text in file_task_texts if _clean_text(text)]
        if needles:
            pattern = "|".join(re.escape(text) for text in needles)
            tasks = tasks.filter(has_text=re.compile(pattern))
        return tasks.get_by_text("下载", exact=True)

    def _click_file_task_download_js(self, page: Any, file_task_texts: Iterable[str], exclude_file_task_ids: set[str] | None = None) -> bool:
        needles = [_clean_text(text) for text in file_task_texts if _clean_text(text)]
        exclude = list(exclude_file_task_ids or [])
        script = """
        ({needles, exclude}) => {
          const visible = (el) => {
            if (!el) return false;
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 8 && rect.height > 8 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el?.innerText || el?.textContent || '').replace(/\\s+/g, ' ').trim();
          const labelOf = (el) => (
            textOf(el) ||
            el?.getAttribute?.('title') ||
            el?.getAttribute?.('aria-label') ||
            el?.getAttribute?.('data-tip') ||
            ''
          ).replace(/\\s+/g, ' ').trim();
          const rowKey = (row) => {
            const id = row.id || row.getAttribute('data-id') || row.getAttribute('data-row-key') || '';
            if (id) return id;
            return `text:${textOf(row).slice(0, 220)}`;
          };
          const addRows = (rows, seen, root, selector) => {
            for (const row of root.querySelectorAll(selector)) {
              if (seen.has(row)) continue;
              seen.add(row);
              rows.push(row);
            }
          };
          const roots = Array.from(document.querySelectorAll(
            '.river-notification-center_notification, .notification-center, .notification-drawer-container, #notification-center'
          ));
          if (!roots.length) roots.push(document);
          const rowSelectors = [
            'li[id^="fileTask"]',
            '[id^="fileTask"]',
            '[id*="fileTask"]',
            '[role="row"]',
            '.next-table-row',
            '.next-list-item',
            '.file-task-item',
            '[class*="fileTask"]',
            '[class*="file-task"]',
            '[class*="task-item"]'
          ];
          const rows = [];
          const seen = new Set();
          for (const root of roots) {
            for (const selector of rowSelectors) addRows(rows, seen, root, selector);
          }
          const ranked = rows
            .map((row, index) => ({row, index, vis: visible(row) || !!row.querySelector('a,button,[role="button"]')}))
            .filter(({row}) => {
              const text = textOf(row);
              return text && text.length < 1200 && needles.some((needle) => needle && text.includes(needle));
            })
            .filter(({row}) => !exclude.includes(rowKey(row)))
            .sort((a, b) => Number(b.vis) - Number(a.vis) || a.index - b.index);
          for (const {row} of ranked) {
            const text = textOf(row);
            if (!needles.some((needle) => needle && text.includes(needle))) continue;
            const nodes = Array.from(row.querySelectorAll(
              'a.next-btn, a.next-btn-text, .file-item-operation a, a, button, [role="button"], [title], [aria-label], span, div'
            ));
            const rowVisible = visible(row);
            const matches = nodes.filter((el) => {
              const label = labelOf(el);
              const downloadLike = /^(下载|立即下载|下载文件|重新下载)$/.test(label) || /下载$/.test(label) || /下载/.test(textOf(el));
              return downloadLike;
            });
            matches.sort((a, b) => a.querySelectorAll('*').length - b.querySelectorAll('*').length);
            let btn = matches[0];
            let clickRightSide = false;
            if (!btn && /下载\\s*$/.test(text)) {
              btn = row;
              clickRightSide = true;
            }
            if (!btn) continue;
            if (visible(btn)) {
              btn.scrollIntoView({block: 'nearest', inline: 'nearest'});
              let rect = btn.getBoundingClientRect();
              const clientX = clickRightSide ? Math.max(rect.left + 12, rect.right - 28) : rect.left + rect.width / 2;
              const clientY = rect.top + rect.height / 2;
              if (clickRightSide) {
                const hit = document.elementFromPoint(clientX, clientY);
                if (hit && row.contains(hit)) {
                  btn = hit.closest('a, button, [role="button"]') || hit;
                  rect = btn.getBoundingClientRect();
                }
              }
              for (const type of ['mouseenter', 'mousedown', 'mouseup', 'click']) {
                btn.dispatchEvent(new MouseEvent(type, {bubbles: true, cancelable: true, view: window, clientX, clientY}));
              }
            }
            if (typeof btn.click === 'function') btn.click();
            return {key: rowKey(row), label: labelOf(btn), row: text.slice(0, 120)};
          }
          return null;
        }
        """
        for scope in self._iter_scopes(page):
            try:
                clicked = scope.evaluate(script, {"needles": needles, "exclude": exclude})
            except Exception:
                continue
            if clicked:
                if isinstance(clicked, dict):
                    print(
                        f"[猫超] 已用脚本点击文件任务下载: "
                        f"{clicked.get('key') or 'ok'} {clicked.get('label') or ''} {clicked.get('row') or ''}"
                    )
                else:
                    print(f"[猫超] 已用脚本点击文件任务下载: {clicked}")
                return True
        return False

    def _file_task_text_candidates(self, task_key: str, file_task_text: str) -> list[str]:
        candidates: list[str] = []

        def add(value: str) -> None:
            text = _clean_text(value)
            if text and text not in candidates:
                candidates.append(text)

        add(file_task_text)
        if file_task_text.startswith("导出 "):
            add(file_task_text.removeprefix("导出 "))
        else:
            add(f"导出 {file_task_text}")

        aliases = {
            "realtime-inventory": ["实时库存", "导出 实时库存"],
            "pincang-detail": ["品仓明细表", "品仓明细", "导出 品仓明细", "导出货品明细", "导出列表"],
            "channel-goods": ["货品生命周期导出结果", "货品生命周期导出", "导出 库位明细", "渠道货品"],
            "po-list": ["PO明细分页导出", "导出 PO明细分页导出"],
            "transfer-order": ["调拨明细数据导出", "导出 调拨单货品明细", "调拨单货品明细", "调拨单明细导出", "导出 调拨单明细导出"],
        }
        for alias in aliases.get(task_key, []):
            add(alias)
        return candidates

    def _latest_matching_download(self, directory: Path, task_key: str, modified_after: float) -> Path | None:
        keywords = self._expected_download_keywords(task_key)
        if not keywords:
            return None
        try:
            files = [
                path
                for path in directory.iterdir()
                if path.is_file()
                and path.suffix not in {".crdownload", ".tmp"}
                and path.stat().st_mtime >= modified_after
                and any(keyword in path.name for keyword in keywords)
            ]
        except Exception:
            return None
        if not files:
            return None
        latest = max(files, key=lambda path: path.stat().st_mtime)
        print(f"[猫超] 文件任务按钮未出现，改用下载目录已生成文件: {latest.name}")
        return latest

    def _expected_download_keywords(self, task_key: str) -> list[str]:
        return {
            "realtime-inventory": ["实时库存"],
            "pincang-detail": ["品仓明细表", "品仓明细", "货品明细"],
            "system-order": ["PO明细确认分页导出"],
            "po-list": ["PO明细分页导出"],
            "channel-goods": ["货品生命周期导出结果", "货品生命周期导出", "渠道货品", "库位明细"],
            "transfer-order": ["调拨明细数据导出", "调拨单货品明细", "调拨单明细导出"],
        }.get(task_key, [])

    def _is_null_download_error(self, exc: RuntimeError) -> bool:
        message = str(exc)
        return (
            "文件任务后台失败" in message
            or "等待文件任务下载按钮超时" in message
            or "下载目录未出现新文件" in message
        )

    def _file_task_failure_detail(
        self,
        page: Any,
        file_task_texts: Iterable[str],
        exclude_file_task_ids: set[str] | None = None,
    ) -> str:
        needles = [_clean_text(text) for text in file_task_texts if _clean_text(text)]
        exclude = list(exclude_file_task_ids or [])
        script = """
        ({needles, exclude}) => {
          const textOf = (el) => (el?.innerText || el?.textContent || '').replace(/\\s+/g, ' ').trim();
          const rowKey = (row) => {
            const id = row.id || row.getAttribute('data-id') || row.getAttribute('data-row-key') || '';
            return id || `text:${textOf(row).slice(0, 220)}`;
          };
          const roots = Array.from(document.querySelectorAll(
            '.river-notification-center_notification, .notification-center, .notification-drawer-container, #notification-center'
          ));
          if (!roots.length) roots.push(document);
          const selectors = [
            'li[id^="fileTask"]',
            '[id^="fileTask"]',
            '[id*="fileTask"]',
            '[role="row"]',
            '.next-table-row',
            '.next-list-item',
            '.file-task-item',
            '[class*="fileTask"]',
            '[class*="file-task"]',
            '[class*="task-item"]'
          ];
          const seen = new Set();
          for (const root of roots) {
            for (const selector of selectors) {
              for (const row of root.querySelectorAll(selector)) {
                if (seen.has(row) || exclude.includes(rowKey(row))) continue;
                seen.add(row);
                const text = textOf(row);
                if (!text || text.length >= 1200) continue;
                if (!needles.some((needle) => needle && text.includes(needle))) continue;
                if (/返回\\s*NULL|返回\\s*null|失败|异常|请联系业务|error/i.test(text)) return text;
              }
            }
          }
          return '';
        }
        """
        for scope in self._iter_scopes(page):
            try:
                detail = _clean_text(scope.evaluate(script, {"needles": needles, "exclude": exclude}))
            except Exception:
                continue
            if detail:
                return detail
        return ""

    def _first_visible_in_locator(
        self,
        locator: Any,
        timeout: int = 1000,
        exclude_file_task_ids: set[str] | None = None,
    ) -> Any | None:
        deadline = time.time() + timeout / 1000
        while time.time() < deadline:
            try:
                count = min(locator.count(), 100)
            except Exception:
                time.sleep(0.2)
                continue
            for idx in range(count):
                item = locator.nth(idx)
                try:
                    if self._element_is_displayed(item):
                        if self._file_task_link_excluded(item, exclude_file_task_ids):
                            continue
                        return item
                except Exception:
                    continue
            time.sleep(0.2)
        return None

    def _file_task_link_excluded(self, locator: Any, exclude_file_task_ids: set[str] | None) -> bool:
        if not exclude_file_task_ids:
            return False
        try:
            file_task_id = locator.evaluate(
                """
                (el) => {
                  const root = el.closest('li[id^="fileTask"]') || el.closest('[id*="fileTask"]') || el;
                  return root && root.id ? root.id : '';
                }
                """
            )
        except Exception:
            return False
        return bool(file_task_id and file_task_id in exclude_file_task_ids)

    def _list_file_task_titles(self, page: Any) -> list[str]:
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 8 && rect.height > 8 &&
              style.display !== 'none' && style.visibility !== 'hidden' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const roots = Array.from(document.querySelectorAll(
            '.river-notification-center_notification, .notification-center, .notification-drawer-container, #notification-center'
          ));
          if (!roots.length) roots.push(document);
          const selectors = [
            'li[id^="fileTask"]',
            '[id^="fileTask"]',
            '[id*="fileTask"]',
            '[role="row"]',
            '.next-table-row',
            '.next-list-item',
            '.file-task-item',
            '[class*="fileTask"]',
            '[class*="file-task"]',
            '[class*="task-item"]'
          ];
          const rows = [];
          const seen = new Set();
          for (const root of roots) {
            for (const selector of selectors) {
              for (const row of root.querySelectorAll(selector)) {
                if (seen.has(row)) continue;
                seen.add(row);
                rows.push(row);
              }
            }
          }
          return rows
            .map((el, index) => ({text: textOf(el), visible: visible(el), index}))
            .filter((item) => item.text && item.text.length < 1200 && /导出|下载|文件|实时库存|PO明细|品仓|调拨|货品/.test(item.text))
            .sort((a, b) => Number(b.visible) - Number(a.visible) || a.index - b.index)
            .map((item) => item.text);
        }
        """
        titles: list[str] = []
        for scope in self._iter_scopes(page):
            try:
                titles.extend(scope.evaluate(script) or [])
            except Exception:
                continue
        seen: set[str] = set()
        unique: list[str] = []
        for title in titles:
            if title in seen:
                continue
            seen.add(title)
            unique.append(title)
        return unique

    def _file_task_ids(self, page: Any, file_task_text: str, task_key: str = "realtime-inventory") -> set[str]:
        file_task_texts = self._file_task_text_candidates(task_key, file_task_text)
        script = """
        (texts) => {
          const values = new Set();
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const rowKey = (row) => {
            const id = row.id || row.getAttribute('data-id') || row.getAttribute('data-row-key') || '';
            if (id) return id;
            return `text:${textOf(row).slice(0, 220)}`;
          };
          const roots = Array.from(document.querySelectorAll(
            '.river-notification-center_notification, .notification-center, .notification-drawer-container, #notification-center'
          ));
          if (!roots.length) roots.push(document);
          const selectors = [
            'li[id^="fileTask"]',
            '[id^="fileTask"]',
            '[id*="fileTask"]',
            '[role="row"]',
            '.next-table-row',
            '.next-list-item',
            '.file-task-item',
            '[class*="fileTask"]',
            '[class*="file-task"]',
            '[class*="task-item"]'
          ];
          const rows = [];
          const seen = new Set();
          for (const root of roots) {
            for (const selector of selectors) {
              for (const row of root.querySelectorAll(selector)) {
                if (seen.has(row)) continue;
                seen.add(row);
                rows.push(row);
              }
            }
          }
          for (const item of rows) {
            const text = textOf(item);
            const title = Array.from(item.querySelectorAll('[title], [aria-label]'))
              .map((el) => el.getAttribute('title') || el.getAttribute('aria-label') || '')
              .join(' ');
            if (texts.some((needle) => needle && (text.includes(needle) || title.includes(needle)))) {
              values.add(rowKey(item));
            }
          }
          return Array.from(values);
        }
        """
        ids: set[str] = set()
        for scope in self._iter_scopes(page):
            try:
                ids.update(scope.evaluate(script, file_task_texts) or [])
            except Exception:
                continue
        return ids

    def _snapshot_file_task_ids(self, page: Any) -> set[str]:
        script = """
        () => {
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const rowKey = (row) => {
            const id = row.id || row.getAttribute('data-id') || row.getAttribute('data-row-key') || '';
            if (id) return id;
            return `text:${textOf(row).slice(0, 220)}`;
          };
          const roots = Array.from(document.querySelectorAll(
            '.river-notification-center_notification, .notification-center, .notification-drawer-container, #notification-center'
          ));
          if (!roots.length) roots.push(document);
          const selectors = [
            'li[id^="fileTask"]',
            '[id^="fileTask"]',
            '[id*="fileTask"]',
            '[role="row"]',
            '.next-table-row',
            '.next-list-item',
            '.file-task-item',
            '[class*="fileTask"]',
            '[class*="file-task"]',
            '[class*="task-item"]'
          ];
          const values = new Set();
          for (const root of roots) {
            for (const selector of selectors) {
              for (const row of root.querySelectorAll(selector)) {
                const text = textOf(row);
                if (!text || text.length >= 1200 || !/导出|下载|文件|实时库存|PO明细|品仓|调拨|货品/.test(text)) continue;
                values.add(rowKey(row));
              }
            }
          }
          return Array.from(values);
        }
        """
        ids: set[str] = set()
        for scope in self._iter_scopes(page):
            try:
                ids.update(scope.evaluate(script) or [])
            except Exception:
                continue
        self._pre_export_file_task_ids = ids
        print(f"[猫超] 导出前已记录文件任务 {len(ids)} 条")
        return ids

    def _restore_notification_center_styles(self, page: Any) -> None:
        script = """
        () => {
          for (const selector of [
            '#notification-center',
            '.notification-center',
            '.notification-drawer-container',
            '#notification-center-mask',
            '.notification-center-mask',
            '.river-notification-center_notification'
          ]) {
            document.querySelectorAll(selector).forEach((el) => {
              if (el.style.display === 'none' || el.style.pointerEvents === 'none') {
                el.style.display = '';
                el.style.pointerEvents = '';
                el.style.visibility = '';
                el.style.opacity = '';
              }
            });
          }
        }
        """
        for scope in self._iter_scopes(page):
            try:
                scope.evaluate(script)
            except Exception:
                continue

    def _file_center_is_open(self, page: Any) -> bool:
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 8 && rect.height > 8 &&
              style.display !== 'none' && style.visibility !== 'hidden' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const roots = Array.from(document.querySelectorAll(
            '.river-notification-center_notification, .notification-center, .notification-drawer-container, #notification-center'
          ));
          const selectors = [
            'li[id^="fileTask"]',
            '[id^="fileTask"]',
            '[id*="fileTask"]',
            '[role="row"]',
            '.next-table-row',
            '.next-list-item',
            '.file-task-item',
            '[class*="fileTask"]',
            '[class*="file-task"]',
            '[class*="task-item"]'
          ];
          for (const root of roots) {
            if (!visible(root)) continue;
            for (const selector of selectors) {
              for (const row of root.querySelectorAll(selector)) {
                const text = textOf(row);
                if (visible(row) && text && /导出|下载|文件|实时库存|PO明细|品仓|调拨|货品/.test(text)) return true;
              }
            }
          }
          return false;
        }
        """
        for scope in self._iter_scopes(page):
            try:
                if scope.evaluate(script):
                    return True
            except Exception:
                continue
        return False

    def _file_overlay_is_open(self, page: Any) -> bool:
        script = """
        () => {
          const roots = Array.from(document.querySelectorAll('.river-notification-center_notification'));
          return roots.some((root) => {
            const rect = root.getBoundingClientRect();
            const style = window.getComputedStyle(root);
            const opened = rect.width > 80 && rect.height > 80 &&
              style.display !== 'none' && style.visibility !== 'hidden' &&
              Number(style.opacity || 1) > 0;
            if (!opened) return false;
            const tabs = Array.from(root.querySelectorAll('li.next-tabs-tab, [role="tab"]'));
            return tabs.some((el) => /文件/.test((el.innerText || '')));
          });
        }
        """
        for scope in self._iter_scopes(page):
            try:
                if scope.evaluate(script):
                    return True
            except Exception:
                continue
        return False

    def _probe_file_center(self, page: Any) -> None:
        script = """
        () => {
          const roots = Array.from(document.querySelectorAll(
            '.river-notification-center_notification, .notification-center, .notification-drawer-container, #notification-center'
          ));
          const selectors = [
            'li[id^="fileTask"]',
            '[id^="fileTask"]',
            '[id*="fileTask"]',
            '[role="row"]',
            '.next-table-row',
            '.next-list-item',
            '.file-task-item',
            '[class*="fileTask"]',
            '[class*="file-task"]',
            '[class*="task-item"]'
          ];
          const rows = new Set();
          for (const root of roots) {
            for (const selector of selectors) {
              for (const row of root.querySelectorAll(selector)) rows.add(row);
            }
          }
          return {
            icon: !!document.querySelector('i.ascp-frame-icon-taskalert, i.river-origin-notification-icon'),
            badge: !!document.querySelector('.badge.rex-count.notification-count'),
            tabs: Array.from(document.querySelectorAll('li.next-tabs-tab'))
              .map((el) => (el.innerText || '').replace(/\\s+/g, ' ').trim())
              .filter((text) => /文件|消息/.test(text)),
            tasks: rows.size
          };
        }
        """
        for scope in self._iter_scopes(page):
            url = str(getattr(scope, "url", "") or "")[:90]
            if url == "about:blank":
                continue
            try:
                info = scope.evaluate(script)
            except Exception as exc:
                print(f"[猫超] 文件中心探测失败: {url or '(no-url)'} {exc}")
                continue
            if not info:
                continue
            if info.get("icon") or info.get("badge") or info.get("tabs") or info.get("tasks"):
                print(
                    f"[猫超] 文件中心探测 {url or '(page)'} "
                    f"icon={info.get('icon')} badge={info.get('badge')} "
                    f"tabs={info.get('tabs')} tasks={info.get('tasks')}"
                )

    def _click_notification_opener_js(self, page: Any) -> str:
        script = """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 80 && rect.height > 80 &&
              style.display !== 'none' && style.visibility !== 'hidden' &&
              Number(style.opacity || 1) > 0;
          };
          const overlays = Array.from(document.querySelectorAll('.river-notification-center_notification'));
          const live = overlays.find((overlay) => visible(overlay) && overlay.querySelector('li.next-tabs-tab'));
          if (live) return 'already-open';
          const icon = document.querySelector('i.ascp-frame-icon-taskalert, i.river-origin-notification-icon');
          if (!icon) return 'missing';
          icon.click();
          return 'clicked';
        }
        """
        for scope in self._iter_scopes(page):
            try:
                result = scope.evaluate(script)
            except Exception as exc:
                print(f"[猫超] 通知入口脚本失败: {exc}")
                continue
            if result and result != "missing":
                return str(result)
        return "missing"

    def _click_file_center_tab(self, page: Any) -> bool:
        script = """
        () => {
          const roots = Array.from(document.querySelectorAll('.river-notification-center_notification'));
          for (const root of roots) {
            const tabs = Array.from(root.querySelectorAll('li.next-tabs-tab, [role="tab"]'));
            const tab = tabs.find((el) => /文件/.test((el.innerText || '').replace(/\\s+/g, ' ')));
            if (!tab) continue;
            const inner = tab.querySelector('.next-tabs-tab-inner') || tab;
            inner.click();
            tab.click();
            return (tab.innerText || '').replace(/\\s+/g, ' ').trim();
          }
          return '';
        }
        """
        for scope in self._iter_scopes(page):
            try:
                clicked = scope.evaluate(script)
            except Exception:
                continue
            if clicked:
                print(f"[猫超] 已切换文件中心页签: {clicked}")
                return True
        print("[猫超] 未点到文件中心页签")
        return False

    def _open_file_notification_center(self, page: Any) -> bool:
        self._restore_notification_center_styles(page)
        if not getattr(self, "_file_center_probed", False):
            self._probe_file_center(page)
            self._file_center_probed = True
        overlay_open = self._file_overlay_is_open(page)
        if overlay_open:
            self._click_file_center_tab(page)
            self._wait_quiet(page, 400)
            opened = self._file_center_is_open(page)
            print(f"[猫超] 文件中心已在页面上，任务列表={'有' if opened else '无'}")
            return opened

        opener = self._click_notification_opener_js(page)
        print(f"[猫超] 通知入口: {opener}")
        if opener == "missing":
            print("[猫超] 未找到通知入口，无法打开文件中心")
            return False
        self._wait_quiet(page, 800)
        self._click_file_center_tab(page)
        self._wait_quiet(page, 500)
        opened = self._file_center_is_open(page)
        print(f"[猫超] 文件任务列表={'已出现' if opened else '仍未出现'}")
        return opened

    def _download_snapshot(self, directory: Path) -> dict[Path, float]:
        directory.mkdir(parents=True, exist_ok=True)
        return {path: path.stat().st_mtime for path in directory.iterdir() if path.is_file()}

    def _wait_new_download(self, directory: Path, before: dict[Path, float], timeout_sec: int | None = None) -> Path:
        deadline = time.time() + (timeout_sec if timeout_sec is not None else self.settings.download_timeout_sec)
        latest: Path | None = None
        while time.time() < deadline:
            candidates = []
            for path in directory.iterdir():
                if not path.is_file():
                    continue
                if path.suffix in {".crdownload", ".tmp"}:
                    continue
                if path not in before or path.stat().st_mtime > before.get(path, 0):
                    candidates.append(path)
            if candidates:
                latest = max(candidates, key=lambda p: p.stat().st_mtime)
                size1 = latest.stat().st_size
                time.sleep(0.8)
                if latest.exists() and latest.stat().st_size == size1:
                    return latest
            time.sleep(self.settings.poll_interval_sec)
        raise RuntimeError(f"下载目录未出现新文件: {directory}")

    def _clean_file(self, task_key: str, raw_file: Path, cleaned_dir: Path) -> Path:
        suffix = raw_file.suffix.lower()
        target = self._unique_path(cleaned_dir / raw_file.name)
        if suffix == ".csv":
            return self._clean_csv(task_key, raw_file, target)
        if suffix == ".xlsx":
            return self._clean_xlsx(task_key, raw_file, target)
        shutil.copy2(raw_file, target)
        print(f"[猫超] 暂不识别该格式，仅复制原文件: {raw_file.name}")
        return target

    def _clean_csv(self, task_key: str, source: Path, target: Path) -> Path:
        text = self._read_text_auto(source)
        rows = list(csv.reader(text.splitlines()))
        rows = self._clean_rows(task_key, rows)
        with target.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        return target

    def _clean_xlsx(self, task_key: str, source: Path, target: Path) -> Path:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("清洗 xlsx 需要 openpyxl，请安装 requirements.txt") from exc

        workbook = load_workbook(source)
        preserve_columns = set(self.settings.cleanup.get("preserve_columns", []))
        for sheet in workbook.worksheets:
            headers = {
                idx: _clean_text(cell.value)
                for idx, cell in enumerate(sheet[1], start=1)
            } if sheet.max_row >= 1 else {}

            for row in sheet.iter_rows():
                for cell in row:
                    header = headers.get(cell.column, "")
                    if isinstance(cell.value, str):
                        cell.value = self._normalize_cell_value(
                            cell.value,
                            preserve=header in preserve_columns,
                        )
        workbook.save(target)
        return target

    def _clean_rows(self, task_key: str, rows: list[list[Any]]) -> list[list[Any]]:
        if not rows:
            return rows
        headers = [_clean_text(v) for v in rows[0]]
        preserve_columns = set(self.settings.cleanup.get("preserve_columns", []))
        keep_rows = [rows[0]]
        for row in rows[1:]:
            cleaned = []
            for idx, value in enumerate(row):
                cleaned.append(
                    self._normalize_cell_value(
                        value,
                        preserve=idx < len(headers) and headers[idx] in preserve_columns,
                    )
                )
            keep_rows.append(cleaned)
        return keep_rows

    def _delete_transfer_rows(self, sheet: Any, headers: dict[int, str]) -> None:
        status_col = None
        for col, header in headers.items():
            if header == self.settings.cleanup.get("transfer_status_column", "调拨单状态"):
                status_col = col
                break
        if status_col is None:
            print("[猫超] 调拨单未找到“调拨单状态”列，跳过状态过滤。")
            return
        drop_statuses = set(self.settings.cleanup.get("transfer_drop_statuses", ["全部出库全部入库"]))
        delete_rows = []
        for row_idx in range(2, sheet.max_row + 1):
            if _clean_text(sheet.cell(row_idx, status_col).value) in drop_statuses:
                delete_rows.append(row_idx)
        for row_idx in reversed(delete_rows):
            sheet.delete_rows(row_idx)
        if delete_rows:
            print(f"[猫超] 调拨单已删除状态过滤行数: {len(delete_rows)}")

    def _transfer_status_index(self, headers: list[str]) -> int | None:
        wanted = self.settings.cleanup.get("transfer_status_column", "调拨单状态")
        for idx, header in enumerate(headers):
            if header == wanted:
                return idx
        print("[猫超] 调拨单 CSV 未找到“调拨单状态”列，跳过状态过滤。")
        return None

    def _normalize_cell_value(self, value: Any, preserve: bool = False) -> Any:
        text = _clean_text(value)
        if preserve or text == "":
            return text
        text = text.replace("\u3000", " ").strip()
        text_no_comma = text.replace(",", "")
        if re.fullmatch(r"-?\d+(\.0+)?", text_no_comma):
            return int(float(text_no_comma))
        if re.fullmatch(r"-?\d+\.\d+", text_no_comma):
            return float(text_no_comma)
        return text

    def _read_text_auto(self, path: Path) -> str:
        raw = path.read_bytes()
        for encoding in ("utf-8-sig", "utf-8", "gb18030", "utf-16"):
            try:
                return raw.decode(encoding)
            except UnicodeDecodeError:
                continue
        return raw.decode("utf-8", errors="replace")

    def _unique_path(self, path: Path) -> Path:
        if not path.exists():
            return path
        stem = path.stem
        suffix = path.suffix
        for idx in range(1, 1000):
            candidate = path.with_name(f"{stem}_{idx}{suffix}")
            if not candidate.exists():
                return candidate
        raise RuntimeError(f"无法生成唯一文件名: {path}")

    def _selector(self, dotted_key: str) -> str:
        value = self._selector_optional(dotted_key)
        if not value:
            raise RuntimeError(f"config 缺少 XPath/selectors.{dotted_key}")
        return value

    def _selector_optional(self, dotted_key: str) -> str:
        node: Any = self._active_selectors
        for part in dotted_key.split("."):
            if not isinstance(node, dict) or part not in node:
                return ""
            node = node[part]
        return _clean_text(node)

    def _list_config(self, dotted_key: str, default: list[str]) -> list[str]:
        node: Any = self._active_selectors
        for part in dotted_key.split("."):
            if not isinstance(node, dict) or part not in node:
                return default
            node = node[part]
        if isinstance(node, list):
            return [_clean_text(v) for v in node if _clean_text(v)]
        return default

    def _click(self, page: Any, selector_key: str, label: str, timeout: int = 10000) -> None:
        selector = self._selector(selector_key)
        locator = self._visible_locator(page, selector, label, timeout=timeout)
        if locator is None:
            raise RuntimeError(f"找不到{label}: selectors.{selector_key}")
        if not self._js_click(locator):
            raise RuntimeError(f"点击{label}失败: selectors.{selector_key}")

    def _click_optional(self, page: Any, selector_key: str, label: str, timeout: int = 2000) -> bool:
        selector = self._selector_optional(selector_key)
        if not selector:
            return False
        try:
            locator = self._visible_locator(page, selector, label, timeout=timeout)
            if locator is None:
                return False
            return self._js_click(locator)
        except Exception:
            return False

    def _click_exact_control(self, page: Any, text: str, timeout: int = 8000, overlay_only: bool = False) -> bool:
        deadline = time.time() + timeout / 1000
        while time.time() < deadline:
            hit = self._js_click_matching_text(page, [text], overlay_only=overlay_only, exact=True)
            if hit:
                self._wait_quiet(page, 400)
                print(f"[猫超] 已精确点击: {text}")
                return True
            time.sleep(0.2)
        return False

    def _click_force(self, page: Any, selector_key: str, label: str, timeout: int = 10000) -> None:
        selector = self._selector(selector_key)
        locator = self._visible_locator(page, selector, label, timeout=timeout)
        if locator is None:
            raise RuntimeError(f"找不到{label}: selectors.{selector_key}")
        try:
            self._js_click(locator)
        except Exception:
            pass

    def _frame_with_selectors(self, page: Any, selector_keys: tuple[str, ...], timeout: int = 5000) -> Any | None:
        selectors = [self._selector_optional(key) for key in selector_keys]
        if any(not selector for selector in selectors):
            return None

        deadline = time.time() + timeout / 1000
        while time.time() < deadline:
            for frame in getattr(page, "frames", []) or []:
                try:
                    if frame is getattr(page, "main_frame", None):
                        continue
                    matched = True
                    for selector in selectors:
                        locator = self._scoped_visible_locator(frame, selector, timeout=200)
                        if locator is None:
                            matched = False
                            break
                    if matched:
                        return frame
                except Exception:
                    continue
            time.sleep(0.2)
        return None

    def _frame_is_displayed(self, frame: Any) -> bool:
        current = frame
        while getattr(current, "parent_frame", None) is not None:
            try:
                if not self._element_is_displayed(current.frame_element()):
                    return False
            except Exception:
                return False
            current = current.parent_frame
        return True

    def _scope_fill(self, scope: Any, selector_key: str, value: str, label: str, timeout: int = 10000) -> None:
        selector = self._selector(selector_key)
        locator = self._scoped_visible_locator(scope, selector, timeout=timeout)
        if locator is None:
            raise RuntimeError(f"找不到{label}: selectors.{selector_key}")
        locator.fill(value, timeout=timeout)

    def _scope_click(self, scope: Any, selector_key: str, label: str, timeout: int = 10000) -> None:
        selector = self._selector(selector_key)
        locator = self._scoped_visible_locator(scope, selector, timeout=timeout)
        if locator is None:
            raise RuntimeError(f"找不到{label}: selectors.{selector_key}")
        if not self._js_click(locator):
            raise RuntimeError(f"点击{label}失败: selectors.{selector_key}")

    def _fill(self, page: Any, selector_key: str, value: str, label: str, timeout: int = 10000) -> None:
        selector = self._selector(selector_key)
        locator = self._visible_locator(page, selector, label, timeout=timeout)
        if locator is None:
            raise RuntimeError(f"找不到{label}: selectors.{selector_key}")
        readonly = False
        try:
            readonly = bool(locator.evaluate("(el) => !!el.readOnly || el.hasAttribute('readonly')"))
        except Exception:
            pass
        if readonly:
            if self._set_input_value(locator, value):
                return
        if self._set_input_value(locator, value):
            return
        try:
            locator.fill(value, timeout=min(timeout, 2000))
            return
        except Exception:
            raise

    def _set_input_value(self, locator: Any, value: str) -> bool:
        try:
            locator.evaluate(
                """
                (el, nextValue) => {
                  const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
                  setter.call(el, nextValue);
                  el.dispatchEvent(new Event('input', { bubbles: true }));
                  el.dispatchEvent(new Event('change', { bubbles: true }));
                }
                """,
                value,
                timeout=800,
            )
            return locator.input_value(timeout=800) == value
        except Exception:
            return False

    def _selector_visible(self, page: Any, selector_key: str, timeout: int = 1000) -> bool:
        selector = self._selector_optional(selector_key)
        if not selector:
            return False
        return self._visible_locator(page, selector, selector_key, timeout=timeout) is not None

    def _iter_scopes(self, page: Any) -> Iterable[Any]:
        yield page
        skip = ("alicdn.com", "1688.com", "taobao.com", "mmstat.com", "google.", "chrome-extension:")
        seen: set[int] = {id(page)}
        main = getattr(page, "main_frame", None)
        if main is not None:
            seen.add(id(main))
        visible_srcs = self._visible_iframe_srcs(page)
        for frame in getattr(page, "frames", []) or []:
            if id(frame) in seen:
                continue
            url = str(getattr(frame, "url", "") or "")
            if not url or url == "about:blank":
                continue
            if any(token in url for token in skip):
                continue
            keep = any(
                token in url
                for token in (
                    "ai_tj_inventory_3",
                    "inventory_realtime_search",
                    "purchase_order_list_v4",
                    "purchase_transfer_order_list_v4",
                    "merchandise_channel_store",
                )
            )
            if visible_srcs and not keep:
                if not any(src and (src in url or url in src) for src in visible_srcs):
                    continue
            yield frame

    def _visible_iframe_srcs(self, page: Any) -> list[str]:
        try:
            values = page.evaluate(
                """
                () => Array.from(document.querySelectorAll('iframe')).filter((el) => {
                  const rect = el.getBoundingClientRect();
                  const style = window.getComputedStyle(el);
                  return rect.width > 80 && rect.height > 80 &&
                    style.display !== 'none' && style.visibility !== 'hidden';
                }).map((el) => el.src || '')
                """
            )
        except Exception:
            return []
        return [str(item) for item in (values or []) if str(item).strip()]

    def _element_is_displayed(self, locator: Any) -> bool:
        try:
            return bool(
                locator.evaluate(
                    """
                    (el) => {
                      const rect = el.getBoundingClientRect();
                      const style = window.getComputedStyle(el);
                      return rect.width > 0 && rect.height > 0 &&
                        style.display !== 'none' &&
                        style.visibility !== 'hidden' &&
                        Number(style.opacity || 1) > 0;
                    }
                    """,
                    timeout=800,
                )
            )
        except Exception:
            return False

    def _js_click(self, locator: Any) -> bool:
        try:
            locator.evaluate("(el) => el.click()", timeout=800)
            return True
        except Exception:
            return False

    def _js_click_matching_text(
        self,
        page: Any,
        texts: Iterable[str],
        overlay_only: bool = False,
        exact: bool = True,
    ) -> str:
        needles = [_clean_text(text) for text in texts if _clean_text(text)]
        if not needles:
            return ""
        script = """
        ({needles, overlayOnly, exact}) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 &&
              style.visibility !== 'hidden' && style.display !== 'none' &&
              Number(style.opacity || 1) > 0;
          };
          const textOf = (el) => (el.innerText || el.textContent || '').replace(/\\s+/g, ' ').trim();
          const roots = overlayOnly
            ? Array.from(document.querySelectorAll('.next-overlay-wrapper.opened, .next-overlay-wrapper, .next-menu, .next-overlay-inner, .next-balloon, [role="menu"], [role="listbox"], .next-select-menu'))
            : [document];
          const nodesSel = overlayOnly
            ? '.next-menu-item, [role="menuitem"], [role="option"], li, button, a, span'
            : 'button, a, [role="button"], li, span, div, label';
          for (const root of roots) {
            if (!root) continue;
            const nodes = Array.from(root.querySelectorAll(nodesSel));
            let best = null;
            let bestLen = 1e9;
            let bestLabel = '';
            for (const el of nodes) {
              if (!visible(el)) continue;
              const label = textOf(el);
              if (!label || label.length > 40) continue;
              const hit = needles.find((needle) => exact ? (label === needle || label.includes(needle)) : label.includes(needle));
              if (!hit) continue;
              const len = el.querySelectorAll('*').length;
              if (len < bestLen) {
                best = el;
                bestLen = len;
                bestLabel = label;
              }
            }
            if (best) {
              best.click();
              return bestLabel;
            }
          }
          return '';
        }
        """
        for scope in self._iter_scopes(page):
            try:
                hit = scope.evaluate(
                    script,
                    {"needles": needles, "overlayOnly": overlay_only, "exact": exact},
                )
            except Exception:
                continue
            if hit:
                return str(hit)
        return ""

    def _visible_locator(self, page: Any, selector: str, label: str, timeout: int = 10000) -> Any | None:
        pw_selector = _pw_selector(selector)
        deadline = time.time() + timeout / 1000
        while time.time() < deadline:
            for scope in self._iter_scopes(page):
                try:
                    locator = scope.locator(pw_selector)
                    count = min(locator.count(), 100)
                except Exception:
                    continue
                for idx in range(count):
                    item = locator.nth(idx)
                    if self._element_is_displayed(item):
                        return item
            time.sleep(0.2)
        return None

    def _scoped_visible_locator(self, scope: Any, selector: str, timeout: int = 10000) -> Any | None:
        pw_selector = _pw_selector(selector)
        deadline = time.time() + timeout / 1000
        while time.time() < deadline:
            try:
                locator = scope.locator(pw_selector)
                count = min(locator.count(), 100)
            except Exception:
                time.sleep(0.2)
                continue
            for idx in range(count):
                item = locator.nth(idx)
                if self._element_is_displayed(item):
                    return item
            time.sleep(0.2)
        return None

    def _click_text(
        self,
        page: Any,
        text: str,
        timeout: int = 5000,
        optional: bool = False,
        loose: bool = False,
    ) -> bool:
        deadline = time.time() + timeout / 1000
        while time.time() < deadline:
            hit = self._js_click_matching_text(page, [text], overlay_only=True, exact=not loose)
            if hit:
                return True
            hit = self._js_click_matching_text(page, [text], overlay_only=False, exact=not loose)
            if hit:
                return True
            time.sleep(0.2)
        if optional:
            return False
        raise RuntimeError(f"找不到文本选项: {text}")

    def _compact_text(self, value: str) -> str:
        text = _clean_text(value)
        if not text:
            return ""
        return re.sub(r"[\s\-_/／·—–,，.。()（）\[\]【】]+", "", text)

    def _page_has_text(self, page: Any, text: str, timeout: int = 1000) -> bool:
        needle = _clean_text(text)
        if not needle:
            return False
        deadline = time.time() + timeout / 1000
        script = "(needle) => ((document.body && document.body.innerText) || '').includes(needle)"
        while time.time() < deadline:
            for scope in self._iter_scopes(page):
                try:
                    if scope.evaluate(script, needle):
                        return True
                except Exception:
                    continue
            time.sleep(0.2)
        return False

    def _wait_quiet(self, page: Any, timeout_ms: int = 5000) -> None:
        ms = max(0, min(int(timeout_ms), 8000))
        try:
            page.wait_for_timeout(ms)
        except Exception:
            time.sleep(ms / 1000)

    def _screenshot(self, page: Any, name: str) -> Path:
        self.settings.screenshot_dir.mkdir(parents=True, exist_ok=True)
        path = self.settings.screenshot_dir / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{_slug(name)}.png"
        try:
            session = page.context.new_cdp_session(page)
            result = session.send("Page.captureScreenshot", {"format": "png", "fromSurface": True})
            data = result.get("data") or ""
            if data:
                path.write_bytes(base64.b64decode(data))
                if path.stat().st_size > 0:
                    print(f"[猫超] 已保存错误截图: {path} ({path.stat().st_size} bytes)")
                    return path
        except Exception as exc:
            print(f"[猫超] CDP 截图失败: {exc}")
        return path

    def _capture_error_screenshot(self, page: Any | None, name: str) -> str:
        if page is None:
            return ""
        try:
            shot = self._screenshot(page, name)
            print(f"[猫超] 已保存错误截图: {shot}")
            return str(shot)
        except Exception as exc:
            print(f"[猫超] 截图失败: {exc}")
            return ""

    def _write_manifest(self, results: list[RunResult]) -> None:
        self.settings.log_dir.mkdir(parents=True, exist_ok=True)
        path = self.settings.log_dir / f"maochao_download_manifest_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        payload = {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "config": str(self.settings.config_path),
            "results": [asdict(item) for item in results],
        }
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[猫超] manifest 已写入: {path}")


def dry_run(settings: Settings, tasks: list[str], account_keys: list[str] | None = None) -> int:
    accounts = settings.accounts
    if account_keys:
        wanted = set(account_keys)
        accounts = [account for account in accounts if account.key in wanted]
    print(f"[dry-run] config: {settings.config_path}")
    print(f"[dry-run] accounts_source: {settings.accounts_source}")
    print(f"[dry-run] accounts_db: {settings.accounts_db_path}")
    print(f"[dry-run] login_url: {settings.login_url or '未配置'}")
    print(f"[dry-run] tasks: {', '.join(tasks)}")
    print(f"[dry-run] accounts: {len(accounts)}")
    seen_ports: dict[int, str] = {}
    for account in accounts:
        task_list = [task for task in tasks if task in account.tasks]
        password_status = "已填" if account.password else "未填"
        print(
            f"  - {account.key}: port={account.port}, tasks={task_list or '无'}, "
            f"profile={account.profile_dir}, password={password_status}"
        )
        if account.port in seen_ports:
            print(f"[dry-run] 端口冲突: {seen_ports[account.port]} 与 {account.key} 都使用 {account.port}")
            return 1
        seen_ports[account.port] = account.key

    missing = []
    for task in tasks:
        for selector_key in REQUIRED_SELECTORS[task]:
            for account in accounts:
                if task not in account.tasks:
                    continue
                context = account_context(
                    {
                        "key": account.key,
                        "name": account.name,
                        "username": account.username,
                        "password": account.password,
                        "port": account.port,
                        "profile_dir": str(account.profile_dir),
                        "download_dir": str(account.download_dir),
                        "supplier_names": account.supplier_names,
                        "tasks": account.tasks,
                        "note": account.note,
                        "xpath_vars": account.xpath_vars,
                        "selector_overrides": account.selector_overrides,
                        "enabled": account.enabled,
                    }
                )
                rendered = deep_merge(render_tree(settings.selectors, context), render_tree(account.selector_overrides, context))
                node: Any = rendered
                ok = True
                for part in selector_key.split("."):
                    if not isinstance(node, dict) or part not in node or not _clean_text(node.get(part)):
                        ok = False
                        break
                    node = node[part]
                if not ok:
                    missing.append(f"{task}: {account.key}: selectors.{selector_key}")
                for unresolved_key in unresolved_templates(rendered):
                    missing.append(f"{task}: {account.key}: template 未展开 -> {unresolved_key}")
    if missing:
        print("[dry-run] 缺少以下 XPath/selector：")
        for item in sorted(set(missing)):
            print(f"  - {item}")
        return 1
    print("[dry-run] 基础配置检查通过。")
    return 0


def add_common_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument(
        "--config",
        default=str(DEFAULT_CONFIG),
        help="配置文件路径，默认使用 config.example.json；实际运行建议复制为 config.local.json",
    )
    parser.add_argument("--task", action="append", help="任务 key/编号，可重复；默认 all")
    parser.add_argument("--account", action="append", help="账号 key，可重复；默认全部配置账号")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="猫超补货数据源下载 RPA")
    sub = parser.add_subparsers(dest="cmd")

    p_dry = sub.add_parser("dry-run", help="只检查配置和 XPath 完整性")
    add_common_args(p_dry)

    p_login = sub.add_parser("login", help="打开/接管账号浏览器，用于首次人工登录")
    add_common_args(p_login)
    p_login.add_argument("--headed", action="store_true", help="强制显示浏览器窗口")

    p_run = sub.add_parser("run", help="顺序执行猫超数据源下载")
    add_common_args(p_run)
    p_run.add_argument("--manual-login", action="store_true", help="登录后暂停，方便处理验证码/手机确认")
    p_run.add_argument("--headed", action="store_true", help="强制显示浏览器窗口")
    p_run.add_argument("--force-account-tasks", action="store_true", help="指定账号时临时执行命令行任务，忽略账号库任务归属")
    p_run.add_argument("--supplier-id", action="append", help="右上角供应商 ID，可重复；一个供应商会依次执行所选任务")
    p_run.add_argument("--supplier-name", action="append", help="右上角供应商名称，可重复；仅在未提供 ID 时作为辅助匹配")
    p_run.add_argument("--use-current-supplier", action="store_true", help="不切换供应商，只处理当前右上角已选中的供应商")

    p_sync = sub.add_parser("sync-suppliers", help="登录后读取右上角可切换供应商清单")
    add_common_args(p_sync)
    p_sync.add_argument("--headed", action="store_true", help="强制显示浏览器窗口")

    args = parser.parse_args(argv)
    if not args.cmd:
        parser.print_help()
        return 0

    settings = load_settings(Path(args.config))
    tasks = selected_tasks(args.task)

    if args.cmd == "dry-run":
        return dry_run(settings, tasks, args.account)

    headless = False if getattr(args, "headed", False) else None
    rpa = MaochaoRPA(
        settings,
        manual_login=getattr(args, "manual_login", False),
        headless=headless,
    )
    if args.cmd == "login":
        rpa.manual_login = True
        rpa.login_only(args.account)
        return 0
    if args.cmd == "sync-suppliers":
        rows = rpa.sync_header_suppliers(args.account)
        print(f"[猫超] 同步到 {len(rows)} 个供应商")
        for item in rows:
            print(f"  - {item['account_key']}: {item['supplier_id']} / {item['supplier_name']}")
        return 0
    if args.cmd == "run":
        suppliers = []
        ids = getattr(args, "supplier_id", None) or []
        names = getattr(args, "supplier_name", None) or []
        for supplier_id, supplier_name in zip(ids, names + [""] * len(ids)):
            suppliers.append({"supplier_id": supplier_id, "supplier_name": supplier_name})
        if len(names) > len(ids):
            for supplier_name in names[len(ids):]:
                suppliers.append({"supplier_id": "", "supplier_name": supplier_name})
        results = rpa.run(
            tasks,
            args.account,
            force_account_tasks=getattr(args, "force_account_tasks", False) or bool(suppliers) or getattr(args, "use_current_supplier", False),
            suppliers=suppliers or None,
            use_current_supplier=getattr(args, "use_current_supplier", False),
        )
        failed = [item for item in results if item.status != "ok"]
        print(f"[猫超] 执行完成: ok={len(results) - len(failed)}, failed={len(failed)}")
        return 1 if failed else 0

    parser.print_help()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
