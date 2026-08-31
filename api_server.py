from __future__ import annotations

import json
import os
import re
import secrets
import socket
import sys
import tempfile
import zipfile
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import unquote

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, PlainTextResponse, Response
from pydantic import BaseModel, Field
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

from account_store import AccountStore
from backend_core import (
    DEFAULT_OPERATOR_PASSWORD,
    RUN_KIND_SYNC_SUPPLIERS,
    RUN_KIND_TASKS,
    BackendStore,
    DEFAULT_CONFIG_PATH,
)
from maochao_rpa import load_settings, normalize_task_name, selected_tasks


app = FastAPI(title="Maochao RPA Backend", version="0.1.0")
store = BackendStore(DEFAULT_CONFIG_PATH)
WEB_ROOT = Path(__file__).resolve().parent / "web"
ADMIN_USERNAME = os.environ.get("RPA_ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("RPA_ADMIN_PASSWORD", "admin123")
AUTH_SESSIONS: dict[str, dict[str, Any]] = {}


def _new_session(payload: dict[str, Any]) -> str:
    token = secrets.token_urlsafe(32)
    AUTH_SESSIONS[token] = payload
    return token


def _revoke_supply_chain_sessions(user_id: str) -> None:
    for token, session in list(AUTH_SESSIONS.items()):
        if session.get("role") == "supply_chain" and session.get("user_id") == user_id:
            AUTH_SESSIONS.pop(token, None)


def _bearer_token(request: Request) -> str:
    header = request.headers.get("authorization") or ""
    if not header.lower().startswith("bearer "):
        return ""
    return header[7:].strip()


def _session_from_request(request: Request) -> dict[str, Any] | None:
    token = _bearer_token(request)
    if not token:
        return None
    session = AUTH_SESSIONS.get(token)
    if session is None:
        return None
    return {**session, "token": token}


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    public_paths = {"/api/auth/login", "/api/auth/operators", "/api/health", "/api/worker", "/health"}
    if path.startswith("/api/") and path not in public_paths:
        session = _session_from_request(request)
        if session is None:
            return JSONResponse({"detail": "未登录"}, status_code=401)
        if session.get("role") == "supply_chain":
            user = store.get_supply_chain_user(str(session.get("user_id") or ""))
            if user is None or not user.get("enabled"):
                AUTH_SESSIONS.pop(str(session.get("token") or ""), None)
                return JSONResponse({"detail": "供应链账号已停用"}, status_code=401)
        request.state.user = session
    return await call_next(request)


@app.get("/", include_in_schema=False, response_class=HTMLResponse)
def root() -> HTMLResponse:
    index_path = WEB_ROOT / "index.html"
    if not index_path.is_file():
        return HTMLResponse(
            "<h1>猫超 RPA Web 管理台</h1><p>前端文件尚未部署，请检查 web/index.html。</p>",
            status_code=503,
        )
    return HTMLResponse(
        index_path.read_text(encoding="utf-8"),
        headers={"Cache-Control": "no-store"},
    )


class AccountPayload(BaseModel):
    key: str | None = None
    name: str | None = None
    username: str | None = None
    password: str | None = None
    port: int | None = None
    profile_dir: str | None = None
    download_dir: str | None = None
    supplier_names: list[str] = Field(default_factory=list)
    tasks: list[str] = Field(default_factory=list)
    note: str | None = ""
    xpath_vars: dict[str, Any] = Field(default_factory=dict)
    selector_overrides: dict[str, Any] = Field(default_factory=dict)
    enabled: bool = True


class AccountPatch(BaseModel):
    name: str | None = None
    username: str | None = None
    password: str | None = None
    port: int | None = None
    profile_dir: str | None = None
    download_dir: str | None = None
    supplier_names: list[str] | None = None
    tasks: list[str] | None = None
    note: str | None = None
    xpath_vars: dict[str, Any] | None = None
    selector_overrides: dict[str, Any] | None = None
    enabled: bool | None = None


class AuthLogin(BaseModel):
    role: str = "member"
    username: str = ""
    password: str = ""
    operator_id: str = ""


class SupplierRefPayload(BaseModel):
    account_key: str = ""
    supplier_id: str = ""
    supplier_name: str = ""


class OperatorCreate(BaseModel):
    name: str
    supply_chain_user_id: str = ""


class OperatorPatch(BaseModel):
    name: str | None = None
    supply_chain_user_id: str | None = None
    active: bool | None = None


class SupplyChainUserCreate(BaseModel):
    username: str
    name: str
    password: str = DEFAULT_OPERATOR_PASSWORD


class SupplyChainUserPatch(BaseModel):
    username: str | None = None
    name: str | None = None
    password: str | None = None
    enabled: bool | None = None


class OperatorSupplierAssign(BaseModel):
    account_key: str
    supplier_ids: list[str] = Field(default_factory=list)


class OperatorPasswordChange(BaseModel):
    old_password: str = ""
    new_password: str = ""


class RunCreate(BaseModel):
    task_keys: list[str] = Field(default_factory=list)
    account_keys: list[str] = Field(default_factory=list)
    force_account_tasks: bool = False
    headed: bool = True
    operator_id: str = ""
    suppliers: list[SupplierRefPayload] = Field(default_factory=list)
    run_kind: str = RUN_KIND_TASKS


class SchedulePayload(BaseModel):
    task_keys: list[str] = Field(default_factory=list)
    operator_id: str = ""
    operator_ids: list[str] = Field(default_factory=list)
    all_operators: bool = False
    enabled: bool = True
    time_of_day: str = "09:00"
    weekdays: list[int] = Field(default_factory=lambda: list(range(7)))
    headed: bool = True

    def normalized_operator_ids(self) -> list[str]:
        if self.operator_ids:
            return self.operator_ids
        return [self.operator_id] if self.operator_id else []


def _public_account(account: dict[str, Any]) -> dict[str, Any]:
    item = dict(account)
    item["username_set"] = bool(item.get("username"))
    item["password_set"] = bool(item.get("password"))
    item["username"] = item.get("username") or ""
    item.pop("password", None)
    return item


def _slug(value: str, fallback: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_]+", "_", value.strip()).strip("_").lower()
    return text or fallback


def _next_account_key(accounts: list[dict[str, Any]], username: str) -> str:
    base = f"tmall_{_slug(username, 'account')}"
    used = {account["key"] for account in accounts}
    if base not in used:
        return base
    index = 2
    while f"{base}_{index:02d}" in used:
        index += 1
    return f"{base}_{index:02d}"


def _next_port(accounts: list[dict[str, Any]], start: int = 9231) -> int:
    used = {int(account["port"]) for account in accounts if account.get("port")}
    port = start
    while port in used:
        port += 1
    return port


def _resolve_file_id(root: Path, file_id: str) -> Path:
    path = (root / file_id).resolve()
    if not path.is_file() or root.resolve() not in path.parents:
        raise HTTPException(status_code=404, detail="file not found")
    return path


def _run_output_files(run_item: dict[str, Any]) -> list[Path]:
    files: list[Path] = []
    seen: set[Path] = set()
    for result in run_item.get("result", []):
        for field in ("raw_file", "cleaned_file"):
            value = result.get(field)
            if not value:
                continue
            path = Path(str(value)).expanduser().resolve()
            if path.is_file() and path not in seen:
                seen.add(path)
                files.append(path)
    return files


def _file_run_index() -> dict[str, str]:
    index: dict[str, str] = {}
    for run_item in store.list_runs():
        for path in _run_output_files(run_item):
            index[str(path)] = run_item["run_id"]
    return index


def _is_member(request: Request) -> bool:
    return (getattr(request.state, "user", {}) or {}).get("role") == "member"


def _is_supply_chain(request: Request) -> bool:
    return (getattr(request.state, "user", {}) or {}).get("role") == "supply_chain"


def _current_user_id(request: Request) -> str:
    return str((getattr(request.state, "user", {}) or {}).get("user_id") or "")


def _require_admin(request: Request) -> None:
    if (getattr(request.state, "user", {}) or {}).get("role") != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可执行此操作")


def _require_owned_operator(operator_id: str, request: Request) -> None:
    if not _is_supply_chain(request):
        return
    operator = store.get_operator(operator_id)
    if operator is None or operator.get("supply_chain_user_id") != _current_user_id(request):
        raise HTTPException(status_code=403, detail="只能管理自己负责的运营组员")


def _member_operator_id(request: Request) -> str:
    return str((getattr(request.state, "user", {}) or {}).get("operator_id") or "")


def _normalized_file_id(value: str) -> str:
    return str(value or "").replace("\\", "/").strip("/")


def _member_files(rows: list[dict[str, Any]], request: Request) -> list[dict[str, Any]]:
    if _is_supply_chain(request):
        owned = {
            item["operator_id"] for item in store.list_operators()
            if item.get("supply_chain_user_id") == _current_user_id(request)
        }
        return [row for row in rows if row.get("operator_id") in owned]
    if not _is_member(request):
        return rows
    operator_id = _member_operator_id(request)
    return [row for row in rows if row.get("operator_id") == operator_id]


def _member_runs(rows: list[dict[str, Any]], request: Request) -> list[dict[str, Any]]:
    if _is_supply_chain(request):
        owned = {
            item["operator_id"] for item in store.list_operators()
            if item.get("supply_chain_user_id") == _current_user_id(request)
        }
        return [row for row in rows if row.get("operator_id") in owned]
    if not _is_member(request):
        return rows
    operator_id = _member_operator_id(request)
    return [row for row in rows if row.get("operator_id") == operator_id]


def _assert_member_run(run_item: dict[str, Any], request: Request) -> None:
    if _is_supply_chain(request):
        _require_owned_operator(str(run_item.get("operator_id") or ""), request)
        return
    if _is_member(request) and run_item.get("operator_id") != _member_operator_id(request):
        raise HTTPException(status_code=404, detail="run not found")


def _get_member_checked_run(run_id: str, request: Request) -> dict[str, Any]:
    try:
        run_item = store.get_run(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc
    _assert_member_run(run_item, request)
    return run_item


CODE_REVISION = "2026-08-31-channel-goods-item-filter-v52.32"


@app.get("/health")
@app.get("/api/health")
def health() -> dict[str, Any]:
    return {
        "status": "ok",
        "pid": os.getpid(),
        "config": str(store.config_path),
        "db": str(store.settings.accounts_db_path),
        "tasks": len(store.list_tasks()),
        "code_revision": CODE_REVISION,
    }


@app.get("/api/ready")
def ready() -> dict[str, Any]:
    checks = {
        "config_exists": store.config_path.exists(),
        "accounts_db_exists": store.settings.accounts_db_path.exists(),
        "data_root_exists": store.settings.data_root.exists(),
        "log_dir_exists": store.settings.log_dir.exists(),
        "screenshot_dir_exists": store.settings.screenshot_dir.exists(),
    }
    return {"status": "ok" if all(checks.values()) else "degraded", "checks": checks}


@app.get("/api/auth/operators")
def auth_operators() -> list[dict[str, Any]]:
    return [
        {"user_id": item["user_id"], "username": item["username"], "name": item["name"]}
        for item in store.list_supply_chain_users()
    ]


@app.post("/api/auth/login")
def auth_login(payload: AuthLogin) -> dict[str, Any]:
    role = (payload.role or "supply_chain").strip().lower()
    if role == "admin":
        username = payload.username.strip()
        if username != ADMIN_USERNAME or payload.password != ADMIN_PASSWORD:
            raise HTTPException(status_code=401, detail="管理员账号或密码错误")
        user = {"role": "admin", "username": username, "operator_id": "", "operator_name": "管理员"}
        token = _new_session(user)
        return {"token": token, "user": user}
    if role == "supply_chain":
        user_item = store.authenticate_supply_chain_user(payload.username, payload.password)
        if user_item is None:
            raise HTTPException(status_code=401, detail="供应链账号或密码错误")
        user = {
            "role": "supply_chain",
            "user_id": user_item["user_id"],
            "username": user_item["username"],
            "name": user_item["name"],
        }
        token = _new_session(user)
        return {"token": token, "user": user}
    raise HTTPException(status_code=400, detail="未知登录类型")


@app.get("/api/auth/me")
def auth_me(request: Request) -> dict[str, Any]:
    user = dict(_session_from_request(request) or getattr(request.state, "user", {}) or {})
    user.pop("token", None)
    return {"user": user}


@app.post("/api/auth/logout")
def auth_logout(request: Request) -> dict[str, str]:
    token = _bearer_token(request)
    if token:
        AUTH_SESSIONS.pop(token, None)
    return {"status": "ok"}


@app.get("/api/worker")
def worker() -> dict[str, Any]:
    return store.worker_status()


@app.get("/api/tasks")
def tasks() -> list[dict[str, Any]]:
    return store.list_tasks()


@app.get("/api/accounts")
def accounts(request: Request, include_disabled: bool = False) -> list[dict[str, Any]]:
    locks = {lock["account_key"]: lock for lock in store.list_account_locks()}
    rows: list[dict[str, Any]] = []
    for account in store.list_accounts(include_disabled=include_disabled):
        item = _public_account(account)
        lock = locks.get(account["key"])
        item["browser_status"] = "占用中" if lock else "空闲"
        item["locked_by_run_id"] = lock["run_id"] if lock else ""
        owner = store.get_account_owner(account["key"])
        item.update(owner)
        item["can_edit"] = not _is_supply_chain(request) or owner.get("creator_user_id") == _current_user_id(request)
        item["can_delete"] = not _is_supply_chain(request)
        rows.append(item)
    return rows


def _port_open(port: int) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.settimeout(0.3)
        return sock.connect_ex(("127.0.0.1", port)) == 0


def _close_browser_via_cdp(port: int) -> None:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError as exc:
        raise RuntimeError("关闭浏览器需要 Playwright 运行时") from exc
    with sync_playwright() as playwright:
        browser = playwright.chromium.connect_over_cdp(
            f"http://127.0.0.1:{port}",
            timeout=5000,
        )
        browser.close()


@app.post("/api/browsers/close-idle")
def close_idle_browsers() -> dict[str, Any]:
    items: list[dict[str, Any]] = []
    closed = 0
    locked_accounts = {item["account_key"] for item in store.list_account_locks()}
    for account in store.list_accounts(include_disabled=True):
        port = int(account.get("port") or 0)
        if port <= 0:
            continue
        if account.get("key") in locked_accounts:
            items.append({"account_key": account.get("key", ""), "port": port, "status": "in_use"})
            continue
        if not _port_open(port):
            items.append({"account_key": account.get("key", ""), "port": port, "status": "already_closed"})
            continue
        try:
            _close_browser_via_cdp(port)
        except Exception as exc:
            items.append({"account_key": account.get("key", ""), "port": port, "status": "close_failed", "error": str(exc)})
            continue
        closed += 1
        items.append({"account_key": account.get("key", ""), "port": port, "status": "closed"})
    return {"closed": closed, "items": items}


@app.post("/api/accounts")
def create_account(payload: AccountPayload, request: Request) -> dict[str, Any]:
    account_store = AccountStore(store.settings.accounts_db_path, store.settings.accounts_db_key_path)
    existing_accounts = store.list_accounts(include_disabled=True)
    data = payload.dict(exclude_none=True)
    username = str(data.get("username") or "").strip()
    if not username:
        raise HTTPException(status_code=400, detail="username is required")
    data.setdefault("key", _next_account_key(existing_accounts, username))
    data.setdefault("name", username)
    data.setdefault("port", _next_port(existing_accounts))
    data.setdefault("profile_dir", f"./browser_profiles/{data['key']}")
    data.setdefault("download_dir", f"./downloads/{data['key']}")
    data["tasks"] = [normalize_task_name(task) for task in data.get("tasks", [])]
    account_store.upsert_account(data, base_dir=store.config_path.parent)
    role = "supply_chain" if _is_supply_chain(request) else "admin"
    store.set_account_owner(data["key"], role, _current_user_id(request) if role == "supply_chain" else "")
    store.settings = load_settings(store.config_path)
    return {"status": "ok", "account_key": data["key"]}


@app.patch("/api/accounts/{account_key}")
def patch_account(account_key: str, payload: AccountPatch, request: Request) -> dict[str, Any]:
    existing = None
    for account in store.list_accounts(include_disabled=True):
        if account["key"] == account_key:
            existing = account
            break
    if existing is None:
        raise HTTPException(status_code=404, detail="account not found")
    if _is_supply_chain(request):
        owner = store.get_account_owner(account_key)
        if owner.get("creator_role") != "supply_chain" or owner.get("creator_user_id") != _current_user_id(request):
            raise HTTPException(status_code=403, detail="只能编辑自己新增的猫超账号")
    data = dict(existing)
    for key, value in payload.dict(exclude_unset=True).items():
        if value is not None:
            data[key] = value
    data["tasks"] = [normalize_task_name(task) for task in data.get("tasks", [])]
    account_store = AccountStore(store.settings.accounts_db_path, store.settings.accounts_db_key_path)
    account_store.upsert_account(data, base_dir=store.config_path.parent)
    store.settings = load_settings(store.config_path)
    return {"status": "ok", "account_key": account_key}


@app.delete("/api/accounts/{account_key}")
def delete_account(account_key: str, request: Request) -> dict[str, Any]:
    _require_admin(request)
    try:
        store.delete_account(account_key)
        store.settings = load_settings(store.config_path)
        return {"status": "ok", "account_key": account_key}
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="account not found") from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.get("/api/operators")
def operators() -> list[dict[str, Any]]:
    return store.list_operators()


@app.get("/api/supply-chain-users")
def supply_chain_users(request: Request, include_disabled: bool = True) -> list[dict[str, Any]]:
    _require_admin(request)
    return store.list_supply_chain_users(include_disabled=include_disabled)


@app.post("/api/supply-chain-users")
def create_supply_chain_user(payload: SupplyChainUserCreate, request: Request) -> dict[str, Any]:
    _require_admin(request)
    try:
        return store.create_supply_chain_user(payload.username, payload.name, payload.password)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.patch("/api/supply-chain-users/{user_id}")
def patch_supply_chain_user(user_id: str, payload: SupplyChainUserPatch, request: Request) -> dict[str, Any]:
    _require_admin(request)
    try:
        result = store.update_supply_chain_user(user_id, **payload.dict(exclude_unset=True))
        if payload.enabled is False or payload.password is not None:
            _revoke_supply_chain_sessions(user_id)
        return result
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="供应链账号不存在") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/supply-chain-users/password/change")
def change_supply_chain_password(payload: OperatorPasswordChange, request: Request) -> dict[str, str]:
    if not _is_supply_chain(request):
        raise HTTPException(status_code=403, detail="仅供应链账号可修改自己的密码")
    user_id = _current_user_id(request)
    current = store.get_supply_chain_user(user_id)
    if current is None or store.authenticate_supply_chain_user(current["username"], payload.old_password) is None:
        raise HTTPException(status_code=400, detail="原密码错误")
    if not payload.new_password:
        raise HTTPException(status_code=400, detail="新密码不能为空")
    store.update_supply_chain_user(user_id, password=payload.new_password)
    return {"status": "ok"}


@app.delete("/api/supply-chain-users/{user_id}")
def delete_supply_chain_user(user_id: str, request: Request) -> dict[str, Any]:
    _require_admin(request)
    try:
        result = store.delete_supply_chain_user(user_id)
        _revoke_supply_chain_sessions(user_id)
        return result
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="供应链账号不存在") from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/operators")
def create_operator(payload: OperatorCreate, request: Request) -> dict[str, Any]:
    try:
        owner_id = _current_user_id(request) if _is_supply_chain(request) else payload.supply_chain_user_id
        return store.create_operator(payload.name, owner_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.patch("/api/operators/{operator_id}")
def patch_operator(operator_id: str, payload: OperatorPatch, request: Request) -> dict[str, Any]:
    _require_owned_operator(operator_id, request)
    values = payload.dict(exclude_unset=True)
    if _is_supply_chain(request):
        values.pop("supply_chain_user_id", None)
    try:
        return store.update_operator(operator_id, **values)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="operator not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/operators/password/change")
def change_operator_password(payload: OperatorPasswordChange, request: Request) -> dict[str, str]:
    raise HTTPException(status_code=410, detail="运营组员不再使用登录密码")


@app.post("/api/operators/{operator_id}/password/reset")
def reset_operator_password(operator_id: str) -> dict[str, Any]:
    raise HTTPException(status_code=410, detail="运营组员不再使用登录密码")


@app.delete("/api/operators/{operator_id}")
def delete_operator(operator_id: str, request: Request) -> dict[str, Any]:
    _require_owned_operator(operator_id, request)
    try:
        deleted = store.delete_operator(operator_id)
        return {"status": "ok", "operator": deleted}
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="operator not found") from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.get("/api/operators/{operator_id}/suppliers")
def operator_suppliers(operator_id: str, request: Request, account_key: str = "") -> list[dict[str, Any]]:
    if _is_member(request) and operator_id != _member_operator_id(request):
        raise HTTPException(status_code=403, detail="只能查看自己的供应商")
    if store.get_operator(operator_id) is None:
        raise HTTPException(status_code=404, detail="operator not found")
    return store.list_operator_suppliers(operator_id, account_key)


@app.put("/api/operators/{operator_id}/suppliers")
def assign_operator_suppliers(operator_id: str, payload: OperatorSupplierAssign, request: Request) -> list[dict[str, Any]]:
    _require_owned_operator(operator_id, request)
    try:
        return store.set_operator_suppliers(operator_id, payload.account_key, payload.supplier_ids)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="operator not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.get("/api/suppliers")
def suppliers(account_key: str = "", include_hidden: bool = False) -> list[dict[str, Any]]:
    return store.list_account_suppliers(account_key, include_hidden=include_hidden)


@app.get("/api/item-id-config")
def item_id_config(request: Request) -> dict[str, Any]:
    _require_admin(request)
    return {"rows": store.list_item_id_config(), "uploads": store.list_item_id_uploads()}


@app.get("/api/item-id-config/template")
def item_id_config_template(request: Request) -> Response:
    _require_admin(request)
    workbook = Workbook()
    guide = workbook.active
    guide.title = "使用说明"
    guide["A1"] = "货品 ID 配置导入教程"
    guide.merge_cells("A1:H1")
    guide["A3"] = "1. 获取猫超账号标识"
    guide["A4"] = "网站进入【维护】，下拉到【猫超账号】，点击对应账号右侧的铅笔编辑按钮，复制弹窗中的【账号标识】。"
    guide.merge_cells("A3:H3")
    guide.merge_cells("A4:H4")
    guide["A21"] = "2. 获取二级供应商名称"
    guide["A22"] = "网站进入【设置】→【分配管理】，依次选择供应链和组员，在【配置供应商】列表中复制完整供应商名称。"
    guide.merge_cells("A21:H21")
    guide.merge_cells("A22:H22")
    guide["A46"] = "3. 填写货品 ID"
    guide["B46"] = "进入第二个工作表【货品ID配置】，每个货品 ID 填写一行；账号标识和完整供应商名称可以向下重复粘贴。"
    guide["A47"] = "4. 简称"
    guide["B47"] = "简称仅方便运营识别，可以自行填写或留空，不参与系统匹配。"
    guide["A48"] = "5. 上传"
    guide["B48"] = "保存为 .xlsx 后，在网站的货品 ID 配置区域上传。系统按【猫超账号标识 + 二级供应商名称】识别供应商。"
    guide["A50"] = "规则"
    guide["B50"] = "说明"
    rules = [
        ("货品ID", "请按文本格式填写，不要使用科学计数法。"),
        ("分批", "货品 ID 超过 30 个时，系统每 30 个自动分批，最终合并成一个库位明细文件。"),
        ("空配置", "未配置货品 ID 的账号和供应商按原逻辑全量导出。"),
        ("名称要求", "供应商名称必须从设置页面复制完整名称；找不到或同名无法唯一识别时，系统会拒绝导入并提示具体行号。"),
    ]
    for row, values in enumerate(rules, start=51):
        guide.cell(row, 1, values[0])
        guide.cell(row, 2, values[1])
    guide["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    for cell in guide[1]:
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for row in (3, 21, 50):
        for cell in guide[row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
    for column in "ABCDEFGH":
        guide.column_dimensions[column].width = 18
    for row in guide.iter_rows(min_row=1, max_row=54, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in range(5, 21):
        guide.row_dimensions[row].height = 19
    for row in range(23, 46):
        guide.row_dimensions[row].height = 19
    for row in range(46, 55):
        guide.row_dimensions[row].height = 32
    guide.row_dimensions[1].height = 28
    image_dir = WEB_ROOT / "assets" / "item-id-guide"
    account_image_path = image_dir / "account-key.png"
    supplier_image_path = image_dir / "supplier-name.png"
    if account_image_path.is_file():
        image = XLImage(account_image_path)
        image.width = 1000
        image.height = 372
        guide.add_image(image, "A5")
    if supplier_image_path.is_file():
        image = XLImage(supplier_image_path)
        image.width = 1000
        image.height = 549
        guide.add_image(image, "A23")
    guide.freeze_panes = "A3"
    config_sheet = workbook.create_sheet("货品ID配置")
    config_sheet.append(["猫超账号标识", "二级供应商名称", "货品ID", "简称"])
    for supplier in store.list_account_suppliers(include_hidden=False):
        config_sheet.append([supplier["account_key"], supplier["supplier_name"], "", ""])
    for column in ("A", "B", "C", "D"):
        for cell in config_sheet[column]:
            cell.number_format = "@"
    for cell in config_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    config_sheet.freeze_panes = "A2"
    config_sheet.column_dimensions["A"].width = 32
    config_sheet.column_dimensions["B"].width = 60
    config_sheet.column_dimensions["C"].width = 24
    config_sheet.column_dimensions["D"].width = 24
    output = BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=item_id_config_template.xlsx"},
    )


@app.post("/api/item-id-config/upload")
async def upload_item_id_config(request: Request) -> dict[str, Any]:
    _require_admin(request)
    content = await request.body()
    if not content:
        raise HTTPException(status_code=400, detail="请选择配置文件")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="配置文件不能超过 10MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        if "货品ID配置" not in workbook.sheetnames:
            raise ValueError("缺少“货品ID配置”工作表")
        sheet = workbook["货品ID配置"]
        header = [str(cell.value or "").strip() for cell in sheet[1]]
        if len(header) < 3 or header[0] != "猫超账号标识" or header[2] != "货品ID":
            raise ValueError("表头必须为: 猫超账号标识 / 二级供应商名称 / 货品ID / 简称")
        if header[1] not in {"二级供应商名称", "二级供应商ID"}:
            raise ValueError("第二列表头必须为“二级供应商名称”")
        supplier_field = "supplier_name" if header[1] == "二级供应商名称" else "supplier_id"
        rows = []
        for values in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
            rows.append({"account_key": values[0], supplier_field: values[1], "item_id": values[2]})
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取配置文件: {exc}") from exc
    user = getattr(request.state, "user", {}) or {}
    result = store.replace_item_id_config(
        rows,
        original_name=unquote(request.headers.get("x-file-name") or "item_id_config.xlsx"),
        uploaded_by_role=str(user.get("role") or ""),
        uploaded_by_user_id=str(user.get("user_id") or ""),
    )
    if result.get("status") == "rejected":
        raise HTTPException(status_code=400, detail=result)
    return result


@app.post("/api/item-id-config/uploads/{upload_id}/rollback")
def rollback_item_id_config(upload_id: str, request: Request) -> dict[str, Any]:
    _require_admin(request)
    try:
        return store.rollback_item_id_config(upload_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="上传版本不存在") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/schedules")
def schedules(operator_id: str = "") -> list[dict[str, Any]]:
    return store.list_schedules(operator_id)


def _single_schedule_task_key(task_key: str) -> str:
    try:
        task_keys = selected_tasks([task_key])
    except KeyError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if len(task_keys) != 1:
        raise HTTPException(status_code=400, detail="schedule task_key must be a single task")
    return task_keys[0]


@app.post("/api/schedules")
def create_schedule(payload: SchedulePayload, request: Request) -> dict[str, Any]:
    _require_admin(request)
    try:
        return store.create_schedule(
            task_keys=payload.task_keys,
            operator_ids=payload.normalized_operator_ids(),
            all_operators=payload.all_operators,
            enabled=payload.enabled,
            time_of_day=payload.time_of_day,
            weekdays=payload.weekdays,
            headed=payload.headed,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.put("/api/schedules/{task_key}")
def put_legacy_task_schedule(task_key: str, payload: SchedulePayload, request: Request) -> dict[str, Any]:
    _require_admin(request)
    normalized_task_key = _single_schedule_task_key(task_key)
    existing = next(
        (
            schedule
            for schedule in store.list_schedules()
            if schedule.get("task_keys") == [normalized_task_key]
        ),
        None,
    )
    try:
        if existing:
            return store.update_schedule(
                existing["schedule_id"],
                task_keys=[normalized_task_key],
                operator_ids=payload.normalized_operator_ids(),
                all_operators=payload.all_operators,
                enabled=payload.enabled,
                time_of_day=payload.time_of_day,
                weekdays=payload.weekdays,
                headed=payload.headed,
            )
        return store.create_schedule(
            task_keys=[normalized_task_key],
            operator_ids=payload.normalized_operator_ids(),
            all_operators=payload.all_operators,
            enabled=payload.enabled,
            time_of_day=payload.time_of_day,
            weekdays=payload.weekdays,
            headed=payload.headed,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.patch("/api/schedules/{schedule_id}")
def update_schedule(schedule_id: str, payload: SchedulePayload, request: Request) -> dict[str, Any]:
    _require_admin(request)
    try:
        return store.update_schedule(
            schedule_id,
            task_keys=payload.task_keys,
            operator_ids=payload.normalized_operator_ids(),
            all_operators=payload.all_operators,
            enabled=payload.enabled,
            time_of_day=payload.time_of_day,
            weekdays=payload.weekdays,
            headed=payload.headed,
        )
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="schedule not found") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.delete("/api/schedules/{schedule_id}")
def delete_schedule(schedule_id: str, request: Request) -> dict[str, str]:
    _require_admin(request)
    try:
        store.delete_schedule(schedule_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="schedule not found") from exc
    return {"status": "ok", "schedule_id": schedule_id}


@app.post("/api/accounts/{account_key}/suppliers/sync")
def sync_account_suppliers(account_key: str) -> dict[str, Any]:
    existing = [account for account in store.list_accounts() if account["key"] == account_key]
    if not existing:
        raise HTTPException(status_code=404, detail="account not found")
    return store.create_run(
        task_keys=[],
        account_keys=[account_key],
        force_account_tasks=True,
        headed=True,
        run_kind=RUN_KIND_SYNC_SUPPLIERS,
    )


@app.post("/api/runs")
def create_run(payload: RunCreate, request: Request) -> dict[str, Any]:
    run_kind = payload.run_kind or RUN_KIND_TASKS
    operator_id = payload.operator_id
    if _is_supply_chain(request):
        if run_kind != RUN_KIND_TASKS:
            raise HTTPException(status_code=403, detail="供应链账号不能执行系统维护任务")
        _require_owned_operator(operator_id, request)
    if _is_member(request):
        if run_kind != RUN_KIND_TASKS:
            raise HTTPException(status_code=403, detail="组员不能执行维护任务")
        member_operator_id = _member_operator_id(request)
        if payload.operator_id and payload.operator_id != member_operator_id:
            raise HTTPException(status_code=403, detail="只能发起自己的下载任务")
        operator_id = member_operator_id
    try:
        task_keys = [] if run_kind == RUN_KIND_SYNC_SUPPLIERS else selected_tasks(payload.task_keys)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    supplier_payloads = [item.dict() for item in payload.suppliers]
    account_keys = payload.account_keys
    if not account_keys:
        supplier_account_keys = [
            str(item.get("account_key") or "")
            for item in supplier_payloads
            if item.get("account_key")
        ]
        account_keys = list(dict.fromkeys(supplier_account_keys)) or [account["key"] for account in store.list_accounts()]
    try:
        return store.create_run(
            task_keys,
            account_keys,
            payload.force_account_tasks if run_kind == RUN_KIND_SYNC_SUPPLIERS else True,
            payload.headed,
            suppliers=supplier_payloads,
            operator_id=operator_id,
            run_kind=run_kind,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/runs")
def runs(request: Request) -> list[dict[str, Any]]:
    return _member_runs(store.list_runs(), request)


@app.get("/api/runs/{run_id}")
def run(run_id: str, request: Request) -> dict[str, Any]:
    return _get_member_checked_run(run_id, request)


@app.post("/api/runs/{run_id}/cancel")
def cancel_run(run_id: str, request: Request) -> dict[str, Any]:
    try:
        _get_member_checked_run(run_id, request)
        return store.cancel_pending_run(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/runs/{run_id}/pause")
def pause_run(run_id: str, request: Request) -> dict[str, Any]:
    try:
        _get_member_checked_run(run_id, request)
        return store.pause_run(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/runs/{run_id}/resume")
def resume_run(run_id: str, request: Request) -> dict[str, Any]:
    try:
        _get_member_checked_run(run_id, request)
        return store.resume_run(run_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/runs/{run_id}/move-up")
def move_run_up(run_id: str, request: Request) -> dict[str, Any]:
    try:
        _get_member_checked_run(run_id, request)
        return store.move_pending_run(run_id, -1)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/runs/{run_id}/move-down")
def move_run_down(run_id: str, request: Request) -> dict[str, Any]:
    try:
        _get_member_checked_run(run_id, request)
        return store.move_pending_run(run_id, 1)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="run not found") from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.get("/api/runs/{run_id}/logs")
def run_logs(run_id: str, request: Request) -> PlainTextResponse:
    _get_member_checked_run(run_id, request)
    log_path = store.settings.log_dir / f"{run_id}.log"
    if not log_path.exists():
        events = store.list_task_events(run_id)
        return PlainTextResponse("\n".join(json.dumps(event, ensure_ascii=False) for event in events))
    return PlainTextResponse(log_path.read_text(encoding="utf-8", errors="replace"))


@app.get("/api/runs/{run_id}/files/download")
def download_run_files(run_id: str, request: Request) -> FileResponse:
    run_item = _get_member_checked_run(run_id, request)
    files = _run_output_files(run_item)
    if not files:
        raise HTTPException(status_code=404, detail="run has no downloadable files")
    zip_path = Path(tempfile.gettempdir()) / f"maochao_run_{run_id[:8]}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in files:
            try:
                arcname = path.relative_to(store.settings.data_root)
            except ValueError:
                arcname = Path(path.name)
            archive.write(path, arcname=str(arcname))
    return FileResponse(zip_path, filename=f"maochao_run_{run_id[:8]}.zip")


@app.get("/api/runs/{run_id}/errors")
def run_errors(run_id: str, request: Request) -> list[dict[str, Any]]:
    run_item = _get_member_checked_run(run_id, request)
    errors = [item for item in run_item.get("result", []) if item.get("status") != "ok" or item.get("error")]
    if run_item.get("error"):
        errors.append({"run_id": run_id, "error": run_item["error"]})
    return errors


@app.get("/api/errors")
def errors(request: Request) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for run_item in _member_runs(store.list_runs(), request):
        if run_item.get("error"):
            rows.append({"run_id": run_item["run_id"], "error": run_item["error"], "updated_at": run_item["updated_at"]})
        for result in run_item.get("result", []):
            if result.get("status") != "ok" or result.get("error"):
                rows.append({"run_id": run_item["run_id"], **result})
    return rows


@app.get("/api/files")
def files(request: Request) -> list[dict[str, Any]]:
    run_index = _file_run_index()
    rows = store.list_files()
    for item in rows:
        if item.get("run_id"):
            continue
        item["run_id"] = run_index.get(str(Path(item["path"]).resolve()), "")
    return _member_files(rows, request)


@app.get("/api/files/{file_id:path}/download")
def download_file(file_id: str, request: Request) -> FileResponse:
    if _is_member(request) or _is_supply_chain(request):
        allowed_ids = {
            _normalized_file_id(item.get("file_id", ""))
            for item in _member_files(store.list_files(), request)
        }
        if _normalized_file_id(file_id) not in allowed_ids:
            raise HTTPException(status_code=404, detail="file not found")
    path = _resolve_file_id(store.settings.data_root, file_id)
    return FileResponse(path, filename=path.name)


@app.get("/api/screenshots/{screenshot_id:path}")
def screenshot(screenshot_id: str) -> FileResponse:
    path = _resolve_file_id(store.settings.screenshot_dir, screenshot_id)
    return FileResponse(path, filename=path.name)


@app.get("/static/{file_path:path}", include_in_schema=False)
def static_file(file_path: str) -> FileResponse:
    path = _resolve_file_id(WEB_ROOT, file_path)
    return FileResponse(path, headers={"Cache-Control": "no-store"})


def main() -> None:
    import uvicorn

    host = os.environ.get("RPA_API_HOST", "0.0.0.0")
    port = int(os.environ.get("RPA_API_PORT", "8000"))
    uvicorn.run("api_server:app", host=host, port=port, reload=False)


if __name__ == "__main__":
    main()
